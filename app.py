"""
工程造价智能助手 - Streamlit 网页界面（Phase 1: 问答功能）
"""
import re
import streamlit as st
from rag_engine import get_engine
from fee_engine import (
    detect_and_calculate, calc_jianli, calc_sheji, calc_huanping,
    calc_cost_consulting_multi, _COST_CONSULTING_SERVICE_ORDER,
    calc_cost_consulting_multi_hebei, _HEBEI_COST_CONSULTING_SERVICE_ORDER,
    _HEBEI_PROFESSIONAL_COEFFICIENTS,
    resolve_dependent_calc, calc_huanping_multi,
    _build_fee_selection_meta, _TIER_DEPS, _calc_all_fees,
    _FEE_LABELS, _is_hebei_project, is_hebei_region,
    ALL_PROVINCES, DEFAULT_REGION,
    JIANLI_PROFESSIONAL_OPTIONS, JIANLI_COMPLEXITY_OPTIONS,
    JIANLI_ELEVATION_OPTIONS,
    SHEJI_PROFESSIONAL_OPTIONS, SHEJI_COMPLEXITY_OPTIONS,
)

# ===== 政策依据可点击链接 =====

_POLICY_URLS: dict[str, str] = {
    # 河北省
    "冀建市研[2017]2号": "https://www.baidu.com/s?wd=冀建市研[2017]2号+工程造价咨询服务收费管理暂行办法",
    "冀价行费[2018]57号": "https://www.baidu.com/s?wd=冀价行费[2018]57号+施工图审查费",
    "发改价格〔2011〕534号": "https://www.baidu.com/s?wd=发改价格〔2011〕534号+降低部分建设项目收费标准",
    # 天津市
    "津价房地[2008]136号": "https://www.baidu.com/s?wd=津价房地[2008]136号+建设工程造价咨询服务",
    "津价管[2011]46号": "https://www.baidu.com/s?wd=津价管[2011]46号+施工图设计文件审查",
    # 国家层面
    "计价格[2002]10号": "https://www.baidu.com/s?wd=计价格[2002]10号+工程勘察设计收费管理规定",
    "发改价格[2007]670号": "https://www.baidu.com/s?wd=发改价格[2007]670号+建设工程监理收费",
    "计价格[2002]125号": "https://www.baidu.com/s?wd=计价格[2002]125号+环境影响咨询收费",
    "计价格[1999]1283号": "https://www.baidu.com/s?wd=计价格[1999]1283号+建设项目前期工作咨询费",
    "计价格[2002]1980号": "https://www.baidu.com/s?wd=计价格[2002]1980号+招标代理服务收费",
    "建市[2007]86号": "https://www.baidu.com/s?wd=建市[2007]86号+工程设计资质标准",
}


def show_policy_badge(policy_id: str):
    """Display a clickable policy badge in the Streamlit UI.

    Renders a styled box with a link to search for the policy document.
    Falls back to st.info() if the policy ID is unknown.
    """
    url = _POLICY_URLS.get(policy_id)
    if url:
        st.markdown(
            f'<div style="background-color:#d4e6f1;padding:10px 14px;'
            f'border-radius:4px;text-align:center;border:1px solid #aed6f1;">'
            f'<a href="{url}" target="_blank" rel="noopener"'
            f' style="color:#0d47a1;text-decoration:none;font-weight:bold;font-size:14px;">'
            f'📋 {policy_id}</a></div>',
            unsafe_allow_html=True,
        )
    else:
        st.info(policy_id)


def _round2(val: float) -> str:
    """Round to 2 decimal places using round-half-up (matches Excel ROUND).

    Python's built-in round() and format() use banker's rounding (round half
    to even), which can differ from Excel at the exact .005 boundary by 0.01.
    """
    import math
    return f"{math.floor(val * 100 + 0.5) / 100:.2f}"


def _basis_with_links(basis_text: str) -> str:
    """Insert clickable HTML links into a 依据 text for panel captions."""
    result = basis_text
    for pid, url in _POLICY_URLS.items():
        if pid in result:
            result = result.replace(
                pid,
                f'<a href="{url}" target="_blank" rel="noopener"'
                f' style="text-decoration:none;color:#0d47a1;">{pid}</a>',
            )
    return result


def _basis_md_links(basis_text: str) -> str:
    """Insert clickable markdown links into a 依据 text for chat messages."""
    result = basis_text
    for pid, url in _POLICY_URLS.items():
        if pid in result:
            result = result.replace(pid, f"[{pid}]({url})")
    return result


# ===== 页面设置 =====
st.set_page_config(
    page_title="造价智能助手",
    page_icon="🌿",
    layout="wide",
)

# ===== 辅助渲染函数 =====


def _get_fee_numeric(fee_result: dict) -> float | None:
    """从二类费计算结果中提取数值金额（统一为万元）。"""
    val = fee_result.get("结果(万元)")
    if isinstance(val, (int, float)):
        return float(val)
    mid = fee_result.get("结果中值(万元)")
    if mid is not None:
        return float(mid)
    yuan = fee_result.get("结果(元)")
    if yuan is not None:
        return round(float(yuan) / 10000.0, 4)
    return None


def _build_cascade_excel(ctx: dict) -> bytes:
    """根据级联计算结果生成 Excel 文件，返回 bytes 供下载。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
    from openpyxl.utils import get_column_letter
    import datetime

    wb = Workbook()
    ws = wb.active
    ws.title = "费用汇总"

    # ── 样式定义 ──
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))
    header_font = Font(name="微软雅黑", bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_w = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    title_font = Font(name="微软雅黑", bold=True, size=14)
    subtotal_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    normal_font = Font(name="微软雅黑", size=10)
    bold_font = Font(name="微软雅黑", bold=True, size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    money_fmt = '#,##0.00'

    preview = ctx.get("preview", {})
    numerical = preview.get("numerical", {}) if preview else {}
    fee_defs = ctx.get("fee_defs", [])
    selected = ctx.get("selected_fees", set())
    custom_fees = ctx.get("custom_fees", [])
    fee_discounts = ctx.get("fee_discounts", {})

    # ── 列宽预设 ──
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 28

    row = 1
    # ── 标题 ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value="建设项目二类费计算汇总表").font = title_font
    ws.cell(row=row, column=1).alignment = center_align
    row += 1

    # ── 项目基本信息 ──
    info_data = [
        ("建安工程费", f"{ctx.get('jianan', 0):.2f} 万元"),
        ("设备购置费", f"{ctx.get('shebei', 0):.2f} 万元"),
        ("第一部分工程费", f"{ctx.get('total_part1', 0):.2f} 万元"),
        ("项目类型", ctx.get("project_type", "")),
        ("计算日期", datetime.date.today().isoformat()),
    ]
    for label, val in info_data:
        ws.cell(row=row, column=1, value=label).font = bold_font
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.cell(row=row, column=2, value=val).font = normal_font
        row += 1
    row += 1

    # ── 表头 ──
    headers = ["序号", "费用名称", "费用（万元）", "打折后（万元）", "备注", "依据"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font_w
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1

    # ── 按层级输出各费种 ──
    seq = 0
    raw_total = 0.0
    discounted_total = 0.0
    tier_names = {0: "第一部分工程费相关", 1: "勘察设计费相关", 2: "总投资相关"}

    for tier in [0, 1, 2]:
        tier_fees = sorted(
            [fd for fd in fee_defs if fd["tier"] == tier and fd["name"] in selected],
            key=lambda fd: fd["name"])
        if not tier_fees:
            continue
        # 层级小标题
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        tier_cell = ws.cell(row=row, column=1, value=tier_names.get(tier, f"Tier {tier}"))
        tier_cell.font = Font(name="微软雅黑", bold=True, size=10, color="4472C4")
        tier_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        row += 1

        for fd in tier_fees:
            fn = fd["name"]
            val = numerical.get(f"{fn}(万元)")
            if val is None or val <= 0:
                continue
            seq += 1
            disc = fee_discounts.get(fn, 1.0)
            disc_val = round(val * disc, 4)
            raw_total += val
            discounted_total += disc_val

            # 备注
            notes = []
            if fn in ctx.get("coef_overrides", {}):
                for k, v in ctx["coef_overrides"][fn].items():
                    if abs(v - 1.0) > 0.005:
                        notes.append(f"{k}={v}")
            if fn in ctx.get("rate_overrides", {}):
                notes.append(f"费率={ctx['rate_overrides'][fn]}")
            if fn in ctx.get("service_selections", {}):
                svcs = ctx["service_selections"][fn]
                notes.append(f"{'、'.join(svcs)}")
            if abs(disc - 1.0) >= 0.005:
                notes.append(f"打折={disc:.2f}")
            note_str = "；".join(notes) if notes else ""
            display_val = disc_val if abs(disc - 1.0) >= 0.005 else val

            ws.cell(row=row, column=1, value=seq).font = normal_font
            ws.cell(row=row, column=1).alignment = center_align
            ws.cell(row=row, column=2, value=fd["label"]).font = normal_font
            ws.cell(row=row, column=3, value=val).font = normal_font
            ws.cell(row=row, column=3).number_format = money_fmt
            ws.cell(row=row, column=3).alignment = right_align
            ws.cell(row=row, column=4, value=display_val).font = normal_font
            ws.cell(row=row, column=4).number_format = money_fmt
            ws.cell(row=row, column=4).alignment = right_align
            ws.cell(row=row, column=5, value=note_str).font = Font(name="微软雅黑", size=9)
            ws.cell(row=row, column=6, value=fd.get("依据", "")).font = Font(name="微软雅黑", size=9)
            for c in range(1, 7):
                ws.cell(row=row, column=c).border = thin_border
            row += 1

    # ── 自定义费用 ──
    if custom_fees:
        for cf in custom_fees:
            seq += 1
            cf_amount = cf["amount_wan"]
            raw_total += cf_amount
            discounted_total += cf_amount
            ws.cell(row=row, column=1, value=seq).font = normal_font
            ws.cell(row=row, column=1).alignment = center_align
            ws.cell(row=row, column=2, value=f"【自定义】{cf['name']}").font = normal_font
            ws.cell(row=row, column=3, value=cf_amount).font = normal_font
            ws.cell(row=row, column=3).number_format = money_fmt
            ws.cell(row=row, column=3).alignment = right_align
            ws.cell(row=row, column=4, value=cf_amount).font = normal_font
            ws.cell(row=row, column=4).number_format = money_fmt
            ws.cell(row=row, column=4).alignment = right_align
            ws.cell(row=row, column=5, value="自定义费用，不打折").font = Font(name="微软雅黑", size=9)
            for c in range(1, 7):
                ws.cell(row=row, column=c).border = thin_border
            row += 1

    # ── 二类费合计 ──
    for c in range(1, 7):
        ws.cell(row=row, column=c).fill = subtotal_fill
        ws.cell(row=row, column=c).border = thin_border
    ws.cell(row=row, column=2, value="二类费合计").font = bold_font
    ws.cell(row=row, column=3, value=round(raw_total, 4)).font = bold_font
    ws.cell(row=row, column=3).number_format = money_fmt
    ws.cell(row=row, column=3).alignment = right_align
    display_disc_total = round(discounted_total, 4)
    ws.cell(row=row, column=4, value=display_disc_total).font = bold_font
    ws.cell(row=row, column=4).number_format = money_fmt
    ws.cell(row=row, column=4).alignment = right_align
    row += 1

    # ── 预备费 ──
    yb_val = preview.get("yubei_total", 0) if preview else 0
    if yb_val > 0:
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = thin_border
        ws.cell(row=row, column=2, value="预备费（基本预备费）").font = bold_font
        ws.cell(row=row, column=3, value=round(yb_val, 4)).font = bold_font
        ws.cell(row=row, column=3).number_format = money_fmt
        ws.cell(row=row, column=3).alignment = right_align
        ws.cell(row=row, column=4, value=round(yb_val, 4)).font = bold_font
        ws.cell(row=row, column=4).number_format = money_fmt
        ws.cell(row=row, column=4).alignment = right_align
        row += 1

    # ── 项目总投资 ──
    project_total = preview.get("project_total_with_custom", 0) if preview else 0
    for c in range(1, 7):
        ws.cell(row=row, column=c).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        ws.cell(row=row, column=c).border = thin_border
    ws.cell(row=row, column=2, value="项目总投资").font = Font(name="微软雅黑", bold=True, size=11)
    ws.cell(row=row, column=3, value=round(project_total, 4)).font = Font(name="微软雅黑", bold=True, size=11)
    ws.cell(row=row, column=3).number_format = money_fmt
    ws.cell(row=row, column=3).alignment = right_align

    # ── 保存到内存 ──
    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _render_discount_section(fee_value_wan: float, default_discount: float, key_suffix: str) -> tuple[float, float]:
    """
    渲染打折系数区域，返回 (打折系数, 打折后费用_万元)。

    参数：
        fee_value_wan: 打折前费用（万元）
        default_discount: 默认打折系数（1.0 = 不打折）
        key_suffix: session_state key 后缀（用于区分不同费种）
    """
    st.markdown("### 💰 费用打折")
    discount = st.number_input(
        "打折系数（1.0 = 不打折，0.8 = 打八折，0.5 = 打五折）",
        min_value=0.01,
        max_value=2.00,
        value=default_discount,
        step=0.05,
        format="%.2f",
        key=f"discount_{key_suffix}",
        help="输入打折系数：1.0 表示不打折，0.8 表示打八折，1.1 表示上浮 10%。",
    )

    discounted = round(fee_value_wan * discount, 4)

    if abs(discount - 1.0) < 0.005:
        st.info(f"**不打折**，最终费用：**{discounted} 万元**")
    elif discount < 1.0:
        st.warning(
            f"打折系数 **{discount:.2f}** → "
            f"{fee_value_wan:.2f} 万 × {discount:.2f} = **{discounted} 万元**"
            f"（节省 {round(fee_value_wan - discounted, 4)} 万元）"
        )
    else:
        st.warning(
            f"上浮系数 **{discount:.2f}** → "
            f"{fee_value_wan:.2f} 万 × {discount:.2f} = **{discounted} 万元**"
            f"（增加 {round(discounted - fee_value_wan, 4)} 万元）"
        )

    st.markdown("---")
    return discount, discounted


def _render_engine_card(fee_result):
    """渲染单个费种的引擎计算结果卡片"""
    params = fee_result.get("参数", {})
    result_val = fee_result.get("结果(万元)") or fee_result.get("结果(元)")
    unit = "万元" if "结果(万元)" in fee_result else "元"

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric(label="费种", value=fee_result.get("费种", ""))
    with col2:
        st.metric(label=f"金额（{unit}）", value=f"{result_val}")
    with col3:
        basis_text = _basis_with_links(fee_result.get('依据', ''))
        st.markdown(f"<small>依据：{basis_text}</small>", unsafe_allow_html=True)

    with st.expander("查看计算明细", expanded=True):
        if params:
            st.markdown("**输入参数**：")
            for k, v in params.items():
                st.markdown(f"- {k}：**{v}**")
        basic = fee_result.get("基本设计收费(万元)")
        if basic is not None:
            st.markdown(f"**基本设计收费**：**{basic} 万元**")
        other = fee_result.get("其他设计收费明细")
        if other:
            st.markdown("**其他设计收费**：")
            for od in other:
                st.markdown(f"- {od['项目']}：**{od['费用(万元)']} 万元**")
        items = fee_result.get("分项明细")
        if items and len(items) > 1:
            st.markdown("**分项明细**：")
            for item in items:
                st.markdown(
                    f"- {item.get('类别', '')}：基数 {item.get('基数(万元)', '')} 万元 "
                    f"→ **{item.get('费用(元)', '')} 元**"
                )
        steps = fee_result.get("计算步骤")
        if steps:
            # 区分两种步骤格式：粗略估算类（有"步骤"/"公式"/"结果"）vs 分档累进类（有"区间"/"费率"）
            if steps[0].get("步骤"):
                st.markdown("**计算过程**：")
                for i, s in enumerate(steps, 1):
                    formula = s.get("公式", "")
                    result_step = s.get("结果", "")
                    st.markdown(
                        f"**{i}. {s.get('步骤', '')}**：{formula} → **{result_step}**"
                    )
            else:
                st.markdown("**分档计算**：")
                for s in steps:
                    st.markdown(
                        f"- {s.get('区间', '')}：{s.get('金额(万元)', '')}万元 "
                        f"× {s.get('费率(%)', '')}% = **{s.get('费用(万元)', '')}万元**"
                    )
        if "分摊" in fee_result:
            st.caption(fee_result["分摊"])
        adjustment = fee_result.get("计费额调整")
        if adjustment and adjustment.get("触发调整"):
            st.info(adjustment.get("说明", ""))


def _render_sheji_static(fee_result):
    """渲染工程设计费的静态确认框（绕过 LLM）"""
    params = fee_result.get("参数", {})
    prof = params.get("专业调整系数", "")
    complexity = params.get("复杂程度系数", "")
    additional = params.get("附加调整系数", "")
    jifei = params.get("计费额(万元)", "")
    jijia = params.get("收费基价(万元)", "")
    basic = fee_result.get("基本设计收费(万元)")

    result_lines = [
        f"以上为程序依据《工程勘察设计收费管理规定》（计价格[2002]10号）精确计算结果：\n",
        f"- 计费额 **{jifei}** 万元",
        f"- 收费基价（附表一内插）**{jijia}** 万元",
        f"- 专业调整系数（附表二）**{prof}**",
        f"- 复杂程度系数 **{complexity}**",
        f"- 附加调整系数 **{additional}**",
        f"- 基本设计收费 = {jijia} × {prof} × {complexity} × {additional} = **{basic} 万元**",
    ]

    other_items = fee_result.get("其他设计收费明细") or []
    other_total = 0.0
    result_lines.append("")
    result_lines.append("**其他设计收费（可选，按基本设计收费比例计取）**：")
    other_types = [
        ("总体设计费", 0.05),
        ("主体设计协调费", 0.05),
        ("施工图预算编制费", 0.10),
        ("竣工图编制费", 0.08),
    ]
    other_selected = {od.get("项目", ""): od.get("费用(万元)", 0) for od in other_items}
    for name, rate in other_types:
        if name in other_selected:
            fee_val = other_selected[name]
            other_total += fee_val
            result_lines.append(f"  - {name}（{int(rate*100)}%）**{fee_val} 万元** ✓已计算")
        else:
            fee_val = round(basic * rate, 4)
            result_lines.append(f"  - {name}（{int(rate*100)}%）{fee_val} 万元")
    for od in other_items:
        if od.get("项目", "") not in dict(other_types):
            fee_val = od.get("费用(万元)", 0)
            other_total += fee_val
            result_lines.append(f"  - {od.get('项目', '')} **{fee_val} 万元** ✓已计算")

    if other_items:
        benchmark = fee_result.get("结果(万元)", basic + other_total)
        result_lines.append(
            f"- 工程设计收费基准价 = {basic} + {other_total} = **{benchmark} 万元**"
        )
    else:
        result_lines.append(
            "  > 💡 以上为参考值，需在提问时说明（如\"含施工图预算\"）才会计入总计。"
        )

    st.success("\n".join(result_lines))


def _build_sheji_text(fee_result):
    """构建工程设计费的对话历史文本"""
    params = fee_result.get("参数", {})
    prof = params.get("专业调整系数", "")
    complexity = params.get("复杂程度系数", "")
    additional = params.get("附加调整系数", "")
    jifei = params.get("计费额(万元)", "")
    jijia = params.get("收费基价(万元)", "")
    basic = fee_result.get("基本设计收费(万元)")

    text = (
        f"根据计价格[2002]10号，计费额{jifei}万元，"
        f"收费基价{jijia}万元（附表一内插），"
        f"专业调整系数{prof}，"
        f"复杂程度系数{complexity}，"
        f"附加调整系数{additional}，"
        f"基本设计收费 **{basic} 万元**。"
    )
    other_items = fee_result.get("其他设计收费明细") or []
    if other_items:
        other_desc = "；".join(
            f"{od.get('项目', '')}{od.get('费用(万元)', 0)}万元"
            for od in other_items
        )
        benchmark = fee_result.get("结果(万元)")
        text += f"其他设计收费：{other_desc}。工程设计收费基准价 **{benchmark} 万元**。"
    else:
        other_types = [
            ("总体设计费", 0.05), ("主体设计协调费", 0.05),
            ("施工图预算编制费", 0.10), ("竣工图编制费", 0.08),
        ]
        other_lines = "，".join(
            f"{n}({int(r*100)}%)={round(basic*r,4)}万元"
            for n, r in other_types
        )
        text += (
            f"其他设计收费（参考）：{other_lines}。"
            f"请在提问时说明需要哪些（如\"含施工图预算\"），会计入基准价。"
        )
    return text


# ===== 多费种迭代计算渲染函数 =====

def _render_cascade_result(result):
    """渲染模式1：多费种联算结果。"""
    import pandas as pd

    st.markdown("## 多费种联算结果（程序精确计算）")

    params = result["输入参数"]
    col1, col2, col3 = st.columns(3)
    col1.metric("建安工程费", f"{params['建安工程费(万元)']} 万元")
    col2.metric("设备购置费", f"{params['设备购置费(万元)']} 万元")
    col3.metric("项目类型", params["项目类型"])

    st.markdown(
        "**依赖层级说明**：\n"
        "- **T0**（仅依赖建安+设备）：监理费、设计费、勘察费、劳安评审费、场地准备费、工程保险费\n"
        "- **T1**（依赖 T0 结果）：施工图审查费、交易服务费\n"
        "- **T2**（依赖总投资）：建设管理费、建设项目前期工作咨询费、环境影响咨询费\n"
        "- **预备费**：（第一部分工程费+二类费）× 5%"
    )

    rows = result["费种合计"]
    tier_colors = {0: "#e8f5e9", 1: "#fff3e0", 2: "#e3f2fd", 3: "#fce4ec"}
    tier_labels = {0: "第一部分工程费相关", 1: "勘察设计费相关",
                   2: "总投资相关", 3: "预备费"}

    for tier in [0, 1, 2, 3]:
        tier_rows = [r for r in rows if r["层级"] == tier]
        if tier_rows:
            st.caption(tier_labels[tier])
            df = pd.DataFrame(tier_rows)
            df = df[["费种", "金额(万元)"]]
            st.dataframe(df, use_container_width=True, hide_index=True)

    # 汇总
    summary = result["结果汇总"]
    st.markdown("---")
    # 额外费用（用户指定）
    extra_fees = result.get("额外费用", [])
    if extra_fees:
        st.markdown("#### 额外费用（用户指定）")
        for e in extra_fees:
            st.markdown(f"- **{e['名称']}**：{e['金额(万元)']} 万元")

    st.markdown("### 费用汇总")
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("第一部分工程费", f"{summary['第一部分工程费(万元)']:.2f} 万元")
    col2.metric("二类费合计", f"{summary['二类费合计(万元)']:.2f} 万元")
    yubei_val = summary.get("预备费(万元)", 0)
    col3.metric("预备费", f"{yubei_val:.2f} 万元")
    col4.metric("项目总投资", f"{summary['项目总投资(万元)']:.2f} 万元")

    # 层级小计
    extra_caption = ""
    if extra_fees:
        extra_total = summary.get("额外费用小计(万元)", 0)
        extra_caption = f" ｜ 额外费用：{extra_total:.2f} 万元"
    st.caption(
        f"T0 小计：{summary['T0小计(万元)']:.2f} 万元 ｜ "
        f"T1 小计：{summary['T1小计(万元)']:.2f} 万元 ｜ "
        f"T2 小计：{summary['T2小计(万元)']:.2f} 万元"
        f"{extra_caption}"
    )

    skipped = result.get("跳过的费种", {})
    if skipped:
        with st.expander("⚠️ 无法自动计算的费种"):
            for name, reason in skipped.items():
                st.warning(f"**{name}**：{reason}")

    with st.expander("查看各费种详细计算步骤"):
        for fee_name, detail in result["明细"].items():
            if not fee_name.startswith("_"):
                _render_engine_card(detail)

    # 构建响应文本
    lines = []
    for r in rows:
        lines.append(f"- **{r['费种']}**：{r['金额(万元)']:.2f} 万元")
    if extra_fees:
        for e in extra_fees:
            lines.append(f"- **{e['名称']}**（用户指定）：{e['金额(万元)']} 万元")
    yubei_text = ""
    yb_val = summary.get("预备费(万元)", 0)
    if yb_val > 0:
        yubei_text = f"\n**预备费（基本预备费）：{yb_val:.2f} 万元**（(一类费+二类费)×5%）"
    return (
        f"## 多费种联算结果\n\n"
        f"计费基数：建安费 {params['建安工程费(万元)']} 万 + 设备费 {params['设备购置费(万元)']} 万 "
        f"= **{params['第一部分工程费(万元)']} 万元**\n\n"
        f"### 各项费用\n\n" + "\n".join(lines) + "\n\n"
        f"**二类费合计：{summary['二类费合计(万元)']:.2f} 万元**"
        f"{yubei_text}\n\n"
        f"**项目总投资：{summary['项目总投资(万元)']:.2f} 万元**"
    )


def _render_iteration_result(result):
    """渲染模式2：迭代收敛结果。"""
    import pandas as pd

    st.markdown("## 迭代计算（总投资收敛）")

    params = result["输入参数"]
    col1, col2 = st.columns(2)
    col1.metric("建安工程费", f"{params['建安工程费(万元)']} 万元")
    col2.metric("设备购置费", f"{params['设备购置费(万元)']} 万元")

    st.info(
        "**迭代原理**：建设管理费、建设项目前期工作咨询费、环境影响咨询费依赖总投资；"
        "总投资又包含这些二类费本身。通过反复迭代使总投资收敛到稳定值。"
    )

    # 收敛过程表
    history = result["迭代过程"]
    steps_data = []
    for h in history:
        steps_data.append({
            "迭代": h["迭代次数"],
            "总投资(万元)": round(h["总投资(万元)"], 2),
            "二类费合计(万元)": round(h["二类费合计(万元)"], 2),
            "变化(万元)": round(h.get("变化(万元)", 0), 4),
        })

    st.markdown("### 收敛过程")
    st.dataframe(pd.DataFrame(steps_data), use_container_width=True, hide_index=True)

    final = result["收敛结果"]
    converged = result["已收敛"]

    # 额外费用提示
    extra_fees = result.get("额外费用", [])
    extra_note = ""
    if extra_fees:
        extra_total = sum(e["金额(万元)"] for e in extra_fees)
        extra_names = "、".join(f"{e['名称']} {e['金额(万元)']}万" for e in extra_fees)
        extra_note = f"\n\n含用户指定额外费用：{extra_names}（已计入合计）"

    if converged:
        yubei_val = final.get("预备费(万元)", 0)
        proj_total = final.get("项目总投资(万元)", final["总投资(万元)"])
        st.success(
            f"✅ 经过 **{result['迭代次数']}** 次迭代已收敛 "
            f"（阈值 {result['收敛阈值(万元)']} 万元）。\n\n"
            f"静态总投资：**{final['总投资(万元)']:.2f} 万元**，"
            f"二类费合计：**{final['二类费合计(万元)']:.2f} 万元**\n\n"
            f"预备费：**{yubei_val:.2f} 万元**（(一类费+二类费)×5%），"
            f"项目总投资：**{proj_total:.2f} 万元**"
            f"{extra_note}"
        )
    else:
        st.warning(
            f"⚠️ 经过 {result['迭代次数']} 次迭代未完全收敛"
        )

    # 最终明细
    st.markdown("### 收敛后各项费用")
    fees = final["各项费用"]
    for fee_key, val in sorted(fees.items()):
        st.markdown(f"- **{fee_key}**：{val:.2f} 万元")
    # 预备费单独显示
    yubei_final_val = final.get("预备费(万元)")
    if yubei_final_val is not None and yubei_final_val > 0:
        st.markdown(f"- **预备费**：{yubei_final_val:.2f} 万元")
    proj_final = final.get("项目总投资(万元)")
    if proj_final is not None:
        st.markdown(f"\n**项目总投资（含预备费）：{proj_final:.2f} 万元**")

    with st.expander("查看每轮迭代详细数据"):
        for h in history:
            st.markdown(f"#### 第 {h['迭代次数']} 轮")
            fees = h["各项费用"]
            for fee_key, val in sorted(fees.items()):
                st.markdown(f"- {fee_key}：{val:.2f} 万元")
            st.caption(f"总投资：{h['总投资(万元)']:.2f} 万元 ｜ 变化：{h['变化(万元)']:.2f} 万元")

    # 响应文本
    yb_val = final.get("预备费(万元)", 0)
    proj_total = final.get("项目总投资(万元)", final["总投资(万元)"])
    yb_text = f"\n预备费：**{yb_val:.2f} 万元**（(一类费+二类费)×5%）" if yb_val > 0 else ""
    return (
        f"## 迭代计算结果\n\n"
        f"经过 **{result['迭代次数']}** 次迭代，静态总投资收敛至 "
        f"**{final['总投资(万元)']:.2f} 万元**，"
        f"二类费合计 **{final['二类费合计(万元)']:.2f} 万元**。"
        f"{yb_text}\n"
        f"项目总投资（含预备费）：**{proj_total:.2f} 万元**。"
    )


def _render_comparison_result(result):
    """渲染模式3：多方案比选结果。"""
    import pandas as pd

    st.markdown("## 多方案比选 / 敏感性分析")

    sweep = result["扫描参数"]
    st.info(
        f"**扫描参数**：{sweep['参数描述']}，"
        f"共 {len(sweep['值列表'])} 个方案："
        f"{', '.join(str(v) + sweep.get('单位', '') for v in sweep['值列表'])}"
    )

    # 对比表
    st.markdown("### 费用对比表（单位：万元）")
    comparison_rows = result["对比表"]
    df = pd.DataFrame(comparison_rows)
    st.dataframe(df, use_container_width=True, hide_index=True)

    # 柱状图
    st.markdown("### 二类费合计对比")
    chart_data = {}
    for s in result["方案列表"]:
        chart_data[s["方案名称"]] = s["二类费合计(万元)"]
    st.bar_chart(chart_data)

    # 方案明细
    with st.expander("查看每个方案的各项费用明细"):
        for s in result["方案列表"]:
            st.markdown(f"#### {s['方案名称']}")
            fees = s["各项费用"]
            for fee_key, val in sorted(fees.items()):
                st.markdown(f"- **{fee_key}**：{val:.2f} 万元")
            st.metric("二类费合计", f"{s['二类费合计(万元)']:.2f} 万元")
            st.metric("总投资", f"{s['总投资(万元)']:.2f} 万元")

    # 响应文本
    return (
        f"## 多方案比选结果\n\n"
        f"扫描参数：{sweep['参数描述']}，共 {len(sweep['值列表'])} 个方案。"
    )


# ===== 侧边栏 =====
with st.sidebar:
    st.title("🏗️ 造价智能助手")
    st.divider()

    # ── 省份选择 ──
    if "selected_region" not in st.session_state:
        st.session_state.selected_region = DEFAULT_REGION

    selected_region = st.selectbox(
        "📍 所在省份",
        ALL_PROVINCES,
        index=ALL_PROVINCES.index(st.session_state.selected_region)
              if st.session_state.selected_region in ALL_PROVINCES else 1,
        key="region_selector",
    )
    st.session_state.selected_region = selected_region
    st.caption("有特殊政策的省份将自动应用对应费率，其余暂按默认（天津）计算。")

    st.divider()

    st.markdown("### 功能导航")
    st.markdown("- 智能问答（已上线）")
    st.markdown("- 二类费计算（已上线）")
    st.markdown("- 指标对比分析（开发中）")
    st.markdown("- 材料价格趋势（开发中）")

    st.divider()

    st.markdown("### 数据状态")
    try:
        engine = get_engine()
        total = len(engine.chunks)
        cats = len(engine.data)
        kb_count = len(engine.knowledge_chunks)
        st.success(f"已加载 {cats} 个类别，共 {total} 条记录")
        st.caption(f"知识库：{kb_count} 个政策文件片段")
    except Exception as e:
        st.error(f"数据加载失败：{e}")

    st.divider()

    st.markdown("### 绿化指标查询")
    examples_green = [
        "白皮松高度3.5米的综合指标是多少？",
        "落叶乔木胸径14cm的有哪些品种？",
        "常绿乔木和落叶乔木的综合指标对比",
        "灌木球类中综合指标最低的是哪个？",
        "银杏的综合指标是多少？",
    ]
    for i, ex in enumerate(examples_green):
        if st.button(ex, use_container_width=True, key=f"green_{i}"):
            st.session_state.current_query = ex

    st.divider()

    st.markdown("### 二类费计算")
    examples_fee = [
        "建安工程费131万，设备费160万，桥梁工程，勘察费多少？",
        "工程总概算8000万，建设管理费多少？",
        "中标金额6000万工程招标代理费",
        "交易服务费 中标额2000万",
        "监理费 计费额8000万",
        "总投资1.2亿建设管理费和监理费",
    ]
    for i, ex in enumerate(examples_fee):
        if st.button(ex, use_container_width=True, key=f"fee_{i}"):
            st.session_state.current_query = ex

    st.divider()

    st.markdown("### 多费种联算 / 迭代 / 比选")
    examples_multi = [
        "建安费131万，设备费160万，桥梁工程，帮我算全部费用",
        "建安费8000万，工程总概算迭代计算",
        "建安费5000万，设备费3000万，方案比选",
    ]
    for i, ex in enumerate(examples_multi):
        if st.button(ex, use_container_width=True, key=f"multi_{i}"):
            st.session_state.current_query = ex

    st.divider()

    # 持久化调试面板
    if "debug_info" in st.session_state:
        st.markdown("### 🔍 引擎调试")
        d = st.session_state.debug_info
        if "error" in d:
            st.error(f"异常: {d['error']}")
        else:
            st.write(f"prompt: `{d.get('prompt','')}`")
            st.write(f"fee_type: `{d['fee_type']}`")
            st.write(f"has_amount: `{d['has_amount']}`")
            st.write(f"费种: {d.get('费种','')}")
            st.write(f"结果: {d.get('结果','')}")
        if st.button("清除调试", key="clear_debug"):
            del st.session_state.debug_info
            st.rerun()

    st.divider()
    st.caption("Powered by DeepSeek v4")

# ===== 主界面 =====
st.title("🏗️ 工程造价智能问答")
st.caption("基于工程造价指标数据库，提供专业造价问答服务 | v2026-07-03")

# 全局调试：检查 current_query 状态
if "current_query" in st.session_state:
    st.warning(f"**:bug: current_query 已设置:** `{st.session_state.current_query[:80]}`")

# 初始化引擎
with st.spinner("正在加载数据库和 AI 模型..."):
    engine = get_engine()

# 初始化聊天历史
if "messages" not in st.session_state:
    st.session_state.messages = [
        {
            "role": "assistant",
            "content": (
                "你好！我是造价智能助手。\n\n"
                "我可以回答以下问题：\n"
                "- 🌱 查询具体苗木品种的综合指标、栽植费用\n"
                "- 📊 对比不同规格、不同品种的造价差异\n"
                "- 🏗️ 计算工程建设二类费（建设管理费、设计费、监理费等）\n"
                "- 📋 查询二类费政策文件和费率表\n\n"
                "请在下方输入你的问题。"
            ),
        }
    ]

# 显示聊天历史
for msg in st.session_state.messages:
    with st.chat_message(msg["role"]):
        st.markdown(msg["content"])

# ===== 交互式费率选择（持久化在聊天区外） =====
if "pending_rate_select" in st.session_state:
    ctx = st.session_state.pending_rate_select
    fee_result = ctx["fee_result"]
    fee_name = fee_result.get("费种", "")
    ft = fee_result.get("fee_type", "")
    basis = fee_result.get("依据", "")
    desc = fee_result.get("说明", "")
    detail = fee_result.get("费率明细", [])
    steps = fee_result.get("计算步骤", [])
    params = fee_result.get("参数", {})

    rate_opts = [d['费率'] for d in detail]
    fee_map = {d['费率']: d['费用(万元)'] for d in detail}
    mid_idx = len(detail) // 2

    st.divider()

    # ── 卡片容器 ──
    with st.container(border=True):
        # 标题行
        col_title, col_badge = st.columns([3, 1])
        with col_title:
            st.markdown(f"## 🎯 {fee_name}")
        with col_badge:
            st.info(f"共 {len(detail)} 档费率")

        st.markdown(f"<small>📜 **依据**：{_basis_with_links(basis)}</small>", unsafe_allow_html=True)

        # 计算过程折叠
        if steps:
            with st.expander("📐 计算过程", expanded=False):
                for i, s in enumerate(steps, 1):
                    st.markdown(
                        f"**{i}. {s.get('步骤', '')}**  \n"
                        f"> {s.get('公式', '')}  \n"
                        f"> → **{s.get('结果', '')}**"
                    )

        st.markdown("---")
        st.markdown("### 📊 选择适用费率")

        # 费率卡片式选择（radio + 可视化费用卡片）
        selected_rate = st.radio(
            "费率（间隔 0.1%）",
            rate_opts,
            index=mid_idx,
            horizontal=True,
            key=f"persist_rate_{ft}",
            label_visibility="collapsed",
        )

        # 选中的费率高亮卡片
        selected_fee = fee_map[selected_rate]
        st.markdown(
            f"""<div style="
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                border-radius: 12px;
                padding: 20px 28px;
                margin: 12px 0;
                color: white;
            ">
                <div style="font-size: 0.85rem; opacity: 0.85; margin-bottom: 4px;">✅ 当前选择</div>
                <div style="display: flex; align-items: baseline; gap: 16px;">
                    <span style="font-size: 1.1rem;">费率</span>
                    <span style="font-size: 2.0rem; font-weight: 700;">{selected_rate}</span>
                    <span style="font-size: 1.1rem; opacity: 0.7;">→ 费用</span>
                    <span style="font-size: 2.0rem; font-weight: 700;">{selected_fee} 万</span>
                </div>
            </div>""",
            unsafe_allow_html=True,
        )

        st.markdown("---")

        # ── 打折系数 ──
        default_discount = fee_result.get("_discount_coef", 1.0)
        discount_coef, discounted_fee = _render_discount_section(selected_fee, default_discount, ft)

        # 操作按钮行
        col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 2])
        with col_btn1:
            if st.button("✅ 确认选择", type="primary", use_container_width=True, key=f"confirm_rate_{ft}"):
                discount_text = ""
                if abs(discount_coef - 1.0) >= 0.005:
                    discount_text = (
                        f"\n\n**打折系数**：{discount_coef:.2f}\n\n"
                        f"**打折后费用**：{discounted_fee} 万元"
                        f"（{selected_fee} 万 × {discount_coef:.2f}）"
                    )
                response = (
                    f"## {fee_name}\n\n"
                    f"**依据**：{_basis_md_links(basis)}\n\n"
                    f"**选定费率**：{selected_rate}\n\n"
                    f"**费用**：{selected_fee} 万元"
                    f"{discount_text}\n\n"
                    f"{desc}"
                )
                # 清理当前费种的打折系数 session state
                st.session_state.pop(f"discount_{ft}", None)
                st.session_state.messages.append({"role": "assistant", "content": response})
                del st.session_state.pending_rate_select
                st.rerun()
        with col_btn2:
            if st.button("🗑 取消", use_container_width=True, key=f"cancel_rate_{ft}"):
                st.session_state.pop(f"discount_{ft}", None)
                del st.session_state.pending_rate_select
                st.rerun()
        with col_btn3:
            pass  # 占位

        # 底部说明
        with st.expander("ℹ️ 费率说明"):
            st.info(desc)

# ===== 交互式系数选择（持久化，在聊天区外渲染） =====
if "pending_coef_select" in st.session_state:
    ctx = st.session_state.pending_coef_select
    meta = ctx["coef_metadata"]
    fee_result = ctx["fee_result"]
    coefs = meta.get("coefs", [])
    base_params = meta.get("base_params", {})
    calc_func = meta.get("calc_func", "")
    fee_name = fee_result.get("费种", "")
    ft = fee_result.get("fee_type", "")
    basis = fee_result.get("依据", "")

    st.divider()

    with st.container(border=True):
        # 标题行
        col_title, col_badge = st.columns([3, 1])
        with col_title:
            st.markdown(f"## 🎛️ {fee_name} — 系数调整")
        with col_badge:
            st.info(f"{len(coefs)} 个系数")

        st.markdown(f"<small>📜 **依据**：{_basis_with_links(basis)}</small>", unsafe_allow_html=True)

        # ── 各系数下拉选择器 ──
        selected_coefs: dict = {}
        for i, coef_def in enumerate(coefs):
            key = coef_def["key"]
            param_name = coef_def["param_name"]
            current_val = float(coef_def["current"])
            current_label = coef_def.get("current_label", str(current_val))
            options: list = coef_def.get("options", [])
            desc = coef_def.get("description", "")

            st.markdown(f"### {key}")
            st.caption(f"{desc}")

            # 构建选项标签（含值）
            option_labels = [f"{label}（{val}）" for label, val in options]
            option_values = [val for _, val in options]

            # 找到当前值对应的索引
            try:
                current_idx = option_values.index(current_val)
            except ValueError:
                current_idx = len(option_values)  # 指向"自定义"

            # 添加"自定义"选项
            option_labels.append("✏️ 自定义…")
            option_values.append(-1.0)  # -1 = 自定义标记

            selected_idx = st.selectbox(
                f"选择{key}",
                range(len(option_labels)),
                index=min(current_idx, len(option_labels) - 1),
                format_func=lambda idx, labels=option_labels: labels[idx],
                key=f"coef_sel_{ft}_{param_name}",
                label_visibility="collapsed",
            )

            chosen_val = option_values[selected_idx]

            if chosen_val == -1.0:
                # 自定义：显示数字输入
                custom_val = st.number_input(
                    f"自定义{key}的值",
                    min_value=0.10,
                    max_value=5.00,
                    value=current_val if current_val > 0.1 else 1.0,
                    step=0.05,
                    format="%.2f",
                    key=f"coef_cust_{ft}_{param_name}",
                )
                selected_coefs[param_name] = float(custom_val)
                selected_coefs[f"{param_name}_label"] = f"自定义（{custom_val:.2f}）"
            else:
                selected_coefs[param_name] = float(chosen_val)
                # 提取纯标签（去掉末尾的系数值括号）
                raw_label = option_labels[selected_idx]
                if "（" in raw_label:
                    raw_label = raw_label.rsplit("（", 1)[0]
                selected_coefs[f"{param_name}_label"] = raw_label

        st.markdown("---")

        # ── 设计费附加项（仅当 calc_func == "calc_sheji" 时显示）──
        sheji_addon_sgys = False
        sheji_addon_jgt = False
        if calc_func == "calc_sheji":
            st.markdown("### 📐 其他设计收费（可选）")
            st.caption("依据计价格[2002]10号 1.0.16条，以下两项可按需勾选：")
            col_sgys, col_jgt = st.columns(2)
            with col_sgys:
                sheji_addon_sgys = st.checkbox(
                    "施工图预算编制费",
                    value=False,
                    key="sheji_sgys",
                    help="按基本设计收费的 10% 收取（1.0.16条）",
                )
                if sheji_addon_sgys:
                    st.caption("→ 基本设计收费 × 10%")
            with col_jgt:
                sheji_addon_jgt = st.checkbox(
                    "竣工图编制费",
                    value=False,
                    key="sheji_jgt",
                    help="按基本设计收费的 8% 收取（1.0.16条）",
                )
                if sheji_addon_jgt:
                    st.caption("→ 基本设计收费 × 8%")
            st.markdown("---")

        # ── 根据所选系数实时重算 ──
        recalc_fee = None
        recalc_desc = ""
        recalc_error = ""
        try:
            if calc_func == "calc_jianli":
                prof = selected_coefs.get("professional_coef", 1.0)
                comp = selected_coefs.get("complexity_coef", 1.0)
                elev = selected_coefs.get("elevation_coef", 1.0)
                jianan = base_params.get("jianan")
                shebei = base_params.get("shebei")
                amount_wan = base_params.get("amount_wan")

                if jianan is not None and shebei is not None:
                    recalc = calc_jianli(
                        jianan=jianan, shebei=shebei,
                        professional_coef=prof, complexity_coef=comp,
                        elevation_coef=elev,
                    )
                elif amount_wan is not None:
                    recalc = calc_jianli(
                        amount_wan=amount_wan,
                        professional_coef=prof, complexity_coef=comp,
                        elevation_coef=elev,
                    )
                else:
                    recalc = None
                    recalc_error = "无法确定计费额"

                recalc_fee = recalc["结果(万元)"] if recalc else None
                recalc_desc = recalc.get("说明", "") if recalc else ""

            elif calc_func == "calc_sheji":
                prof = selected_coefs.get("professional_coef", 1.0)
                comp = selected_coefs.get("complexity_coef", 1.0)
                addi = selected_coefs.get("additional_coef", 1.0)
                amount_wan = base_params.get("amount_wan")

                if amount_wan is not None:
                    addi_list = [addi] if abs(addi - 1.0) > 0.005 else None
                    recalc = calc_sheji(amount_wan, prof, comp, additional_coefs=addi_list,
                                        shigongtu_yusuan=sheji_addon_sgys,
                                        jungongtu=sheji_addon_jgt)
                else:
                    recalc = None
                    recalc_error = "无法确定计费额"

                recalc_fee = recalc["结果(万元)"] if recalc else None
                recalc_desc = recalc.get("说明", "") if recalc else ""

            elif calc_func == "calc_huanping":
                ind = selected_coefs.get("industry_coef", 1.0)
                sens = selected_coefs.get("sensitivity_coef", 1.0)
                amount_wan = base_params.get("amount_wan", 0)
                svc = base_params.get("service_type", "编制报告书")
                ind_name = selected_coefs.get("industry_coef_label", "市政（默认）")

                recalc = calc_huanping(
                    amount_wan, svc,
                    industry_coef=ind, industry_name=ind_name,
                    sensitivity_coef=sens,
                )
                recalc_fee = recalc.get("结果中值(万元)")
                recalc_desc = recalc.get("说明", "")
            else:
                recalc_error = f"未知计算类型：{calc_func}"
        except Exception as e:
            recalc_error = f"计算出错：{e}"
            import traceback
            recalc_error += f"\n```\n{traceback.format_exc()}\n```"

        # ── 结果卡片 ──
        if recalc_fee is not None:
            st.markdown(
                f"""<div style="
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    border-radius: 12px;
                    padding: 20px 28px;
                    margin: 12px 0;
                    color: white;
                ">
                    <div style="font-size: 0.85rem; opacity: 0.85; margin-bottom: 4px;">✅ 调整后费用</div>
                    <div style="display: flex; align-items: baseline; gap: 16px;">
                        <span style="font-size: 2.0rem; font-weight: 700;">{recalc_fee} 万元</span>
                    </div>
                </div>""",
                unsafe_allow_html=True,
            )
            if recalc_desc:
                with st.expander("📐 查看计算过程", expanded=False):
                    st.markdown(recalc_desc)
        elif recalc_error:
            st.error(recalc_error)

        # ── 打折系数（仅当有有效计算结果时显示）──
        discount_coef = 1.0
        discounted_fee = None
        if recalc_fee is not None:
            default_discount = fee_result.get("_discount_coef", 1.0)
            st.markdown("---")
            discount_coef, discounted_fee = _render_discount_section(recalc_fee, default_discount, ft)
        else:
            st.markdown("---")

        # ── 操作按钮 ──
        col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 2])
        with col_btn1:
            if st.button("✅ 确认选择", type="primary", use_container_width=True, key=f"confirm_coef_{ft}"):
                if recalc_fee is None:
                    st.warning("请先选择有效的系数值")
                    st.stop()
                coef_summary = "、".join(
                    f"{cd['key']}={selected_coefs.get(cd['param_name'], cd['current'])}"
                    for cd in coefs
                )
                discount_text = ""
                final_fee = recalc_fee
                if abs(discount_coef - 1.0) >= 0.005 and discounted_fee is not None:
                    discount_text = (
                        f"\n\n**打折系数**：{discount_coef:.2f}\n\n"
                        f"**打折后费用**：{discounted_fee} 万元"
                        f"（{recalc_fee} 万 × {discount_coef:.2f}）"
                    )
                    final_fee = discounted_fee
                response = (
                    f"## {fee_name}\n\n"
                    f"**依据**：{_basis_md_links(basis)}\n\n"
                    f"**调整后系数**：{coef_summary}\n\n"
                    f"**费用**：{recalc_fee} 万元"
                    f"{discount_text}\n\n"
                    f"---\n{recalc_desc}"
                )
                st.session_state.pop(f"discount_{ft}", None)
                st.session_state.messages.append({"role": "assistant", "content": response})
                del st.session_state.pending_coef_select
                st.rerun()
        with col_btn2:
            if st.button("🗑 取消", use_container_width=True, key=f"cancel_coef_{ft}"):
                st.session_state.pop(f"discount_{ft}", None)
                del st.session_state.pending_coef_select
                st.rerun()

# ===== 简单费种打折调整（持久化在聊天区外） =====
if "pending_simple_fee" in st.session_state:
    ctx = st.session_state.pending_simple_fee
    fee_result = ctx["fee_result"]
    fee_name = fee_result.get("费种", "")
    ft = fee_result.get("fee_type", "")
    basis = fee_result.get("依据", "")
    desc = fee_result.get("说明", "")
    default_discount = ctx.get("default_discount", 1.0)
    base_fee_wan = _get_fee_numeric(fee_result)
    result_val = fee_result.get("结果(万元)") or fee_result.get("结果(元)")
    unit = "万元" if "结果(万元)" in fee_result else "元"

    st.divider()

    with st.container(border=True):
        col_title, col_badge = st.columns([3, 1])
        with col_title:
            st.markdown(f"## 💰 {fee_name} — 费用打折")
        with col_badge:
            st.info(f"基准：{base_fee_wan} 万元")

        st.markdown(f"<small>📜 **依据**：{_basis_with_links(basis)}</small>", unsafe_allow_html=True)

        # ── 计算步骤 ──
        steps = fee_result.get("计算步骤", [])
        if steps:
            if steps[0].get("步骤"):
                st.markdown("### 计算过程")
                for i, s in enumerate(steps, 1):
                    step_name = s.get("步骤", "")
                    formula = s.get("公式", "")
                    result_step = s.get("结果", "")
                    st.markdown(f"**{i}. {step_name}**：{formula} → **{result_step}**")
            else:
                st.markdown("### 分档计算")
                table_rows = []
                for s in steps:
                    qujian = s.get("区间", "")
                    amt = s.get("金额(万元)", "")
                    rate = s.get("费率(‰)", s.get("费率(%)", s.get("费率", "")))
                    rate_unit = "‰" if "费率(‰)" in s else ("%" if "费率(%)" in s else "")
                    fee_val = s.get("费用(万元)", "")
                    table_rows.append(f"| {qujian} | {amt} | {rate}{rate_unit} | **{fee_val}** |")
                st.markdown(
                    "| 区间（万元） | 金额（万元） | 费率 | 费用（万元） |\n"
                    "|:--|:--|:--|:--|\n" + "\n".join(table_rows)
                )

        st.markdown("---")

        # ── 打折系数 ──
        if base_fee_wan is not None:
            discount_coef = st.number_input(
                "打折系数（1.0 = 不打折，0.8 = 打八折）",
                min_value=0.01, max_value=2.00,
                value=default_discount, step=0.05,
                format="%.2f",
                key=f"simple_discount_{ft}",
                help="输入打折系数调整最终费用。",
            )
            discounted_fee_wan = round(base_fee_wan * discount_coef, 4)

            discount_text = ""
            if abs(discount_coef - 1.0) < 0.005:
                st.info(f"**不打折**，最终费用：**{discounted_fee_wan} 万元**")
            elif discount_coef < 1.0:
                st.warning(
                    f"打折系数 **{discount_coef:.2f}** → "
                    f"{base_fee_wan:.2f} 万 × {discount_coef:.2f} = "
                    f"**{discounted_fee_wan} 万元**"
                    f"（节省 {round(base_fee_wan - discounted_fee_wan, 4)} 万元）"
                )
                discount_text = (
                    f"\n\n**打折系数**：{discount_coef:.2f}\n\n"
                    f"**打折后费用**：{discounted_fee_wan} 万元"
                    f"（{base_fee_wan:.2f} 万 × {discount_coef:.2f}）"
                )
            else:
                st.warning(
                    f"上浮系数 **{discount_coef:.2f}** → "
                    f"{base_fee_wan:.2f} 万 × {discount_coef:.2f} = "
                    f"**{discounted_fee_wan} 万元**"
                    f"（增加 {round(discounted_fee_wan - base_fee_wan, 4)} 万元）"
                )
                discount_text = (
                    f"\n\n**上浮系数**：{discount_coef:.2f}\n\n"
                    f"**上浮后费用**：{discounted_fee_wan} 万元"
                    f"（{base_fee_wan:.2f} 万 × {discount_coef:.2f}）"
                )
        else:
            discount_coef = 1.0
            discounted_fee_wan = base_fee_wan
            discount_text = ""
            st.warning("无法获取费用数值")

        # ── 构建响应文本（用于确认按钮）──
        steps_md = ""
        if steps:
            if steps[0].get("步骤"):
                steps_md = "### 计算过程\n\n"
                for i, s in enumerate(steps, 1):
                    step_name = s.get("步骤", "")
                    formula = s.get("公式", "")
                    result_step = s.get("结果", "")
                    steps_md += f"**{i}. {step_name}**：{formula} → **{result_step}**\n\n"
            else:
                steps_md = "### 分档计算\n\n"
                steps_md += "| 区间（万元） | 金额（万元） | 费率 | 费用（万元） |\n"
                steps_md += "|:--|:--|:--|:--|\n"
                for s in steps:
                    qujian = s.get("区间", "")
                    amt = s.get("金额(万元)", "")
                    rate = s.get("费率(‰)", s.get("费率(%)", s.get("费率", "")))
                    rate_unit = "‰" if "费率(‰)" in s else ("%" if "费率(%)" in s else "")
                    fee_val = s.get("费用(万元)", "")
                    steps_md += f"| {qujian} | {amt} | {rate}{rate_unit} | **{fee_val}** |\n"
                steps_md += "\n"

        final_response = (
            f"## {fee_name}\n\n"
            f"**依据**：{_basis_md_links(basis)}\n\n"
            f"{steps_md}"
            f"---\n\n"
            f"### 计算结果\n\n"
            f"**{discounted_fee_wan} 万元**{discount_text}\n\n"
            f"{desc}"
        )

        # ── 操作按钮 ──
        col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 2])
        with col_btn1:
            if st.button("✅ 确认", type="primary", use_container_width=True, key=f"confirm_simple_{ft}"):
                st.session_state.messages.append({"role": "assistant", "content": final_response})
                st.session_state.pop(f"simple_discount_{ft}", None)
                del st.session_state.pending_simple_fee
                st.rerun()
        with col_btn2:
            if st.button("🗑 取消", use_container_width=True, key=f"cancel_simple_{ft}"):
                st.session_state.pop(f"simple_discount_{ft}", None)
                del st.session_state.pending_simple_fee
                st.rerun()

# ===== 造价咨询服务多选面板（持久化在聊天区外） =====
if "pending_cost_consulting" in st.session_state:
    ctx = st.session_state.pending_cost_consulting
    fee_result = ctx["fee_result"]
    default_discount = ctx.get("default_discount", 1.0)
    query = ctx.get("query", "")

    # 提取原始输入值
    jianan = fee_result.get("_jianan")
    shebei = fee_result.get("_shebei")
    total_invest = fee_result.get("_total_invest")
    base_amount = fee_result.get("_base_amount", 0) or 0
    is_hebei = fee_result.get("_is_hebei", False)

    st.divider()

    with st.container(border=True):
        col_title, col_badge = st.columns([3, 1])
        with col_title:
            st.markdown("## 📋 造价咨询服务选择")
        with col_badge:
            if is_hebei:
                show_policy_badge("冀建市研[2017]2号")
            else:
                show_policy_badge("津价房地[2008]136号")

        if is_hebei:
            # ─── 河北省模式 ───
            st.caption(f"**建安费**：{base_amount} 万元（设备费不计入取费基数）")
            if total_invest:
                st.caption(f"**工程总投资**：{total_invest} 万元（用于投资估算/概算编制/概算审核/竣工决算编制/造价鉴定）")
            st.caption("请勾选需要的服务子项，每项独立按差额定率分档累进计算后求和。")

            # 专业工程调整系数
            st.markdown("### 🔧 专业工程调整系数")
            prof_options = list(_HEBEI_PROFESSIONAL_COEFFICIENTS.keys())
            # 尝试从查询自动检测专业类型
            default_prof_idx = 12  # 默认"其他工程" (1.0)
            for j, pname in enumerate(prof_options):
                if pname in query:
                    default_prof_idx = j
                    break
            selected_prof = st.selectbox(
                "选择专业工程类别（附件2调整系数）",
                options=prof_options,
                index=default_prof_idx,
                key="cc_hebei_prof",
                help="根据冀建市研[2017]2号附件2，不同专业工程适用不同调整系数。",
            )
            professional_coef = _HEBEI_PROFESSIONAL_COEFFICIENTS[selected_prof]
            st.caption(f"当前系数：**{professional_coef}**（{selected_prof}）")

            st.markdown("---")

            # ── 分类展示服务子项（河北省）──
            st.markdown("### 编制类")
            selected_services: list[str] = []
            bianzhi_hebei = [
                "投资估算", "经济评价", "概算编制",
                "预算编制", "工程量清单编制(审核)",
                "招标控制价编制(审核)", "结算编制",
                "竣工决算编制",
            ]
            cols = st.columns(3)
            for i, svc in enumerate(bianzhi_hebei):
                with cols[i % 3]:
                    label = svc
                    if svc in ("投资估算", "经济评价"):
                        label = f"{svc}（基数=投资估算造价）"
                    elif svc == "概算编制":
                        label = f"{svc}（基数=设计概算造价）"
                    elif svc == "竣工决算编制":
                        label = f"{svc}（基数=总投资）"
                    elif svc in ("预算编制", "工程量清单编制(审核)", "招标控制价编制(审核)", "结算编制"):
                        label = f"{svc}（基数=建安费）"
                    if st.checkbox(label, value=(svc == "预算编制"), key=f"cc_hb_{svc}"):
                        selected_services.append(svc)

            st.markdown("### 审核类")
            shenhe_hebei = ["概算审核", "预算审核", "结算审核"]
            cols2 = st.columns(3)
            for i, svc in enumerate(shenhe_hebei):
                with cols2[i % 3]:
                    label = svc
                    if svc == "概算审核":
                        label = f"{svc}（基数=设计概算造价）"
                    elif svc in ("预算审核", "结算审核"):
                        label = f"{svc}（基数=建安费）"
                    if st.checkbox(label, key=f"cc_hb_{svc}"):
                        selected_services.append(svc)

            st.markdown("### 全过程 / 其他")
            quanguocheng_hebei = [
                "投标报价分析(清标)", "施工阶段造价咨询",
                "全过程造价咨询", "工程造价鉴定",
            ]
            cols3 = st.columns(3)
            for i, svc in enumerate(quanguocheng_hebei):
                with cols3[i % 3]:
                    label = svc
                    if svc == "投标报价分析(清标)":
                        label = f"{svc}（基数=最高投标限价）"
                    elif svc in ("施工阶段造价咨询", "全过程造价咨询"):
                        label = f"{svc}（基数=建安费）"
                    elif svc == "工程造价鉴定":
                        label = f"{svc}（基数=鉴定标的额）"
                    if st.checkbox(label, key=f"cc_hb_{svc}"):
                        selected_services.append(svc)

            st.markdown("---")

            # ── 需要总投资的子项 ──
            calc_total_invest = total_invest
            needs_total = [s for s in selected_services if s in ("投资估算", "经济评价", "概算编制", "概算审核", "竣工决算编制", "工程造价鉴定")]
            if needs_total and total_invest is None:
                st.warning(
                    f"⚠️ **{'、'.join(needs_total)}**需要**工程总投资/鉴定标的额**，"
                    "请在下方面输入金额，否则该子项将无法计算。"
                )
                user_total = st.number_input(
                    "总投资 / 鉴定标的额（万元）",
                    min_value=0.0,
                    value=float(base_amount) if base_amount else 0.0,
                    step=1.0,
                    format="%.2f",
                    key="cc_hebei_total_invest",
                    help="竣工决算编制以总投资为基数，工程造价鉴定以鉴定标的额为基数。",
                )
                calc_total_invest = user_total if user_total > 0 else None

            # ── 计算选中服务（河北省）──
            if selected_services:
                try:
                    multi_result = calc_cost_consulting_multi_hebei(
                        selected_services,
                        base_amount,
                        total_investment=calc_total_invest,
                        professional_coef=professional_coef,
                        discount_coef=1.0,  # 在打折步骤统一处理
                    )
                    detail_list = multi_result["明细"]
                    total_fee = multi_result["合计(万元)"]

                    warnings = multi_result.get("参数", {}).get("警告", [])
                    if warnings:
                        for w in warnings:
                            st.warning(w)

                    # 明细表
                    st.markdown("### 费用明细")
                    detail_rows = []
                    for d in detail_list:
                        svc_name = d["服务类型"]
                        svc_base = d["计费基数(万元)"]
                        svc_fee = d["费用(万元)"]
                        detail_rows.append(f"| **{svc_name}** | {svc_base} | **{svc_fee}** |")
                    st.markdown(
                        "| 服务类型 | 计费基数（万元） | 费用（万元） |\n"
                        "|:--|:--|:--|\n" + "\n".join(detail_rows)
                    )
                    prof_note = f"（专业系数 {professional_coef}）" if abs(professional_coef - 1.0) > 0.001 else ""
                    st.markdown(f"### 💰 合计：**{total_fee} 万元** {prof_note}")

                    # 各子项分档计算展开
                    with st.expander("📐 查看各子项分档计算过程"):
                        for d in detail_list:
                            svc_name = d["服务类型"]
                            svc_fee = d["费用(万元)"]
                            svc_steps = d["计算步骤"]
                            st.markdown(f"#### {svc_name}（{svc_fee} 万元）")
                            if svc_steps:
                                step_rows = []
                                for s in svc_steps:
                                    qujian = s.get("区间", "")
                                    amt = s.get("金额(万元)", "")
                                    rate = s.get("费率(‰)", s.get("费率(%)", ""))
                                    rate_unit = "‰" if "费率(‰)" in s else "%"
                                    fee_v = s.get("费用(万元)", "")
                                    step_rows.append(f"| {qujian} | {amt} | {rate}{rate_unit} | **{fee_v}** |")
                                st.markdown(
                                    "| 区间（万元） | 金额（万元） | 费率 | 费用（万元） |\n"
                                    "|:--|:--|:--|:--|\n" + "\n".join(step_rows)
                                )
                            st.markdown("")

                    # ── 打折系数 ──
                    st.markdown("---")
                    st.markdown("### 🏷️ 费用打折")
                    discount_coef = st.number_input(
                        "打折系数（1.0 = 不打折，0.8 = 打八折）",
                        min_value=0.01, max_value=2.00,
                        value=default_discount, step=0.05,
                        format="%.2f",
                        key=f"cc_hb_discount",
                        help="输入打折系数调整最终总费用。",
                    )
                    discounted_total = round(total_fee * discount_coef, 4)
                    discount_text = ""
                    if abs(discount_coef - 1.0) < 0.005:
                        st.info(f"**不打折**，最终总费用：**{discounted_total} 万元**")
                    elif discount_coef < 1.0:
                        st.warning(
                            f"打折系数 **{discount_coef:.2f}** → "
                            f"{total_fee} 万 × {discount_coef:.2f} = "
                            f"**{discounted_total} 万元**"
                            f"（节省 {round(total_fee - discounted_total, 4)} 万元）"
                        )
                        discount_text = (
                            f"\n\n**打折系数**：{discount_coef:.2f}\n\n"
                            f"**打折后总费用**：{discounted_total} 万元"
                            f"（{total_fee} 万 × {discount_coef:.2f}）"
                        )
                    else:
                        st.warning(
                            f"上浮系数 **{discount_coef:.2f}** → "
                            f"{total_fee} 万 × {discount_coef:.2f} = "
                            f"**{discounted_total} 万元**"
                            f"（增加 {round(discounted_total - total_fee, 4)} 万元）"
                        )
                        discount_text = (
                            f"\n\n**上浮系数**：{discount_coef:.2f}\n\n"
                            f"**上浮后总费用**：{discounted_total} 万元"
                            f"（{total_fee} 万 × {discount_coef:.2f}）"
                        )

                    # ── 构建最终响应（河北省）──
                    detail_md = "### 费用明细\n\n"
                    detail_md += "| 服务类型 | 计费基数（万元） | 费用（万元） |\n"
                    detail_md += "|:--|:--|:--|\n"
                    for d in detail_list:
                        detail_md += f"| **{d['服务类型']}** | {d['计费基数(万元)']} | **{d['费用(万元)']}** |\n"
                    detail_md += f"\n### 💰 合计：**{total_fee} 万元**\n"

                    prof_text = ""
                    if abs(professional_coef - 1.0) > 0.001:
                        prof_text = f"\n**专业工程调整系数**：{professional_coef}（{selected_prof}）\n"

                    final_response = (
                        f"## 造价咨询费（河北省）\n\n"
                        f"**依据**：{_basis_md_links('《河北省建设工程造价咨询服务收费管理暂行办法》（冀建市研[2017]2号）')}\n\n"
                        f"**建安费**：{base_amount} 万元（不含设备费）\n"
                        f"（投资估算/概算编制/概算审核/竣工决算编制以总投资为基数，详见明细）\n"
                        f"{prof_text}\n"
                        f"{detail_md}\n"
                        f"---\n\n"
                        f"{discount_text}\n\n"
                        f"共 {len(selected_services)} 项服务，"
                        f"最终总费用 **{discounted_total} 万元**"
                        f"（可下浮 ≤20%）"
                    )

                except Exception as e:
                    st.error(f"计算出错：{e}")
                    import traceback
                    st.code(traceback.format_exc())
                    total_fee = 0
                    discounted_total = 0
                    discount_text = ""
                    final_response = ""
            else:
                st.warning("请至少选择一项服务")
                total_fee = 0
                discounted_total = 0
                discount_text = ""
                final_response = ""

        else:
            # ─── 天津市模式（原有逻辑）───
            st.caption(f"**工程费用**：{base_amount} 万元（建安 {jianan} 万 + 设备 {shebei or 0} 万）")
            if total_invest:
                st.caption(f"**工程总投资**：{total_invest} 万元（用于审核概算）")
            st.caption("请勾选需要的服务子项，每项独立按差额定率分档累进计算后求和。")

            # ── 分类展示服务子项 ──
            st.markdown("### 编制类（基数 = 工程费用）")
            selected_services: list[str] = []
            bianzhi_gongcheng = [
                "编制工程量清单", "编制标底(含清单)", "编制施工图预算",
                "编制竣工结算", "施工阶段全过程造价控制",
            ]
            cols = st.columns(3)
            for i, svc in enumerate(bianzhi_gongcheng):
                with cols[i % 3]:
                    if st.checkbox(svc, value=(svc == "编制施工图预算"), key=f"cc_{svc}"):
                        selected_services.append(svc)

            st.markdown("### 审核类")
            shenhe_services = ["审核概算", "审核预算、标底", "审核竣工结算"]
            cols2 = st.columns(3)
            for i, svc in enumerate(shenhe_services):
                with cols2[i % 3]:
                    label = svc
                    if svc == "审核概算":
                        label = f"{svc}（基数=总投资）"
                    if st.checkbox(label, key=f"cc_{svc}"):
                        selected_services.append(svc)

            st.markdown("### 其他（基数 = 建安工程费用）")
            other_services = ["编制项目投资估算", "编制设计概算"]
            cols3 = st.columns(3)
            for i, svc in enumerate(other_services):
                with cols3[i % 3]:
                    label = f"{svc}（基数=建安费）"
                    if st.checkbox(label, key=f"cc_{svc}"):
                        selected_services.append(svc)

            st.markdown("---")

            # ── 审核概算需要总投资 ──
            calc_total_invest = total_invest
            if "审核概算" in selected_services and total_invest is None:
                st.warning(
                    "⚠️ **审核概算**的计费基数是**工程总投资**，请在下方面输入总投资金额，"
                    "否则该子项将无法计算。"
                )
                user_total = st.number_input(
                    "工程总投资（万元）",
                    min_value=0.0,
                    value=float(base_amount) if base_amount else 0.0,
                    step=1.0,
                    format="%.2f",
                    key="cc_total_invest",
                    help="审核概算以工程总投资为计费基数，非工程费用。",
                )
                calc_total_invest = user_total if user_total > 0 else None

            # ── 计算选中服务 ──
            if selected_services:
                try:
                    multi_result = calc_cost_consulting_multi(
                        selected_services,
                        base_amount,
                        jianan_only=jianan,
                        total_investment=calc_total_invest,
                    )
                    detail_list = multi_result["明细"]
                    total_fee = multi_result["合计(万元)"]

                    # 显示警告（如审核概算总投资未知）
                    warnings = multi_result.get("参数", {}).get("警告", [])
                    if warnings:
                        for w in warnings:
                            st.warning(w)

                    # 明细表
                    st.markdown("### 费用明细")
                    detail_rows = []
                    for d in detail_list:
                        svc_name = d["服务类型"]
                        svc_base = d["计费基数(万元)"]
                        svc_fee = d["费用(万元)"]
                        detail_rows.append(f"| **{svc_name}** | {svc_base} | **{svc_fee}** |")
                    st.markdown(
                        "| 服务类型 | 计费基数（万元） | 费用（万元） |\n"
                        "|:--|:--|:--|\n" + "\n".join(detail_rows)
                    )
                    st.markdown(f"### 💰 合计：**{total_fee} 万元**")

                    # 各子项分档计算展开
                    with st.expander("📐 查看各子项分档计算过程"):
                        for d in detail_list:
                            svc_name = d["服务类型"]
                            svc_fee = d["费用(万元)"]
                            svc_steps = d["计算步骤"]
                            st.markdown(f"#### {svc_name}（{svc_fee} 万元）")
                            if svc_steps:
                                step_rows = []
                                for s in svc_steps:
                                    qujian = s.get("区间", "")
                                    amt = s.get("金额(万元)", "")
                                    rate = s.get("费率(‰)", s.get("费率(%)", ""))
                                    rate_unit = "‰" if "费率(‰)" in s else "%"
                                    fee_v = s.get("费用(万元)", "")
                                    step_rows.append(f"| {qujian} | {amt} | {rate}{rate_unit} | **{fee_v}** |")
                                st.markdown(
                                    "| 区间（万元） | 金额（万元） | 费率 | 费用（万元） |\n"
                                    "|:--|:--|:--|:--|\n" + "\n".join(step_rows)
                                )
                            st.markdown("")

                    # ── 打折系数 ──
                    st.markdown("---")
                    st.markdown("### 🏷️ 费用打折")
                    discount_coef = st.number_input(
                        "打折系数（1.0 = 不打折，0.8 = 打八折）",
                        min_value=0.01, max_value=2.00,
                        value=default_discount, step=0.05,
                        format="%.2f",
                        key=f"cc_discount",
                        help="输入打折系数调整最终总费用。",
                    )
                    discounted_total = round(total_fee * discount_coef, 4)
                    discount_text = ""
                    if abs(discount_coef - 1.0) < 0.005:
                        st.info(f"**不打折**，最终总费用：**{discounted_total} 万元**")
                    elif discount_coef < 1.0:
                        st.warning(
                            f"打折系数 **{discount_coef:.2f}** → "
                            f"{total_fee} 万 × {discount_coef:.2f} = "
                            f"**{discounted_total} 万元**"
                            f"（节省 {round(total_fee - discounted_total, 4)} 万元）"
                        )
                        discount_text = (
                            f"\n\n**打折系数**：{discount_coef:.2f}\n\n"
                            f"**打折后总费用**：{discounted_total} 万元"
                            f"（{total_fee} 万 × {discount_coef:.2f}）"
                        )
                    else:
                        st.warning(
                            f"上浮系数 **{discount_coef:.2f}** → "
                            f"{total_fee} 万 × {discount_coef:.2f} = "
                            f"**{discounted_total} 万元**"
                            f"（增加 {round(discounted_total - total_fee, 4)} 万元）"
                        )
                        discount_text = (
                            f"\n\n**上浮系数**：{discount_coef:.2f}\n\n"
                            f"**上浮后总费用**：{discounted_total} 万元"
                            f"（{total_fee} 万 × {discount_coef:.2f}）"
                        )

                    # ── 构建最终响应 ──
                    detail_md = "### 费用明细\n\n"
                    detail_md += "| 服务类型 | 计费基数（万元） | 费用（万元） |\n"
                    detail_md += "|:--|:--|:--|\n"
                    for d in detail_list:
                        detail_md += f"| **{d['服务类型']}** | {d['计费基数(万元)']} | **{d['费用(万元)']}** |\n"
                    detail_md += f"\n### 💰 合计：**{total_fee} 万元**\n"

                    final_response = (
                        f"## 造价咨询费\n\n"
                        f"**依据**：{_basis_md_links('《天津市建设工程造价咨询服务项目和价格标准》（津价房地[2008]136号）')}\n\n"
                        f"**计费基数**：工程费用 {base_amount} 万元（建安 {jianan} 万 + 设备 {shebei or 0} 万）\n\n"
                        f"{detail_md}\n"
                        f"---\n\n"
                        f"{discount_text}\n\n"
                        f"共 {len(selected_services)} 项服务，"
                        f"最终总费用 **{discounted_total} 万元**"
                    )

                except Exception as e:
                    st.error(f"计算出错：{e}")
                    import traceback
                    st.code(traceback.format_exc())
                    total_fee = 0
                    discounted_total = 0
                    discount_text = ""
                    final_response = ""
            else:
                st.warning("请至少选择一项服务")
                total_fee = 0
                discounted_total = 0
                discount_text = ""
                final_response = ""

        # ── 操作按钮 ──
        col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 2])
        with col_btn1:
            confirm_key = "confirm_cc_hb" if is_hebei else "confirm_cc"
            if st.button("✅ 确认", type="primary", use_container_width=True, key=confirm_key):
                if not selected_services:
                    st.warning("请先选择至少一项服务")
                    st.stop()
                if final_response:
                    st.session_state.messages.append({"role": "assistant", "content": final_response})
                # 清理所有 cc_ 前缀的 session state keys
                for key in list(st.session_state.keys()):
                    if key.startswith("cc_"):
                        st.session_state.pop(key, None)
                del st.session_state.pending_cost_consulting
                st.rerun()
        with col_btn2:
            cancel_key = "cancel_cc_hb" if is_hebei else "cancel_cc"
            if st.button("🗑 取消", use_container_width=True, key=cancel_key):
                for key in list(st.session_state.keys()):
                    if key.startswith("cc_"):
                        st.session_state.pop(key, None)
                del st.session_state.pending_cost_consulting
                st.rerun()

# ===== 依赖费种交互式配置（招标代理费 & 施工图审查费）=====
if "pending_dependent_fee" in st.session_state:
    ctx = st.session_state.pending_dependent_fee
    target_name = ctx.get("target_fee_name", "")
    dep_fees = ctx.get("dependent_fees", [])
    base_params = ctx.get("base_params", {})
    step = ctx.get("step", "config")
    discount_coef = ctx.get("discount_coef", 1.0)

    st.divider()

    if step == "config":
        # ── 配置阶段：依次展示各依赖费种 ──
        with st.container(border=True):
            col_t, col_b = st.columns([3, 1])
            with col_t:
                st.markdown(f"## 🔗 {target_name} — 依赖费种配置")
            with col_b:
                st.info(f"{len(dep_fees)} 个依赖费种")

            st.caption(
                f"计算 **{target_name}** 前需先确定以下费种的参数。"
                f"请为每个依赖费种选择合适的系数或费率，然后点击「确认并计算」。"
            )

            # ── 施工图审查费：审查费率下拉选择 ──
            if ctx.get("target_fee") == "施工图审查费":
                from fee_engine import (SHIGONG_SHENCHA_RATES, SHIGONG_SHENCHA_ZHUZHAI,
                                        HEBEI_SHENCHA_RATE)
                _query = ctx.get("query", "")
                pt = base_params.get("project_type_shencha", "公建")
                size = base_params.get("size", "中型")

                if is_hebei_region(st.session_state.get("selected_region")):
                    # 河北省：发改价格〔2011〕534号，统一 6.5%
                    rate_options = [
                        (f"河北省 — {HEBEI_SHENCHA_RATE}%（发改价格〔2011〕534号）",
                         f"河北|—|{HEBEI_SHENCHA_RATE}",
                         str(HEBEI_SHENCHA_RATE), "河北", "—"),
                    ]
                    default_idx = 0
                else:
                    # 天津/默认：津价管[2011]46号
                    rate_options: list[tuple[str, str, str, str, str]] = []
                    for r_pt, sizes in SHIGONG_SHENCHA_RATES.items():
                        for r_sz, r_val in sizes.items():
                            label = f"{r_pt} · {r_sz} — {r_val}%"
                            key = f"{r_pt}|{r_sz}|{r_val}"
                            rate_options.append((label, key, str(r_val), r_pt, r_sz))
                    # 住宅类：按建筑面积 × 单价（元/m²）
                    for r_sz, r_val in SHIGONG_SHENCHA_ZHUZHAI.items():
                        label = f"住宅 · {r_sz} — {r_val} 元/m²"
                        key = f"住宅|{r_sz}|{r_val}"
                        rate_options.append((label, key, str(r_val), "住宅", r_sz))

                    # 默认选中当前检测到的项目和规模
                    default_key = f"{pt}|{size}|{SHIGONG_SHENCHA_RATES.get(pt, {}).get(size, SHIGONG_SHENCHA_ZHUZHAI.get(size, 1.7))}"
                    default_idx = 0
                    for i, (_, rk, _, _, _) in enumerate(rate_options):
                        if rk == default_key:
                            default_idx = i
                            break

                st.markdown("#### 📊 审查费率")
                st.caption("选择项目类型和规模对应的审查费率：")
                selected_label = st.selectbox(
                    "审查费率",
                    range(len(rate_options)),
                    index=default_idx,
                    format_func=lambda i: rate_options[i][0] if i < len(rate_options) else "",
                    key="shencha_rate_select",
                    label_visibility="collapsed",
                )
                _, sel_key, sel_rate, sel_pt, sel_sz = rate_options[selected_label]
                # 将选中的费率和类型/规模存入 ctx
                ctx["shencha_rate_override"] = float(sel_rate)
                ctx["shencha_pt_override"] = sel_pt
                ctx["shencha_size_override"] = sel_sz

                if sel_pt == "住宅":
                    st.info(f"当前选择：**住宅 · {sel_sz}** → **{sel_rate} 元/m²**\n\n"
                            f"住宅类审查费按建筑面积计费。请在下方输入建筑面积。")
                    prev_area_sf = ctx.get("shencha_area_override") or 0.0
                    shencha_area_sf = st.number_input(
                        "建筑面积（m²）",
                        min_value=0.0,
                        value=float(prev_area_sf) if prev_area_sf else 0.0,
                        step=100.0,
                        format="%.0f",
                        key="shencha_area_input",
                        help="住宅类施工图审查费 = 建筑面积 × 单价\n"
                             "大型 1.9 元/m² / 中型 1.7 元/m² / 小型 1.3 元/m²",
                    )
                    ctx["shencha_area_override"] = shencha_area_sf
                elif sel_pt == "河北":
                    st.info(f"当前选择：**河北省 — {sel_rate}%**\n\n"
                            f"依据：发改价格〔2011〕534号，计费基数 = 勘察费 + 设计费")
                else:
                    st.info(f"当前选择：**{sel_pt} · {sel_sz}** → **{sel_rate}%**")
                st.markdown("---")

            # 收集各依赖费种的配置
            all_configs: dict = {}
            dep_previews: dict = {}

            for i, dep in enumerate(dep_fees):
                dep_type = dep.get("fee_type", "")
                dep_label = dep.get("fee_label", "")
                config_type = dep.get("config_type", "coef")

                st.markdown("---")
                st.markdown(f"### 📋 依赖费种 {i + 1}：{dep_label}")

                # ── 自定义金额选项 ──
                dep_custom_amounts = ctx.setdefault("dep_custom_amounts", {})
                use_custom = st.checkbox(
                    f"✏️ 使用自定义金额（覆盖标准计算）",
                    value=dep_type in dep_custom_amounts,
                    key=f"dep_use_custom_{dep_type}",
                    help=f"勾选后将直接输入 {dep_label} 的金额，跳过系数/费率计算",
                )

                if use_custom:
                    custom_val = st.number_input(
                        f"{dep_label} 自定义金额（万元）",
                        min_value=0.0, step=0.1,
                        value=float(dep_custom_amounts.get(dep_type, 0.0)),
                        format="%.2f",
                        key=f"dep_custom_val_{dep_type}",
                    )
                    if custom_val > 0:
                        dep_custom_amounts[dep_type] = custom_val
                        dep_previews[dep_type] = custom_val
                        all_configs[dep_type] = {"_custom_amount": custom_val}
                        st.info(f"💡 自定义金额：**{custom_val:.2f} 万元**")

                        # 自定义金额也可打折
                        dep_discounts = ctx.setdefault("dep_discounts", {})
                        cur_dep_disc = dep_discounts.get(dep_type, 1.0)
                        dep_disc = st.number_input(
                            f"{dep_label} 打折系数",
                            min_value=0.01, max_value=2.00,
                            value=float(cur_dep_disc), step=0.05,
                            format="%.2f",
                            key=f"dep_discount_{dep_type}",
                            help=f"1.0 = 不打折，将对 {dep_label} 的金额打折",
                        )
                        dep_discounts[dep_type] = dep_disc
                        if abs(dep_disc - 1.0) >= 0.005:
                            disc_preview = round(custom_val * dep_disc, 4)
                            st.caption(f"  打折后：{custom_val:.2f} × {dep_disc:.2f} = **{disc_preview:.2f}** 万元")
                    else:
                        dep_custom_amounts.pop(dep_type, None)
                elif config_type == "coef":
                    # ── 系数下拉选择 ──
                    coef_meta = dep.get("coef_metadata", {})
                    coefs = coef_meta.get("coefs", [])
                    dep_base = dep.get("base_params", {})
                    selected_coefs: dict = {}

                    for j, coef_def in enumerate(coefs):
                        key = coef_def["key"]
                        param_name = coef_def["param_name"]
                        current_val = float(coef_def["current"])
                        options = coef_def.get("options", [])
                        desc_text = coef_def.get("description", "")

                        st.markdown(f"**{key}**")
                        if desc_text:
                            st.caption(desc_text)

                        # 构建选项
                        option_labels = [f"{label}（{val}）" for label, val in options]
                        option_values = [val for _, val in options]

                        try:
                            current_idx = option_values.index(current_val)
                        except ValueError:
                            current_idx = len(option_values)

                        option_labels.append("✏️ 自定义…")
                        option_values.append(-1.0)

                        sel_idx = st.selectbox(
                            f"选择{key}",
                            range(len(option_labels)),
                            index=min(current_idx, len(option_labels) - 1),
                            format_func=lambda idx, labels=option_labels: labels[idx],
                            key=f"dep_coef_{dep_type}_{param_name}",
                            label_visibility="collapsed",
                        )

                        chosen_val = option_values[sel_idx]
                        if chosen_val == -1.0:
                            custom_val = st.number_input(
                                f"自定义{key}的值",
                                min_value=0.10, max_value=5.00,
                                value=current_val if current_val > 0.1 else 1.0,
                                step=0.05, format="%.2f",
                                key=f"dep_coef_cust_{dep_type}_{param_name}",
                            )
                            selected_coefs[param_name] = float(custom_val)
                        else:
                            selected_coefs[param_name] = float(chosen_val)

                    all_configs[dep_type] = selected_coefs

                    # 实时预览
                    try:
                        if dep_type == "监理费":
                            prof = selected_coefs.get("professional_coef", 1.0)
                            comp = selected_coefs.get("complexity_coef", 1.0)
                            elev = selected_coefs.get("elevation_coef", 1.0)
                            jn = dep_base.get("jianan", 0) or 0
                            sb = dep_base.get("shebei", 0) or 0
                            aw = dep_base.get("amount_wan", jn + sb)
                            if jn > 0 or sb > 0:
                                preview_r = calc_jianli(jianan=jn, shebei=sb,
                                                        professional_coef=prof,
                                                        complexity_coef=comp,
                                                        elevation_coef=elev)
                            else:
                                preview_r = calc_jianli(amount_wan=aw,
                                                        professional_coef=prof,
                                                        complexity_coef=comp,
                                                        elevation_coef=elev)
                            preview_fee = preview_r["结果(万元)"]
                        elif dep_type == "工程设计费":
                            prof = selected_coefs.get("professional_coef", 1.0)
                            comp = selected_coefs.get("complexity_coef", 1.0)
                            addi = selected_coefs.get("additional_coef", 1.0)
                            aw = dep_base.get("amount_wan", 0)
                            addi_list = [addi] if abs(addi - 1.0) > 0.005 else None
                            preview_r = calc_sheji(aw, prof, comp, additional_coefs=addi_list)
                            preview_fee = preview_r["结果(万元)"]
                        else:
                            preview_fee = dep.get("preview_fee", 0)

                        dep_previews[dep_type] = preview_fee
                        st.info(f"💡 预览：{dep_label} = **{preview_fee:.2f} 万元**")
                    except Exception as e:
                        st.warning(f"⚠️ 预览计算失败：{e}")
                        dep_previews[dep_type] = 0

                    dep_discounts = ctx.setdefault("dep_discounts", {})
                    cur_dep_disc = dep_discounts.get(dep_type, 1.0)
                    dep_disc = st.number_input(
                        f"{dep_label} 打折系数",
                        min_value=0.01, max_value=2.00,
                        value=float(cur_dep_disc), step=0.05,
                        format="%.2f",
                        key=f"dep_discount_{dep_type}",
                        help=f"1.0 = 不打折，将对 {dep_label} 的计算结果打折",
                    )
                    dep_discounts[dep_type] = dep_disc
                    if abs(dep_disc - 1.0) >= 0.005 and dep_type in dep_previews:
                        disc_preview = round(dep_previews[dep_type] * dep_disc, 4)
                        st.caption(f"  打折后：{dep_previews[dep_type]:.2f} × {dep_disc:.2f} = **{disc_preview:.2f}** 万元")

                elif config_type == "rate":
                    # ── 费率单选 ──
                    rate_options = dep.get("rate_options", [])
                    default_idx = dep.get("default_rate_index", len(rate_options) // 2)
                    pt = dep.get("project_type", "通用")

                    st.caption(f"项目类型：**{pt}**")

                    # 构建选项
                    rate_labels = [
                        f"{opt['label']} → {opt['fee']:.2f} 万元"
                        for opt in rate_options
                    ]
                    chosen_label = st.radio(
                        "选择勘察费费率",
                        rate_labels,
                        index=min(default_idx, len(rate_labels) - 1),
                        key=f"dep_rate_sel_{dep_type}",
                    )
                    chosen_idx = rate_labels.index(chosen_label)
                    chosen_rate = rate_options[chosen_idx]["rate"]
                    chosen_fee = rate_options[chosen_idx]["fee"]

                    all_configs[dep_type] = {
                        "rate": chosen_rate,
                        "project_type": pt,
                    }
                    dep_previews[dep_type] = chosen_fee
                    st.info(f"💡 预览：勘察费 = **{chosen_fee:.2f} 万元**"
                            f"（费率 {chosen_rate}%）")

                    dep_discounts = ctx.setdefault("dep_discounts", {})
                    cur_dep_disc = dep_discounts.get(dep_type, 1.0)
                    dep_disc = st.number_input(
                        f"{dep_label} 打折系数",
                        min_value=0.01, max_value=2.00,
                        value=float(cur_dep_disc), step=0.05,
                        format="%.2f",
                        key=f"dep_discount_{dep_type}",
                        help=f"1.0 = 不打折，将对 {dep_label} 的计算结果打折",
                    )
                    dep_discounts[dep_type] = dep_disc
                    if abs(dep_disc - 1.0) >= 0.005:
                        disc_preview = round(chosen_fee * dep_disc, 4)
                        st.caption(f"  打折后：{chosen_fee:.2f} × {dep_disc:.2f} = **{disc_preview:.2f}** 万元")

            # ── 底部按钮 ──
            st.markdown("---")
            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                if st.button("✅ 确认并计算", use_container_width=True, key="confirm_dep_cfg"):
                    # 保存配置到 session state
                    ctx["all_configs"] = all_configs
                    ctx["dep_previews"] = dep_previews
                    ctx["step"] = "result"
                    # 调用引擎计算最终结果
                    try:
                        # 施工图审查费：传入用户选择的费率/类型/规模
                        if ctx.get("target_fee") == "施工图审查费":
                            base_params = dict(base_params)
                            base_params["shencha_rate_override"] = ctx.get("shencha_rate_override")
                            base_params["shencha_pt_override"] = ctx.get("shencha_pt_override")
                            base_params["shencha_size_override"] = ctx.get("shencha_size_override")
                            if ctx.get("shencha_pt_override") == "住宅":
                                base_params["shencha_area_override"] = ctx.get("shencha_area_override")

                        final = resolve_dependent_calc(
                            ctx["target_fee"],
                            all_configs,
                            base_params,
                            dep_discounts=ctx.get("dep_discounts"),
                            region=st.session_state.get("selected_region"),
                        )
                        ctx["final_result"] = final
                    except Exception as e:
                        ctx["final_result"] = {"error": str(e)}
                    ctx["discount_coef"] = 1.0
                    st.rerun()
            with col_btn2:
                if st.button("🗑 取消", use_container_width=True, key="cancel_dep_cfg"):
                    del st.session_state.pending_dependent_fee
                    st.rerun()

    elif step == "result":
        # ── 结果阶段：展示最终费用 + 打折 ──
        final = ctx.get("final_result", {})
        all_configs = ctx.get("all_configs", {})
        dep_previews = ctx.get("dep_previews", {})

        if final.get("error"):
            st.error(f"计算失败：{final['error']}")
            if st.button("🔙 返回重新配置"):
                ctx["step"] = "config"
                st.rerun()
        else:
            with st.container(border=True):
                st.markdown(f"## 💰 {target_name} — 计算结果")

                # 依赖费种回顾
                st.markdown("### 依赖费种（用户配置）")
                dep_detail = final.get("_dependent_details", {})
                dep_discounts = ctx.get("dep_discounts", {})
                dep_cols = st.columns(len(dep_detail) if dep_detail else 1)
                dep_discounted: dict = {}
                for idx, (dtype, dr) in enumerate(dep_detail.items()):
                    fee_val = dr.get("结果(万元)", dr.get("结果中值(万元)", 0))
                    ddisc = dep_discounts.get(dtype, 1.0)
                    ddisc_val = round(fee_val * ddisc, 4)
                    dep_discounted[dtype] = ddisc_val
                    label_map = {
                        "监理费": "施工监理服务费",
                        "工程设计费": "工程设计费",
                        "勘察费": "工程勘察费",
                    }
                    with dep_cols[idx]:
                        if abs(ddisc - 1.0) >= 0.005:
                            st.metric(label=label_map.get(dtype, dtype),
                                      value=f"{ddisc_val} 万元",
                                      delta=f"折前 {fee_val:.2f}")
                        else:
                            st.metric(label=label_map.get(dtype, dtype), value=f"{fee_val} 万元")

                st.markdown("---")

                # 目标费种结果
                is_zb = final.get("is_zhaobiao_multi", False)
                if is_zb:
                    # 招标代理费多类型展示
                    detail_list = final.get("明细", [])
                    total_fee = final.get("合计(万元)", 0)

                    st.markdown("### 招标代理费明细")
                    detail_rows = []
                    for d in detail_list:
                        dtype = d.get("类型", "")
                        dbase = d.get("基数(万元)", 0)
                        dsrc = d.get("基数来源", "")
                        dfee = d.get("费用(万元)", 0)
                        note = d.get("说明", "")
                        if note:
                            detail_rows.append(
                                f"| **{dtype}** | {dsrc} | {dbase:.2f} | ⚠️ {note} |"
                            )
                        else:
                            detail_rows.append(
                                f"| **{dtype}** | {dsrc} | {dbase:.2f} | **{dfee}** |"
                            )
                    st.markdown(
                        "| 类型 | 基数来源 | 基数（万元） | 费用（万元） |\n"
                        "|:--|:--|:--|:--|\n" + "\n".join(detail_rows)
                    )

                    # 分档计算过程
                    with st.expander("📐 查看各子项分档计算过程"):
                        for d in detail_list:
                            steps = d.get("计算步骤", [])
                            if steps:
                                st.markdown(f"**{d.get('类型', '')}**")
                                for s in steps:
                                    st.caption(
                                        f"{s.get('步骤', '')}：{s.get('公式', '')} → {s.get('结果', '')}"
                                    )
                                st.markdown("---")

                    base_fee = total_fee
                    st.markdown(f"### 💰 招标代理费合计：**{total_fee} 万元**")
                else:
                    # 施工图审查费展示
                    steps = final.get("计算步骤", [])
                    if steps:
                        st.markdown("### 计算步骤")
                        for s in steps:
                            st.caption(
                                f"**{s.get('步骤', '')}**：{s.get('公式', '')} → {s.get('结果', '')}"
                            )
                    base_fee = final.get("结果(万元)", 0)
                    st.markdown(f"### 💰 审查费：**{base_fee} 万元**")

                # ── 打折系数 ──
                st.markdown("---")
                st.markdown("### 💰 费用打折")
                discount_coef = st.number_input(
                    "打折系数（1.0 = 不打折，0.8 = 打八折）",
                    min_value=0.01, max_value=2.00,
                    value=ctx.get("discount_coef", 1.0), step=0.05,
                    format="%.2f",
                    key="discount_dep_final",
                    help="输入打折系数调整最终费用。",
                )
                ctx["discount_coef"] = discount_coef
                discounted_fee = round(base_fee * discount_coef, 4)

                if abs(discount_coef - 1.0) < 0.005:
                    st.info(f"**不打折**，最终费用：**{discounted_fee} 万元**")
                elif discount_coef < 1.0:
                    st.warning(
                        f"打折系数 **{discount_coef:.2f}** → "
                        f"{base_fee:.2f} 万 × {discount_coef:.2f} = "
                        f"**{discounted_fee} 万元**"
                        f"（节省 {round(base_fee - discounted_fee, 4)} 万元）"
                    )
                else:
                    st.warning(
                        f"上浮系数 **{discount_coef:.2f}** → "
                        f"{base_fee:.2f} 万 × {discount_coef:.2f} = "
                        f"**{discounted_fee} 万元**"
                        f"（增加 {round(discounted_fee - base_fee, 4)} 万元）"
                    )

                st.markdown("---")

                # ── 确认和取消按钮 ──
                col_btn1, col_btn2, col_btn3 = st.columns(3)
                with col_btn1:
                    if st.button("✅ 确认结果", use_container_width=True, key="confirm_dep_result"):
                        # 构建最终响应文本
                        fee_name = final.get("费种", target_name)
                        basis = final.get("依据", "")
                        desc = final.get("说明", "")

                        # 添加依赖费种信息（含自定义金额和各自打折）
                        dep_parts = []
                        dep_custom_amounts = ctx.get("dep_custom_amounts", {})
                        for dtype, dr in dep_detail.items():
                            label_map = {
                                "监理费": "施工监理服务费",
                                "工程设计费": "工程设计费",
                                "勘察费": "工程勘察费",
                            }
                            dl = label_map.get(dtype, dtype)
                            fv = dr.get("结果(万元)", dr.get("结果中值(万元)", 0))
                            ddisc = dep_discounts.get(dtype, 1.0)

                            if dtype in dep_custom_amounts:
                                base_note = f"自定义 {dep_custom_amounts[dtype]:.2f} 万"
                            else:
                                base_note = f"{fv:.2f} 万"

                            if abs(ddisc - 1.0) >= 0.005:
                                ddisc_val = round(fv * ddisc, 4)
                                dep_parts.append(
                                    f"- {dl}：**{ddisc_val} 万元**"
                                    f"（{base_note} × {ddisc:.2f}）"
                                )
                            else:
                                dep_parts.append(f"- {dl}：**{fv}** 万元 ({base_note})")

                        discount_text = ""
                        if abs(discount_coef - 1.0) >= 0.005:
                            discount_text = (
                                f"\n\n**目标费打折系数**：{discount_coef:.2f}\n\n"
                                f"**打折后目标费**：{round(base_fee * discount_coef, 4)} 万元"
                                f"（{base_fee:.2f} 万 × {discount_coef:.2f}）"
                            )

                        final_response = (
                            f"## {fee_name}\n\n"
                            f"**依据**：{_basis_md_links(basis)}\n\n"
                            f"### 依赖费种\n\n"
                            + "\n".join(dep_parts) +
                            f"\n\n---\n\n"
                            f"{desc}"
                            f"{discount_text}"
                            f"\n\n**最终费用**：**{discounted_fee} 万元**"
                        )
                        st.session_state.messages.append({
                            "role": "assistant",
                            "content": final_response,
                        })
                        del st.session_state.pending_dependent_fee
                        st.rerun()
                with col_btn2:
                    if st.button("🔙 返回配置", use_container_width=True, key="back_dep_result"):
                        ctx["step"] = "config"
                        st.rerun()
                with col_btn3:
                    if st.button("🗑 取消", use_container_width=True, key="cancel_dep_result"):
                        del st.session_state.pending_dependent_fee
                        st.rerun()

# ===== 环评费多服务类型选择 =====
if "pending_huanping" in st.session_state:
    ctx = st.session_state.pending_huanping
    amount_wan = ctx.get("amount_wan", 0)
    ind_coef = ctx.get("industry_coef", 1.0)
    ind_name = ctx.get("industry_name", "")
    sens_coef = ctx.get("sensitivity_coef", 1.0)
    discount_coef = ctx.get("discount_coef", 1.0)
    has_explicit = ctx.get("has_explicit_investment", False)
    estimated_investment = ctx.get("estimated_investment", amount_wan)

    st.divider()

    with st.container(border=True):
        st.markdown("## 🌿 环评费 — 服务类型选择")

        st.caption(
            "环境影响咨询费包含 **4 项服务类型**。"
            "请勾选需要计算的服务类型，调整系数后点击确认。"
        )

        # ── 计费基数：项目总投资 ──
        st.markdown("### 💰 计费基数 — 项目总投资")
        if not has_explicit:
            st.warning(
                "⚠️ 环评费计费基数为**项目总投资**（非建安费或设备费）。"
                "请在下框中输入项目总投资金额。"
            )
        investment_input = st.number_input(
            "项目总投资（万元）",
            min_value=0.0,
            value=float(estimated_investment),
            step=10.0,
            format="%.2f",
            key="huanping_investment",
            help="环评费以项目总投资为计费基数（计价格[2002]125号）",
        )
        ctx["estimated_investment"] = investment_input

        # ── 服务类型多选 ──
        st.markdown("### 📋 选择服务类型")
        all_services = ["编制报告书", "编制报告表", "评估报告书", "评估报告表"]
        selected = []
        svc_col1, svc_col2 = st.columns(2)
        for i, svc in enumerate(all_services):
            col = svc_col1 if i < 2 else svc_col2
            with col:
                checked = st.checkbox(svc, value=True, key=f"huanping_svc_{i}")
                if checked:
                    selected.append(svc)

        if not selected:
            st.warning("⚠️ 请至少选择一项服务类型")
        else:
            st.markdown("---")

            # ── 系数调整 ──
            st.markdown("### 🎛️ 系数调整")

            # 行业调整系数
            from fee_engine import HUANPING_INDUSTRY_OPTIONS, HUANPING_SENSITIVITY_OPTIONS

            industry_labels = [f"{label}（{val}）" for label, val in HUANPING_INDUSTRY_OPTIONS]
            industry_values = [val for _, val in HUANPING_INDUSTRY_OPTIONS]
            try:
                cur_ind_idx = industry_values.index(ind_coef)
            except ValueError:
                cur_ind_idx = 0
            ind_label = st.selectbox(
                "行业调整系数",
                range(len(industry_labels)),
                index=cur_ind_idx,
                format_func=lambda idx: industry_labels[idx],
                key="huanping_ind_coef",
            )
            chosen_ind_coef = industry_values[ind_label]
            chosen_ind_name = HUANPING_INDUSTRY_OPTIONS[ind_label][0]

            # 环境敏感程度系数
            sens_labels = [f"{label}（{val}）" for label, val in HUANPING_SENSITIVITY_OPTIONS]
            sens_values = [val for _, val in HUANPING_SENSITIVITY_OPTIONS]
            try:
                cur_sens_idx = sens_values.index(sens_coef)
            except ValueError:
                cur_sens_idx = 0
            sens_label = st.selectbox(
                "环境敏感程度系数",
                range(len(sens_labels)),
                index=cur_sens_idx,
                format_func=lambda idx: sens_labels[idx],
                key="huanping_sens_coef",
            )
            chosen_sens_coef = sens_values[sens_label]

            st.markdown("---")

            # ── 实时费用预览 ──
            st.markdown("### 💡 费用预览")
            try:
                preview = calc_huanping_multi(
                    investment_input,
                    selected,
                    industry_coef=chosen_ind_coef,
                    industry_name=chosen_ind_name,
                    sensitivity_coef=chosen_sens_coef,
                )
                detail_list = preview.get("明细", [])
                total_fee = preview.get("合计(万元)", 0)

                # 明细表
                detail_rows = []
                for d in detail_list:
                    detail_rows.append(
                        f"| **{d['服务类型']}** "
                        f"| {d['结果(万元)']} "
                        f"| **{d['结果中值(万元)']}** |"
                    )
                st.markdown(
                    "| 服务类型 | 费用范围（万元） | 中值（万元） |\n"
                    "|----------|:--:|:--:|\n" + "\n".join(detail_rows)
                )

                st.markdown(f"### 💰 环评费合计（中值）：**{total_fee} 万元**")

                # 分档计算详情
                with st.expander("📐 查看各项计算过程"):
                    for d in detail_list:
                        st.markdown(f"**{d['服务类型']}**")
                        steps = d.get("计算步骤", [])
                        for s in steps:
                            st.caption(
                                f"{s.get('步骤', '')}：{s.get('公式', '')} → {s.get('结果', '')}"
                            )
                        st.markdown("---")

            except Exception as e:
                st.error(f"预览计算失败：{e}")
                total_fee = 0

            st.markdown("---")

            # ── 打折系数 ──
            st.markdown("### 💰 费用打折")
            discount_coef = st.number_input(
                "打折系数（1.0 = 不打折，0.8 = 打八折）",
                min_value=0.01, max_value=2.00,
                value=ctx.get("discount_coef", 1.0), step=0.05,
                format="%.2f",
                key="discount_huanping",
                help="输入打折系数调整最终费用。",
            )
            ctx["discount_coef"] = discount_coef
            discounted_total = round(total_fee * discount_coef, 4)

            if abs(discount_coef - 1.0) < 0.005:
                st.info(f"**不打折**，最终费用：**{discounted_total} 万元**")
            elif discount_coef < 1.0:
                st.warning(
                    f"打折系数 **{discount_coef:.2f}** → "
                    f"{total_fee:.2f} 万 × {discount_coef:.2f} = "
                    f"**{discounted_total} 万元**"
                    f"（节省 {round(total_fee - discounted_total, 4)} 万元）"
                )
            else:
                st.warning(
                    f"上浮系数 **{discount_coef:.2f}** → "
                    f"{total_fee:.2f} 万 × {discount_coef:.2f} = "
                    f"**{discounted_total} 万元**"
                    f"（增加 {round(discounted_total - total_fee, 4)} 万元）"
                )

            st.markdown("---")

            # ── 确认 / 取消 ──
            col_btn1, col_btn2 = st.columns(2)
            discount_text = ""
            if abs(discount_coef - 1.0) >= 0.005:
                discount_text = (
                    f"\n\n**打折系数**：{discount_coef:.2f}\n\n"
                    f"**打折后费用**：{discounted_total} 万元"
                    f"（{total_fee:.2f} 万 × {discount_coef:.2f}）"
                )

            with col_btn1:
                if st.button("✅ 确认结果", use_container_width=True, key="confirm_huanping"):
                    # 构建详情文本
                    detail_parts = []
                    for d in detail_list:
                        detail_parts.append(
                            f"- **{d['服务类型']}**：{d['结果(万元)']} 万元"
                            f"（中值 **{d['结果中值(万元)']}** 万元）"
                        )

                    final_response = (
                        f"## 环境影响咨询费\n\n"
                        f"**依据**：{_basis_md_links('《关于规范环境影响咨询收费有关问题的通知》（计价格[2002]125号）')}"
                        f"（计价格[2002]125号）\n\n"
                        f"**参数**："
                        f"项目总投资 {investment_input:.2f} 万元，"
                        f"行业「{chosen_ind_name}」系数 {chosen_ind_coef}，"
                        f"环境敏感程度系数 {chosen_sens_coef}\n\n"
                        f"### 服务类型明细\n\n"
                        + "\n".join(detail_parts) +
                        f"\n\n### 💰 合计（中值）：**{total_fee} 万元**"
                        f"{discount_text}"
                    )
                    st.session_state.messages.append({
                        "role": "assistant",
                        "content": final_response,
                    })
                    del st.session_state.pending_huanping
                    st.rerun()
            with col_btn2:
                if st.button("🗑 取消", use_container_width=True, key="cancel_huanping"):
                    del st.session_state.pending_huanping
                    st.rerun()

# ===== 建设项目前期工作咨询费多服务类型选择（单费种提问） =====
if "pending_keyan" in st.session_state:
    ctx = st.session_state.pending_keyan
    amount_yi = ctx.get("amount_yi", 0)
    ind_coef = ctx.get("industry_coef", 1.0)
    ind_name = ctx.get("industry_name", "")
    comp_coef = ctx.get("complexity_coef", 1.0)
    discount_coef = ctx.get("discount_coef", 1.0)

    st.divider()

    with st.container(border=True):
        st.markdown("## 📊 建设项目前期工作咨询费 — 服务类型选择")

        st.caption(
            "前期工作咨询费（建设项目前期工作咨询费）包含 **4 项服务类型**。"
            "请勾选需要计算的服务类型，调整系数后点击确认。"
        )

        # ── 服务类型多选 ──
        st.markdown("### 📋 选择服务类型")
        all_services = ["编制项目建议书", "编制可研报告", "评估项目建议书", "评估可研报告"]
        selected = []
        svc_col1, svc_col2 = st.columns(2)
        for i, svc in enumerate(all_services):
            col = svc_col1 if i < 2 else svc_col2
            with col:
                checked = st.checkbox(svc, value=True, key=f"keyan_svc_{i}")
                if checked:
                    selected.append(svc)

        if not selected:
            st.warning("⚠️ 请至少选择一项服务类型")
        else:
            st.markdown("---")

            # ── 系数调整 ──
            st.markdown("### 🎛️ 系数调整")

            # 行业调整系数（按系数分组）
            ky_ind_groups = [
                ("石化、化工、钢铁", 1.3),
                ("石油、天然气、水利、水电、交通（水运）、化纤", 1.2),
                ("有色、黄金、纺织、轻工、邮电、广播电视、医药、煤炭、"
                 "火电（含核电）、机械（含船舶、航空、航天、兵器）", 1.0),
                ("林业、商业、粮食、建筑", 0.8),
                ("建材、交通（公路）、铁道、市政公用工程", 0.7),
            ]
            ky_ind_labels = [f"{label}（{val}）" for label, val in ky_ind_groups]
            ky_ind_values = [val for _, val in ky_ind_groups]
            try:
                cur_ind_idx = ky_ind_values.index(ind_coef)
            except ValueError:
                cur_ind_idx = 2  # 默认 1.0
            ind_label = st.selectbox(
                "行业调整系数",
                range(len(ky_ind_labels)),
                index=cur_ind_idx,
                format_func=lambda idx: ky_ind_labels[idx],
                key="keyan_ind_coef",
            )
            chosen_ind_coef = ky_ind_values[ind_label]
            chosen_ind_name = ky_ind_groups[ind_label][0]

            # 复杂程度系数
            comp_options = [("简单", 0.8), ("一般", 1.0), ("复杂", 1.2)]
            comp_labels = [f"{label}（{val}）" for label, val in comp_options]
            comp_values = [val for _, val in comp_options]
            try:
                cur_comp_idx = comp_values.index(comp_coef)
            except ValueError:
                cur_comp_idx = 1
            comp_label = st.selectbox(
                "复杂程度系数",
                range(len(comp_labels)),
                index=cur_comp_idx,
                format_func=lambda idx: comp_labels[idx],
                key="keyan_comp_coef",
            )
            chosen_comp_coef = comp_values[comp_label]

            st.markdown("---")

            # ── 实时费用预览 ──
            st.markdown("### 💡 费用预览")
            try:
                from fee_engine import calc_keyan_multi

                preview = calc_keyan_multi(
                    amount_yi,
                    selected,
                    industry_coef=chosen_ind_coef,
                    industry_name=chosen_ind_name,
                    complexity_coef=chosen_comp_coef,
                )
                detail_list = preview.get("明细", [])
                total_fee = preview.get("合计(万元)", 0)

                # 明细表
                detail_rows = []
                for d in detail_list:
                    detail_rows.append(
                        f"| **{d['服务类型']}** "
                        f"| {d.get('基准价(万元)', '-')} "
                        f"| {d.get('系数', '-')} "
                        f"| **{d['费用(万元)']}** |"
                    )
                st.markdown(
                    "| 服务类型 | 基准价（万元） | 调整系数 | 费用（万元） |\n"
                    "|----------|:--:|:--:|:--:|\n" + "\n".join(detail_rows)
                )

                st.markdown(f"### 💰 建设项目前期工作咨询费合计：**{total_fee} 万元**")

                # 分档计算详情
                with st.expander("📐 查看各项计算过程"):
                    for d in detail_list:
                        st.markdown(f"**{d['服务类型']}**")
                        steps = d.get("计算步骤", [])
                        if steps:
                            for s in steps:
                                st.caption(
                                    f"{s.get('步骤', '')}：{s.get('公式', '')} → {s.get('结果', '')}"
                                )
                        else:
                            st.caption(f"费用：{d['费用(万元)']} 万元")
                        st.markdown("---")

            except Exception as e:
                st.error(f"预览计算失败：{e}")
                total_fee = 0
                detail_list = []

            st.markdown("---")

            # ── 打折系数 ──
            st.markdown("### 💰 费用打折")
            discount_coef = st.number_input(
                "打折系数（1.0 = 不打折，0.8 = 打八折）",
                min_value=0.01, max_value=2.00,
                value=ctx.get("discount_coef", 1.0), step=0.05,
                format="%.2f",
                key="discount_keyan",
                help="输入打折系数调整最终费用。",
            )
            ctx["discount_coef"] = discount_coef
            discounted_total = round(total_fee * discount_coef, 4)

            if abs(discount_coef - 1.0) < 0.005:
                st.info(f"**不打折**，最终费用：**{discounted_total} 万元**")
            elif discount_coef < 1.0:
                st.warning(
                    f"打折系数 **{discount_coef:.2f}** → "
                    f"{total_fee:.2f} 万 × {discount_coef:.2f} = "
                    f"**{discounted_total} 万元**"
                    f"（节省 {round(total_fee - discounted_total, 4)} 万元）"
                )
            else:
                st.warning(
                    f"上浮系数 **{discount_coef:.2f}** → "
                    f"{total_fee:.2f} 万 × {discount_coef:.2f} = "
                    f"**{discounted_total} 万元**"
                    f"（增加 {round(discounted_total - total_fee, 4)} 万元）"
                )

            st.markdown("---")

            # ── 确认 / 取消 ──
            col_btn1, col_btn2 = st.columns(2)
            discount_text = ""
            if abs(discount_coef - 1.0) >= 0.005:
                discount_text = (
                    f"\n\n**打折系数**：{discount_coef:.2f}\n\n"
                    f"**打折后费用**：{discounted_total} 万元"
                    f"（{total_fee:.2f} 万 × {discount_coef:.2f}）"
                )

            with col_btn1:
                if st.button("✅ 确认结果", use_container_width=True, key="confirm_keyan"):
                    # 构建详情文本
                    detail_parts = []
                    for d in detail_list:
                        detail_parts.append(
                            f"- **{d['服务类型']}**：{d['费用(万元)']} 万元"
                        )

                    final_response = (
                        f"## 建设项目前期工作咨询费\n\n"
                        f"**依据**：{_basis_md_links('《建设项目前期工作咨询收费暂行规定》（计价格[1999]1283号）')}"
                        f"（计价格[1999]1283号）\n\n"
                        f"**参数**："
                        f"估算投资额 {amount_yi:.4f} 亿元（{amount_yi * 10000:.0f} 万元），"
                        f"行业「{chosen_ind_name}」系数 {chosen_ind_coef}，"
                        f"复杂程度系数 {chosen_comp_coef}\n\n"
                        f"### 服务类型明细\n\n"
                        + "\n".join(detail_parts) +
                        f"\n\n### 💰 合计：**{total_fee} 万元**"
                        f"{discount_text}"
                    )
                    st.session_state.messages.append({
                        "role": "assistant",
                        "content": final_response,
                    })
                    del st.session_state.pending_keyan
                    st.rerun()
            with col_btn2:
                if st.button("🗑 取消", use_container_width=True, key="cancel_keyan"):
                    del st.session_state.pending_keyan
                    st.rerun()

# ===== 水土保持补偿费交互式输入 =====

if "pending_shuibao_compensation" in st.session_state:
    ctx = st.session_state.pending_shuibao_compensation

    st.divider()

    with st.container(border=True):
        st.markdown("## 🏗️ 水土保持补偿费 — 参数输入")
        st.caption(
            "依据《天津市水土保持补偿费征收标准》"
            "（津发改价综〔2020〕351号 / 发改价格[2017]1186号）"
        )

        # ── 计算类型选择 ──
        st.markdown("### 📋 选择计征类型")
        calc_types = [
            ("general", "一般性生产建设项目 — 1.4 元/m²（按征占土地面积一次性计征）"),
            ("mining_construction", "矿产资源开采（建设期）— 1.4 元/m²（一次性）"),
            ("mining_oil_gas", "矿产资源开采（油气生产期）— 每口井 2000m² × 1.4 元/m²/年"),
            ("mining_other", "矿产资源开采（其他矿产）— 0.3 元/m³（按开采量）"),
            ("material_extraction", "取土/挖砂/采石/烧制砖瓦瓷石灰 — 0.3 元/m³"),
            ("waste_disposal", "排放废弃土石渣 — 0.3 元/m³"),
        ]
        type_labels = [label for _, label in calc_types]
        type_keys = [key for key, _ in calc_types]

        cur_type = ctx.get("calc_type", "general")
        try:
            cur_idx = type_keys.index(cur_type)
        except ValueError:
            cur_idx = 0

        chosen_idx = st.selectbox(
            "计征类型",
            range(len(type_labels)),
            index=cur_idx,
            format_func=lambda i: type_labels[i],
            key="shuibao_comp_type",
        )
        chosen_type = type_keys[chosen_idx]
        ctx["calc_type"] = chosen_type

        st.markdown("---")

        # ── 参数输入（根据类型动态显示）──
        st.markdown("### 📐 输入参数")

        land_m2 = 0.0
        well_cnt = 0
        add_wells = 0
        extract_vol = 0.0
        material_vol = 0.0
        waste_vol = 0.0

        if chosen_type in ("general", "mining_construction"):
            st.caption("请输入征占土地面积（支持亩、公顷、平方米）")
            unit_col, val_col = st.columns([1, 3])
            with unit_col:
                area_unit = st.selectbox(
                    "面积单位",
                    ["m²", "亩", "公顷"],
                    key="shuibao_area_unit",
                )
            with val_col:
                land_input = st.number_input(
                    f"征占土地面积（{area_unit}）",
                    min_value=0.0,
                    value=float(ctx.get("land_input", 0.0)),
                    step=1.0 if area_unit == "m²" else 0.1,
                    format="%.2f",
                    key="shuibao_land_input",
                )
            ctx["land_input"] = land_input
            if area_unit == "亩":
                land_m2 = round(land_input * 666.67, 2)
            elif area_unit == "公顷":
                land_m2 = land_input * 10000
            else:
                land_m2 = land_input
            if land_input > 0:
                st.caption(f"换算为 **{land_m2:,.0f} m²**")

        elif chosen_type == "mining_oil_gas":
            col1, col2 = st.columns(2)
            with col1:
                well_cnt = st.number_input(
                    "油气生产井数量（口）",
                    min_value=0,
                    value=int(ctx.get("well_cnt", 0)),
                    step=1,
                    key="shuibao_well_cnt",
                )
                ctx["well_cnt"] = well_cnt
            with col2:
                add_wells = st.number_input(
                    "丛式井增加井数（口）",
                    min_value=0,
                    value=int(ctx.get("add_wells", 0)),
                    step=1,
                    key="shuibao_add_wells",
                    help="每增加一口井，增加计征面积 400m²",
                )
                ctx["add_wells"] = add_wells
            if well_cnt > 0:
                total_area = well_cnt * 2000 + add_wells * 400
                st.caption(
                    f"计费面积 = {well_cnt}口 × 2000m²"
                    + (f" + {add_wells}口 × 400m²" if add_wells > 0 else "")
                    + f" = **{total_area:,} m²**"
                )

        elif chosen_type == "mining_other":
            extract_vol = st.number_input(
                "开采量（m³ / 采掘、采剥量）",
                min_value=0.0,
                value=float(ctx.get("extract_vol", 0.0)),
                step=100.0,
                format="%.0f",
                key="shuibao_extract_vol",
            )
            ctx["extract_vol"] = extract_vol

        elif chosen_type == "material_extraction":
            material_vol = st.number_input(
                "取土/挖砂/采石/烧制量（m³）",
                min_value=0.0,
                value=float(ctx.get("material_vol", 0.0)),
                step=100.0,
                format="%.0f",
                key="shuibao_material_vol",
            )
            ctx["material_vol"] = material_vol

        elif chosen_type == "waste_disposal":
            waste_vol = st.number_input(
                "排放废弃土石渣量（m³）",
                min_value=0.0,
                value=float(ctx.get("waste_vol", 0.0)),
                step=100.0,
                format="%.0f",
                key="shuibao_waste_vol",
            )
            ctx["waste_vol"] = waste_vol

        st.markdown("---")

        # ── 实时预览 ──
        st.markdown("### 💰 费用预览")
        can_calc = (
            (chosen_type in ("general", "mining_construction") and land_m2 > 0)
            or (chosen_type == "mining_oil_gas" and well_cnt > 0)
            or (chosen_type == "mining_other" and extract_vol > 0)
            or (chosen_type == "material_extraction" and material_vol > 0)
            or (chosen_type == "waste_disposal" and waste_vol > 0)
        )

        if can_calc:
            from fee_engine import calc_shuibao_compensation

            preview = calc_shuibao_compensation(
                calc_type=chosen_type,
                land_area_m2=land_m2,
                well_count=well_cnt,
                additional_wells=add_wells,
                extraction_volume_m3=extract_vol,
                material_volume_m3=material_vol,
                waste_volume_m3=waste_vol,
            )

            fee_yuan = preview.get("结果(元)", 0)
            fee_wan = preview.get("结果(万元)", 0)
            central = preview.get("中央收入(元)", 0)
            local = preview.get("地方收入(元)", 0)
            calc_formula = preview.get("计算公式", "")

            st.markdown(f"**计算公式**：{calc_formula}")
            st.markdown(f"**补偿费合计**：**{fee_yuan:,.2f} 元**（{fee_wan:.2f} 万元）")
            st.markdown(f"- 中央收入（10%）：{central:,.2f} 元")
            st.markdown(f"- 地方收入（90%）：{local:,.2f} 元")

            # 重复计征提醒
            if chosen_type in ("material_extraction", "waste_disposal"):
                st.info("⚠️ 对缴纳义务人已按前几种方式计征水土保持补偿费的，不再重复计征。")
            if chosen_type != "mining_oil_gas":
                st.caption("💡 水利水电工程建设项目，水库淹没区不在计征范围之内。")

            ctx["_preview"] = preview
        else:
            st.info("👆 请在上方输入参数后查看费用预览")

        st.markdown("---")

        # ── 确认/取消按钮 ──
        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            if st.button("✅ 确认结果", use_container_width=True, key="confirm_shuibao_comp",
                         disabled=not can_calc):
                preview = ctx.get("_preview", {})
                if preview:
                    result_text = (
                        f"## 水土保持补偿费\n\n"
                        f"**计征类型**：{preview.get('参数', {}).get('计算类型', '')}\n\n"
                        f"**计算过程**：\n"
                        f"- 计算公式：{preview.get('计算公式', '')}\n"
                        f"- 补偿费合计：**{preview.get('结果(元)', 0):,.2f} 元**"
                        f"（{preview.get('结果(万元)', 0):.4f} 万元）\n"
                        f"- 其中中央收入（10%）：{preview.get('中央收入(元)', 0):,.2f} 元\n"
                        f"- 地方收入（90%）：{preview.get('地方收入(元)', 0):,.2f} 元\n\n"
                        f"> 依据：{preview.get('依据', '')}"
                    )
                    st.session_state.messages.append({
                        "role": "assistant",
                        "content": result_text,
                    })
                    del st.session_state.pending_shuibao_compensation
                    st.rerun()
        with col_btn2:
            if st.button("🗑 取消", use_container_width=True, key="cancel_shuibao_comp"):
                del st.session_state.pending_shuibao_compensation
                st.rerun()

def _recalc_jiaoyi_with_deps(fee_result, jianli_fee_new, sheji_fee_new):
    """根据调整后的依赖费种值重新计算交易服务费各项金额。"""
    from fee_engine import calc_jiaoyi_fuwu as _calc_jiaoyi
    categories = fee_result.get("分项明细", [])
    jianan = None
    shebei = None
    for c in categories:
        if c["类别"] == "施工":
            jianan = c["基数(万元)"]
        elif c["类别"] == "设备":
            shebei = c["基数(万元)"]
    if jianan is not None:
        return _calc_jiaoyi(
            jianan=jianan, shebei=shebei,
            jianli_fee=jianli_fee_new, sheji_fee=sheji_fee_new,
        )
    elif categories:
        amt = categories[0]["基数(万元)"]
        return _calc_jiaoyi(amount_wan=amt)
    return {}


def _render_coef_select(label, options, current_val, key_prefix, help_text=""):
    """渲染一个系数选择器（selectbox + 自定义输入）。"""
    option_labels = [f"{lbl}（{val}）" for lbl, val in options]
    option_values = [val for _, val in options]

    # 找到当前值的索引
    try:
        idx = option_values.index(current_val)
    except ValueError:
        idx = len(option_values)  # 指向"自定义"

    # 添加"自定义"选项
    option_labels.append("✏️ 自定义…")
    option_values.append(-1.0)

    selected_idx = st.selectbox(
        label,
        range(len(option_labels)),
        index=min(idx, len(option_labels) - 1),
        format_func=lambda i, labels=option_labels: labels[i],
        key=f"{key_prefix}_{label}",
        help=help_text,
    )

    chosen_val = option_values[selected_idx]
    if chosen_val == -1.0:
        chosen_val = st.number_input(
            f"自定义{label}的值",
            min_value=0.10, max_value=5.00,
            value=float(current_val) if current_val > 0.1 else 1.0,
            step=0.05, format="%.2f",
            key=f"{key_prefix}_{label}_cust",
        )
    return float(chosen_val)


# ===== 交易服务费计费方选择 =====

if "pending_jiaoyi_party" in st.session_state:
    ctx = st.session_state.pending_jiaoyi_party
    fee_result = ctx["fee_result"]

    st.divider()

    with st.container(border=True):
        st.markdown("## 🏛️ 交易服务费 — 计费方选择")
        st.caption(
            "依据津发改价管[2017]979号，交易服务费由**招标方承担 60%**、"
            "**中标方承担 40%**。请选择您的身份。"
        )

        # 显示费用明细
        total_yuan = fee_result.get("合计(元)", fee_result.get("结果(元)", 0))
        zb_yuan = fee_result.get("招标方(元)", round(total_yuan * 0.6, 2))
        zb_wan = round(zb_yuan / 10000.0, 4)
        zhongb_yuan = fee_result.get("中标方(元)", round(total_yuan * 0.4, 2))
        zhongb_wan = round(zhongb_yuan / 10000.0, 4)

        # 依赖费种展示（监理费、设计费作为计算基数）
        deps = fee_result.get("依赖费种", {})
        # 初始化 per-fee 系数、自定义金额和打折状态
        ctx.setdefault("dep_coefs", {})
        ctx.setdefault("dep_custom_amounts", {})
        ctx.setdefault("dep_discounts", {})
        dep_coefs = ctx["dep_coefs"]
        dep_custom_amounts = ctx["dep_custom_amounts"]
        dep_discounts = ctx["dep_discounts"]

        # 从 fee_result 提取建安/设备费用于重算监理费
        categories = fee_result.get("分项明细", [])
        _jianan = None
        _shebei = None
        for c in categories:
            if c["类别"] == "施工":
                _jianan = c["基数(万元)"]
            elif c["类别"] == "设备":
                _shebei = c["基数(万元)"]

        if deps:
            st.markdown("### 📎 依赖费种")

            # ========== 监理费 ==========
            jl_params = deps.get("监理费_参数", {})
            jl_custom_key = "监理费"
            # 初始化监理费系数
            jl_coefs = dep_coefs.setdefault("监理费", {
                "professional_coef": jl_params.get("专业调整系数", 1.0),
                "complexity_coef": jl_params.get("复杂程度系数", 1.0),
                "elevation_coef": jl_params.get("高程调整系数", 1.0),
            })

            with st.expander("#### 监理费 — 系数调整", expanded=True):
                # 系数选择器
                jl_prof = _render_coef_select(
                    "专业调整系数", JIANLI_PROFESSIONAL_OPTIONS,
                    jl_coefs["professional_coef"],
                    key_prefix="jiaoyi_dep_jl",
                    help_text="发改价格[2007]670号 附表三",
                )
                jl_comp = _render_coef_select(
                    "复杂程度系数", JIANLI_COMPLEXITY_OPTIONS,
                    jl_coefs["complexity_coef"],
                    key_prefix="jiaoyi_dep_jl_comp",
                    help_text="发改价格[2007]670号 1.0.9条",
                )
                jl_elev = _render_coef_select(
                    "高程调整系数", JIANLI_ELEVATION_OPTIONS,
                    jl_coefs["elevation_coef"],
                    key_prefix="jiaoyi_dep_jl_elev",
                    help_text="发改价格[2007]670号 1.0.9条",
                )
                jl_coefs["professional_coef"] = jl_prof
                jl_coefs["complexity_coef"] = jl_comp
                jl_coefs["elevation_coef"] = jl_elev

                # 用当前系数重算监理费
                try:
                    if _jianan is not None:
                        jl_r = calc_jianli(
                            jianan=_jianan, shebei=_shebei,
                            professional_coef=jl_prof,
                            complexity_coef=jl_comp,
                            elevation_coef=jl_elev,
                        )
                    else:
                        jl_r = calc_jianli(
                            amount_wan=deps.get("监理费(万元)", 0),
                            professional_coef=jl_prof,
                            complexity_coef=jl_comp,
                            elevation_coef=jl_elev,
                        )
                    jl_val = jl_r["结果(万元)"]
                except Exception:
                    jl_val = deps.get("监理费(万元)", 0)

                col_a, col_b = st.columns([3, 1])
                with col_a:
                    use_custom_jl = st.checkbox("使用自定义金额", key="jiaoyi_dep_cust_jl",
                                                value=jl_custom_key in dep_custom_amounts)
                    if use_custom_jl:
                        custom_jl = st.number_input(
                            "自定义金额（万元）", min_value=0.0, step=0.1,
                            value=float(dep_custom_amounts.get(jl_custom_key, jl_val)),
                            format="%.4f", key="jiaoyi_dep_cust_jl_val",
                        )
                        dep_custom_amounts[jl_custom_key] = custom_jl
                        jl_effective = custom_jl
                    else:
                        dep_custom_amounts.pop(jl_custom_key, None)
                        jl_effective = jl_val
                    jl_disc = st.number_input(
                        "打折系数", min_value=0.01, max_value=2.00,
                        value=float(dep_discounts.get(jl_custom_key, 1.0)),
                        step=0.05, format="%.2f", key="jiaoyi_dep_disc_jl",
                    )
                    dep_discounts[jl_custom_key] = jl_disc
                with col_b:
                    jl_final = round(jl_effective * jl_disc, 4)
                    st.metric("监理费", f"{jl_final} 万元")
                    if abs(jl_disc - 1.0) >= 0.005:
                        st.caption(f"= {jl_effective:.2f} × {jl_disc:.2f}")
                    elif use_custom_jl:
                        st.caption(f"= 自定义 {jl_effective} 万元")
                    else:
                        st.caption(f"= {jl_val:.2f} 万元")

            # ========== 设计费 ==========
            sj_params = deps.get("设计费_参数", {})
            sj_custom_key = "设计费"
            sj_coefs = dep_coefs.setdefault("设计费", {
                "professional_coef": sj_params.get("专业调整系数", 1.0),
                "complexity_coef": sj_params.get("复杂程度系数", 1.0),
                "additional_coef": sj_params.get("附加调整系数", 0.0),
            })

            with st.expander("#### 设计费 — 系数调整", expanded=True):
                sj_prof = _render_coef_select(
                    "专业调整系数", SHEJI_PROFESSIONAL_OPTIONS,
                    sj_coefs["professional_coef"],
                    key_prefix="jiaoyi_dep_sj",
                    help_text="计价格[2002]10号 附表二",
                )
                sj_comp = _render_coef_select(
                    "复杂程度系数", SHEJI_COMPLEXITY_OPTIONS,
                    sj_coefs["complexity_coef"],
                    key_prefix="jiaoyi_dep_sj_comp",
                    help_text="计价格[2002]10号 1.0.9.2",
                )
                sj_addi = st.number_input(
                    "附加调整系数",
                    min_value=0.10, max_value=5.00,
                    value=float(sj_coefs.get("additional_coef", 1.0)),
                    step=0.05, format="%.2f",
                    key="jiaoyi_dep_sj_addi",
                    help="计价格[2002]10号 1.0.9.3（多个系数合并 = 相加 − 个数 + 1）",
                )
                sj_coefs["professional_coef"] = sj_prof
                sj_coefs["complexity_coef"] = sj_comp
                sj_coefs["additional_coef"] = sj_addi

                # 用当前系数重算设计费
                try:
                    if _jianan is not None:
                        sj_r = calc_sheji(
                            _jianan + (_shebei or 0),
                            professional_coef=sj_prof,
                            complexity_coef=sj_comp,
                            additional_coefs=[sj_addi] if abs(sj_addi - 1.0) >= 0.005 else None,
                        )
                    else:
                        sj_r = calc_sheji(
                            deps.get("设计费(万元)", 0),
                            professional_coef=sj_prof,
                            complexity_coef=sj_comp,
                            additional_coefs=[sj_addi] if abs(sj_addi - 1.0) >= 0.005 else None,
                        )
                    sj_val = sj_r["结果(万元)"]
                except Exception:
                    sj_val = deps.get("设计费(万元)", 0)

                col_a, col_b = st.columns([3, 1])
                with col_a:
                    use_custom_sj = st.checkbox("使用自定义金额", key="jiaoyi_dep_cust_sj",
                                                value=sj_custom_key in dep_custom_amounts)
                    if use_custom_sj:
                        custom_sj = st.number_input(
                            "自定义金额（万元）", min_value=0.0, step=0.1,
                            value=float(dep_custom_amounts.get(sj_custom_key, sj_val)),
                            format="%.4f", key="jiaoyi_dep_cust_sj_val",
                        )
                        dep_custom_amounts[sj_custom_key] = custom_sj
                        sj_effective = custom_sj
                    else:
                        dep_custom_amounts.pop(sj_custom_key, None)
                        sj_effective = sj_val
                    sj_disc = st.number_input(
                        "打折系数", min_value=0.01, max_value=2.00,
                        value=float(dep_discounts.get(sj_custom_key, 1.0)),
                        step=0.05, format="%.2f", key="jiaoyi_dep_disc_sj",
                    )
                    dep_discounts[sj_custom_key] = sj_disc
                with col_b:
                    sj_final = round(sj_effective * sj_disc, 4)
                    st.metric("设计费", f"{sj_final} 万元")
                    if abs(sj_disc - 1.0) >= 0.005:
                        st.caption(f"= {sj_effective:.2f} × {sj_disc:.2f}")
                    elif use_custom_sj:
                        st.caption(f"= 自定义 {sj_effective} 万元")
                    else:
                        st.caption(f"= {sj_val:.2f} 万元")

            st.markdown("---")

            # 根据依赖费种调整值重新计算交易服务费
            _recalc_jiaoyi = _recalc_jiaoyi_with_deps(
                fee_result, jl_final, sj_final
            )
            total_yuan = _recalc_jiaoyi.get("合计(元)", fee_result.get("合计(元)", total_yuan))
            zb_yuan = _recalc_jiaoyi.get("招标方(元)", round(total_yuan * 0.6, 2))
            zb_wan = round(zb_yuan / 10000.0, 4)
            zhongb_yuan = _recalc_jiaoyi.get("中标方(元)", round(total_yuan * 0.4, 2))
            zhongb_wan = round(zhongb_yuan / 10000.0, 4)
            # 更新 fee_result 供确认按钮使用
            fee_result["_adjusted_total_yuan"] = total_yuan
            fee_result["_adjusted_zb_yuan"] = zb_yuan
            fee_result["_adjusted_zhongb_yuan"] = zhongb_yuan
            fee_result["_adjusted_dep_fees"] = {"监理费": jl_final, "设计费": sj_final}
            # 更新依赖费种参数（供确认响应使用）
            fee_result["_dep_coefs"] = dict(dep_coefs)
            fee_result["_dep_discounts"] = dict(dep_discounts)
            fee_result["_dep_custom_amounts"] = dict(dep_custom_amounts)

        # 费用预览卡片
        col_zb, col_zhongb = st.columns(2)
        with col_zb:
            st.markdown(
                f"""<div style="
                    background: linear-gradient(135deg, #43e97b 0%, #38f9d7 100%);
                    border-radius: 12px; padding: 16px 20px; color: #333;
                ">
                    <div style="font-size: 0.85rem; opacity: 0.7;">🏢 招标方（60%）</div>
                    <div style="font-size: 1.6rem; font-weight: 700;">{zb_yuan:,.0f} 元</div>
                    <div style="font-size: 0.85rem; opacity: 0.7;">{zb_wan:.4f} 万元</div>
                </div>""",
                unsafe_allow_html=True,
            )
        with col_zhongb:
            st.markdown(
                f"""<div style="
                    background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
                    border-radius: 12px; padding: 16px 20px; color: white;
                ">
                    <div style="font-size: 0.85rem; opacity: 0.85;">🎯 中标方（40%）</div>
                    <div style="font-size: 1.6rem; font-weight: 700;">{zhongb_yuan:,.0f} 元</div>
                    <div style="font-size: 0.85rem; opacity: 0.85;">{zhongb_wan:.4f} 万元</div>
                </div>""",
                unsafe_allow_html=True,
            )

        st.markdown("---")

        # 计费方选择
        party = st.radio(
            "请选择您的身份",
            ["招标方", "中标方"],
            format_func=lambda x: f"{x}（{'60%' if x == '招标方' else '40%'}）",
            key="jiaoyi_party_radio",
        )

        st.markdown("---")

        # 确认按钮
        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            if st.button("✅ 确认结果", use_container_width=True, key="confirm_jiaoyi_party"):
                chosen_fee = zb_yuan if party == "招标方" else zhongb_yuan
                chosen_fee_wan = zb_wan if party == "招标方" else zhongb_wan
                party_pct = "60%" if party == "招标方" else "40%"

                # 重新调用 calc_jiaoyi_fuwu 以获得包含 party 的结果描述
                from fee_engine import calc_jiaoyi_fuwu as _calc_jiaoyi
                # 从原始结果中提取参数
                categories = fee_result.get("分项明细", [])
                jianan = None
                shebei = None
                jianli_fee_orig = None
                sheji_fee_orig = None
                for c in categories:
                    if c["类别"] == "施工":
                        jianan = c["基数(万元)"]
                    elif c["类别"] == "设备":
                        shebei = c["基数(万元)"]
                    elif c["类别"] == "监理":
                        jianli_fee_orig = c["基数(万元)"]
                    elif c["类别"] == "设计":
                        sheji_fee_orig = c["基数(万元)"]
                # 使用调整后的依赖费种值（如果用户自定义了）
                adjusted_deps = fee_result.get("_adjusted_dep_fees", {})
                if adjusted_deps:
                    jianli_fee_for_calc = adjusted_deps.get("监理费", jianli_fee_orig or 0)
                    sheji_fee_for_calc = adjusted_deps.get("设计费", sheji_fee_orig or 0)
                else:
                    jianli_fee_for_calc = jianli_fee_orig
                    sheji_fee_for_calc = sheji_fee_orig

                if jianan is not None:
                    # 重算以获得 party 特定的结果
                    party_result = _calc_jiaoyi(
                        jianan=jianan, shebei=shebei,
                        jianli_fee=jianli_fee_for_calc, sheji_fee=sheji_fee_for_calc,
                        party=party,
                    )
                else:
                    amt = categories[0]["基数(万元)"] if categories else 0
                    party_result = _calc_jiaoyi(amount_wan=amt, party=party)

                # 使用调整后的金额（如果存在）
                adjusted_zb = fee_result.get("_adjusted_zb_yuan")
                adjusted_zhongb = fee_result.get("_adjusted_zhongb_yuan")
                if adjusted_zb is not None:
                    chosen_fee = adjusted_zb if party == "招标方" else adjusted_zhongb
                    chosen_fee_wan = round(chosen_fee / 10000.0, 4)

                # 依赖费种详细文本（使用用户实际选择的系数/折扣/自定义）
                user_dep_coefs = fee_result.get("_dep_coefs", {})
                user_dep_discs = fee_result.get("_dep_discounts", {})
                user_dep_custs = fee_result.get("_dep_custom_amounts", {})
                deps_info = fee_result.get("依赖费种", {})
                dep_text = ""
                if adjusted_deps:
                    dep_lines = []
                    # 监理费
                    jl_final_val = adjusted_deps.get("监理费", 0)
                    jl_coefs = user_dep_coefs.get("监理费", {})
                    jl_cust = user_dep_custs.get("监理费")
                    jl_disc = user_dep_discs.get("监理费", 1.0)
                    jl_line = f"- 监理费：**{jl_final_val} 万元**"
                    jl_notes = []
                    if jl_cust is not None:
                        jl_notes.append(f"自定义={jl_cust}")
                    elif jl_coefs:
                        for cn, ckey in [("专业系数", "professional_coef"), ("复杂系数", "complexity_coef"), ("高程系数", "elevation_coef")]:
                            cv = jl_coefs.get(ckey, 1.0)
                            if cv is not None and abs(cv - 1.0) >= 0.005:
                                jl_notes.append(f"{cn}={cv}")
                    if abs(jl_disc - 1.0) >= 0.005:
                        jl_notes.append(f"打折={jl_disc:.2f}")
                    if jl_notes:
                        jl_line += f"（{'，'.join(jl_notes)}）"
                    dep_lines.append(jl_line)
                    # 设计费
                    sj_final_val = adjusted_deps.get("设计费", 0)
                    sj_coefs = user_dep_coefs.get("设计费", {})
                    sj_cust = user_dep_custs.get("设计费")
                    sj_disc = user_dep_discs.get("设计费", 1.0)
                    sj_line = f"- 设计费：**{sj_final_val} 万元**"
                    sj_notes = []
                    if sj_cust is not None:
                        sj_notes.append(f"自定义={sj_cust}")
                    elif sj_coefs:
                        for cn, ckey in [("专业系数", "professional_coef"), ("复杂系数", "complexity_coef"), ("附加系数", "additional_coef")]:
                            cv = sj_coefs.get(ckey, 1.0)
                            if cv is not None and abs(cv) >= 0.005:
                                sj_notes.append(f"{cn}={cv}")
                    if abs(sj_disc - 1.0) >= 0.005:
                        sj_notes.append(f"打折={sj_disc:.2f}")
                    if sj_notes:
                        sj_line += f"（{'，'.join(sj_notes)}）"
                    dep_lines.append(sj_line)
                    dep_text = "### 📎 依赖费种（计算基数）\n\n" + "\n".join(dep_lines) + "\n\n"
                elif deps_info:
                    # 没有调整时使用原始值
                    dep_lines = []
                    if deps_info.get("监理费(万元)"):
                        dep_lines.append(f"- 监理费：**{deps_info['监理费(万元)']} 万元**")
                    if deps_info.get("设计费(万元)"):
                        dep_lines.append(f"- 设计费：**{deps_info['设计费(万元)']} 万元**")
                    if dep_lines:
                        dep_text = "### 📎 依赖费种（计算基数）\n\n" + "\n".join(dep_lines) + "\n\n"

                result_text = (
                    f"## 工程建设交易服务费\n\n"
                    f"{dep_text}"
                    f"**计费方**：{party}（{party_pct}）\n\n"
                    f"**应承担费用**：**{chosen_fee:,.0f} 元**"
                    f"（{chosen_fee_wan:.4f} 万元）\n\n"
                    f"> 分摊规则：招标方 60%（{zb_yuan:,.0f} 元），"
                    f"中标方 40%（{zhongb_yuan:,.0f} 元）\n\n"
                    f"> 依据：津发改价管[2017]979号"
                )
                st.session_state.messages.append({
                    "role": "assistant",
                    "content": result_text,
                })
                del st.session_state.pending_jiaoyi_party
                st.rerun()
        with col_btn2:
            if st.button("🗑 取消", use_container_width=True, key="cancel_jiaoyi_party"):
                del st.session_state.pending_jiaoyi_party
                st.rerun()

# ===== 全费用交互式选择 =====
if "pending_fee_selection" in st.session_state:
    ctx = st.session_state.pending_fee_selection

    # 安全检查：若最近一条用户消息并非当前 cascade 查询，说明用户已提新问题，
    # pending_fee_selection 应已被新 prompt 的清理逻辑移除。若未移除说明是残留状态。
    last_user_msgs = [m for m in st.session_state.messages if m.get("role") == "user"]
    if last_user_msgs:
        last_user_content = last_user_msgs[-1].get("content", "")
        if last_user_content != ctx.get("query", ""):
            del st.session_state.pending_fee_selection
            st.rerun()

    st.divider()

    with st.container(border=True):
        # ── Header ──
        st.markdown("## 📋 二类费选择 — 交互式联算")
        st.caption(
            f"计费基数：建安费 **{ctx['jianan']}** 万元 + 设备费 **{ctx['shebei']}** 万元 "
            f"= **{ctx['total_part1']}** 万元 ｜ "
            f"项目类型：**{ctx['project_type']}**"
        )

        # ── Fee Selection Checkboxes (grouped by tier) ──
        st.markdown("### 请选择需要计算的二类费")
        st.caption("勾选需要计算的费种，取消勾选将从最终汇总中移除该费种。")

        tier_labels = {0: "第一部分工程费相关", 1: "勘察设计费相关",
                       2: "总投资相关", 3: "预备费", 4: "独立费种（非投资额相关）"}
        tier_colors = {0: "#e8f5e9", 1: "#fff3e0", 2: "#e3f2fd", 3: "#fce4ec", 4: "#f3e5f5"}

        # 收集当前勾选状态
        new_selected: set[str] = set()
        prev_selected = set(ctx.get("selected_fees", set()))

        for tier in [0, 1, 2, 3, 4]:
            tier_fees = [fd for fd in ctx["fee_defs"] if fd["tier"] == tier]
            if not tier_fees:
                continue

            # Tier 标题（小字低调）
            _bg = tier_colors.get(tier, "#f5f5f5")
            _tl = tier_labels.get(tier, "")
            st.markdown(
                f"<span style='background:{_bg};padding:1px 6px;"
                f"border-radius:3px;font-size:0.8em;color:#666;'>{_tl}</span>",
                unsafe_allow_html=True,
            )

            cols = st.columns(3)
            for i, fd in enumerate(tier_fees):
                fee_name = fd["name"]
                checked = fee_name in prev_selected
                with cols[i % 3]:
                    new_checked = st.checkbox(
                        f"**{fd['label']}**",
                        value=checked,
                        key=f"fee_sel_{fee_name}",
                        help=f"默认值：{fd['default_value_wan']:.2f} 万元 | "
                             f"依据：{_FEE_LABELS.get(fee_name, fee_name)}",
                    )
                    if new_checked:
                        new_selected.add(fee_name)
                    if fd["default_value_wan"] > 0:
                        st.caption(f"≈ {fd['default_value_wan']:.2f} 万元")
                    else:
                        st.caption("需满足前提条件")

        # ── 依赖关系检查：自动取消依赖已取消费种的费种 ──
        removed = prev_selected - new_selected
        if removed:
            # 找出所有依赖被移除费种的费种
            auto_removed: set[str] = set()
            for fd in ctx["fee_defs"]:
                if fd["name"] not in new_selected:
                    continue
                deps = fd.get("depends_on", [])
                for dep in deps:
                    if dep in removed and fd["name"] in new_selected:
                        auto_removed.add(fd["name"])
                        break

            if auto_removed:
                new_selected -= auto_removed
                st.warning(
                    f"⚠️ 以下费种依赖已取消选择的费种，已自动取消勾选："
                    f"**{'、'.join(auto_removed)}**"
                )

        # ── 依赖警告：已选中但依赖未选中的费种 ──
        for fd in ctx["fee_defs"]:
            if fd["name"] not in new_selected:
                continue
            deps = fd.get("depends_on", [])
            missing = [d for d in deps if d not in new_selected and not d.startswith("__")]
            if missing:
                st.info(
                    f"ℹ️ **{fd['label']}** 依赖 **{'、'.join(missing)}**，"
                    f"但后者未选中。计算时该项费用可能为 0。"
                )

        # 持久化勾选状态
        ctx["selected_fees"] = new_selected
        st.session_state.pending_fee_selection["selected_fees"] = new_selected

        st.markdown("---")

        # ── Coefficient Controls (expandable, per fee) ──
        has_coef_fees = [
            fd for fd in ctx["fee_defs"]
            if fd["has_coefs"] and fd["name"] in new_selected
        ]
        if has_coef_fees:
            st.markdown("### 🎛️ 系数调整")
            st.caption("展开可调整对应费种的系数，实时影响预览结果。")

            for fd in has_coef_fees:
                fee_name = fd["name"]
                config = fd["coef_config"]
                if not config:
                    continue

                with st.expander(
                    f"{fd['label']} — 系数调整",
                    expanded=False,
                ):
                    overrides = ctx["coef_overrides"].get(fee_name, {})
                    for coef_def in config["coefs"]:
                        coef_key = coef_def["key"]
                        param_name = coef_def["param_name"]
                        options = coef_def["options"]
                        current_val = overrides.get(param_name, coef_def["current"])

                        if options:
                            # 有预设选项 → selectbox + 自定义
                            option_labels = [f"{label}（{val}）" for label, val in options]
                            option_values = [val for _, val in options]

                            # 找到当前值的索引
                            try:
                                idx = option_values.index(current_val)
                            except ValueError:
                                idx = len(option_values)  # 指向"自定义"

                            # 添加"自定义"选项
                            option_labels.append("✏️ 自定义…")
                            option_values.append(-1.0)

                            selected_idx = st.selectbox(
                                coef_key,
                                range(len(option_labels)),
                                index=min(idx, len(option_labels) - 1),
                                format_func=lambda i, labels=option_labels: labels[i],
                                key=f"cascade_coef_{fee_name}_{param_name}",
                                help=coef_def.get("description", ""),
                            )

                            chosen_val = option_values[selected_idx]
                            if chosen_val == -1.0:
                                # 自定义值
                                chosen_val = st.number_input(
                                    f"自定义{coef_key}的值",
                                    min_value=0.10, max_value=5.00,
                                    value=float(current_val) if current_val > 0.1 else 1.0,
                                    step=0.05, format="%.2f",
                                    key=f"cascade_coef_cust_{fee_name}_{param_name}",
                                )
                            overrides[param_name] = float(chosen_val)
                        else:
                            # 无预设选项 → 纯自定义输入（如附加调整系数）
                            st.caption(coef_def.get("description", ""))
                            custom_val = st.number_input(
                                coef_key,
                                min_value=0.10, max_value=5.00,
                                value=float(current_val),
                                step=0.05, format="%.2f",
                                key=f"cascade_coef_{fee_name}_{param_name}",
                                help=coef_def.get("description", ""),
                            )
                            overrides[param_name] = float(custom_val)

                    # ── 设计费附加项（计价格[2002]10号 1.0.16条）──
                    if fee_name == "工程设计费":
                        st.markdown("---")
                        st.caption(
                            "📐 **其他设计收费**（计价格[2002]10号 第1.0.16条）："
                        )
                        col_sg, col_jg = st.columns(2)
                        with col_sg:
                            sgys = st.checkbox(
                                "施工图预算编制费",
                                value=overrides.get("shigongtu_yusuan", False),
                                key=f"cascade_sgys_{fee_name}",
                                help="按基本设计收费的 10% 收取",
                            )
                            overrides["shigongtu_yusuan"] = sgys
                        with col_jg:
                            jgt = st.checkbox(
                                "竣工图编制费",
                                value=overrides.get("jungongtu", False),
                                key=f"cascade_jgt_{fee_name}",
                                help="按基本设计收费的 8% 收取",
                            )
                            overrides["jungongtu"] = jgt

                    ctx["coef_overrides"][fee_name] = overrides
                    st.session_state.pending_fee_selection["coef_overrides"] = ctx["coef_overrides"]

            st.markdown("---")

        # ── Service Selection (always visible, for ALL fees with has_services) ──
        svc_fees = [
            fd for fd in ctx["fee_defs"]
            if fd.get("has_services") and fd["name"] in new_selected
        ]
        if svc_fees:
            st.markdown("### 📋 服务类型选择")
            for fd in svc_fees:
                fee_name = fd["name"]
                service_config = fd["service_config"]
                if not service_config:
                    continue
                is_cost_consulting = (fee_name == "造价咨询费")
                is_huanping = (fee_name == "环境影响咨询费")
                is_keyan = (fee_name == "可行性研究费")

                # 各服务组之间用浅分隔线隔开
                st.markdown("---")
                if is_cost_consulting:
                    st.markdown("#### 🌿 工程造价咨询服务子项")
                    if is_hebei_region(st.session_state.get("selected_region")):
                        all_svcs = service_config.get("services_hebei", [])
                        default_svcs = service_config.get("default_selected_hebei", ["预算编制"])
                        st.caption("勾选需要计算的河北省造价咨询服务子项（冀建市研[2017]2号）。")
                    else:
                        all_svcs = service_config.get("services_tianjin", [])
                        default_svcs = service_config.get("default_selected_tianjin", ["编制施工图预算"])
                        st.caption("勾选需要计算的造价咨询服务子项（津价房地[2008]136号）。")
                elif is_huanping:
                    st.markdown("#### 🌍 环境影响咨询服务子项")
                    all_svcs = service_config["services"]
                    default_svcs = service_config.get("default_selected", ["编制报告书"])
                    st.caption("勾选需要计算的环境影响咨询服务类型（计价格[2002]125号）。")
                elif is_keyan:
                    st.markdown("#### 📊 建设项目前期工作咨询服务子项")
                    all_svcs = service_config["services"]
                    default_svcs = service_config.get("default_selected", ["编制可研报告"])
                    st.caption("勾选需要计算的建设项目前期工作咨询服务类型（计价格[1999]1283号）。")
                else:
                    all_svcs = service_config.get("services", [])
                    default_svcs = service_config.get("default_selected", [])
                    st.caption("勾选需要计算的服务类型。")

                prev_selected_svcs = ctx.get("service_selections", {}).get(fee_name, default_svcs)

                selected_svcs: list[str] = []
                n_items = len(all_svcs)
                n_cols = 3 if n_items >= 6 else (2 if n_items > 2 else 1)
                svc_cols = st.columns(n_cols)
                for i, svc_info in enumerate(all_svcs):
                    svc_name = svc_info["name"]
                    svc_label = svc_info["label"]
                    checked = svc_name in prev_selected_svcs
                    with svc_cols[i % n_cols]:
                        if st.checkbox(
                            svc_label,
                            value=checked,
                            key=f"cascade_svc_{fee_name}_{svc_name}",
                        ):
                            selected_svcs.append(svc_name)

                if not selected_svcs:
                    st.warning("⚠️ 请至少选择一项服务类型")
                elif is_cost_consulting:
                    try:
                        # 构建 param_overrides（与确认区一致的费率覆盖）
                        _svc_skip = set(fd["name"] for fd in ctx["fee_defs"]) - new_selected
                        _svc_param = {}
                        _rate_map = {
                            "劳动安全卫生评审费": "劳动安全卫生评审费费率",
                            "场地准备费及临时设施费": "场地准备费费率",
                            "工程保险费": "工程保险费费率",
                        }
                        for _fn, _rv in ctx.get("rate_overrides", {}).items():
                            _pk = _rate_map.get(_fn)
                            if _pk:
                                _svc_param[_pk] = float(_rv.replace("%", ""))
                        # 施工图审查费：使用复合键（如 "公建|中型|2.9"）
                        _shencha_rv2 = ctx.get("rate_overrides", {}).get("施工图审查费", "")
                        if _shencha_rv2 and "|" in str(_shencha_rv2):
                            _parts2 = str(_shencha_rv2).split("|")
                            if len(_parts2) >= 3:
                                _svc_param["施工图审查费项目类型"] = _parts2[0]
                                _svc_param["施工图审查费项目规模"] = _parts2[1]
                                _svc_param["施工图审查费费率"] = float(
                                    _parts2[2].replace("%", "").replace("元/m²", "").strip())
                                _shencha_area2 = ctx.get("_shencha_area")
                                if _shencha_area2:
                                    _svc_param["施工图审查费建筑面积"] = float(_shencha_area2)
                        _svc_raw = _calc_all_fees(
                            jianan=ctx["jianan"], shebei=ctx["shebei"],
                            project_type=ctx["project_type"], query=ctx["query"],
                            skip_fees=_svc_skip if _svc_skip else None,
                            coef_overrides=ctx.get("coef_overrides") or None,
                            param_overrides=_svc_param or None,
                            region=st.session_state.get("selected_region"),
                        )
                        _svc_approx_total = (
                            _svc_raw.get("项目总投资(万元)", 0)
                            + sum(cf["amount_wan"]
                                  for cf in ctx.get("custom_fees", []))
                        )
                        if is_hebei_region(st.session_state.get("selected_region")):
                            from fee_engine import calc_cost_consulting_multi_hebei
                            _cc_prof = (ctx.get("coef_overrides", {})
                                        .get("造价咨询费", {})
                                        .get("professional_coef", 1.0))
                            cc_preview = calc_cost_consulting_multi_hebei(
                                selected_svcs,
                                ctx["jianan"],
                                total_investment=_svc_approx_total if _svc_approx_total > 0 else None,
                                professional_coef=_cc_prof,
                                discount_coef=1.0,
                            )
                        else:
                            from fee_engine import calc_cost_consulting_multi
                            cc_preview = calc_cost_consulting_multi(
                                selected_svcs,
                                ctx["total_part1"],
                                jianan_only=ctx["jianan"],
                                total_investment=_svc_approx_total if _svc_approx_total > 0 else None,
                            )
                        cc_total = cc_preview.get("合计(万元)", 0)
                        cc_details = cc_preview.get("明细", [])
                        st.markdown("**预览**：")
                        # 表格化展示（两列紧凑布局）
                        cc_cols = st.columns(2)
                        for i, d in enumerate(cc_details):
                            with cc_cols[i % 2]:
                                st.markdown(
                                    f"**{d['服务类型']}**：{d['费用(万元)']} 万元"
                                )
                        st.caption(f"造价咨询费合计：**{cc_total:.2f}** 万元")
                    except Exception:
                        pass
                elif is_huanping:
                    from fee_engine import calc_huanping_multi
                    hp_coefs = ctx.get("coef_overrides", {}).get(fee_name, {})
                    hp_ind_coef = hp_coefs.get("industry_coef", 1.0)
                    hp_sens_coef = hp_coefs.get("sensitivity_coef", 1.0)
                    try:
                        # 环评费基数为项目总投资，从 cascade 引擎获取近似值
                        _hp_skip = set(fd["name"] for fd in ctx["fee_defs"]) - new_selected
                        _hp_raw = _calc_all_fees(
                            jianan=ctx["jianan"], shebei=ctx["shebei"],
                            project_type=ctx["project_type"], query=ctx["query"],
                            skip_fees=_hp_skip if _hp_skip else None,
                            coef_overrides=ctx.get("coef_overrides") or None,
                            region=st.session_state.get("selected_region"),
                        )
                        _hp_approx_total = (
                            _hp_raw.get("项目总投资(万元)", 0)
                            + sum(cf["amount_wan"]
                                  for cf in ctx.get("custom_fees", []))
                        )
                        hp_preview = calc_huanping_multi(
                            _hp_approx_total if _hp_approx_total > 0 else ctx["total_part1"],
                            selected_svcs,
                            industry_coef=hp_ind_coef,
                            sensitivity_coef=hp_sens_coef,
                        )
                        hp_total = hp_preview.get("合计(万元)", 0)
                        hp_details = hp_preview.get("明细", [])
                        st.markdown("**预览**：")
                        for d in hp_details:
                            st.markdown(
                                f"- {d['服务类型']}：**{d['结果(万元)']}** "
                                f"（中值 **{d['结果中值(万元)']}** 万元）"
                            )
                        st.caption(f"环评费合计：**{hp_total:.2f}** 万元")
                    except Exception:
                        pass
                elif is_keyan:
                    from fee_engine import calc_keyan_multi
                    _ky_coefs = ctx.get("coef_overrides", {}).get(fee_name, {})
                    _ky_ind_coef = _ky_coefs.get("industry_coef", 1.0)
                    _ky_comp_coef = _ky_coefs.get("complexity_coef", 1.0)
                    try:
                        # 可研费基数为总投资（亿元），用 cascade 引擎近似总投
                        _ky_skip = set(fd["name"] for fd in ctx["fee_defs"]) - new_selected
                        _ky_raw = _calc_all_fees(
                            jianan=ctx["jianan"], shebei=ctx["shebei"],
                            project_type=ctx["project_type"], query=ctx["query"],
                            skip_fees=_ky_skip if _ky_skip else None,
                            coef_overrides=ctx.get("coef_overrides") or None,
                            region=st.session_state.get("selected_region"),
                        )
                        _ky_total = _ky_raw.get("项目总投资(万元)", 0)
                        _ky_amount_yi = _ky_total / 10000.0
                        ky_preview = calc_keyan_multi(
                            _ky_amount_yi,
                            selected_svcs,
                            industry_coef=_ky_ind_coef,
                            complexity_coef=_ky_comp_coef,
                        )
                        ky_total = ky_preview.get("合计(万元)", 0)
                        ky_details = ky_preview.get("明细", [])
                        st.markdown("**预览**：")
                        for d in ky_details:
                            st.markdown(
                                f"- {d['服务类型']}：**{d['费用(万元)']}** 万元"
                            )
                        st.caption(f"可研费合计：**{ky_total:.2f}** 万元")
                    except Exception:
                        pass

                if "service_selections" not in ctx:
                    ctx["service_selections"] = {}
                ctx["service_selections"][fee_name] = selected_svcs
                st.session_state.pending_fee_selection["service_selections"] = ctx["service_selections"]

            st.markdown("---")

        # ── Rate Selection (per fee, expandable) ──
        has_rate_fees = [
            fd for fd in ctx["fee_defs"]
            if fd.get("has_rates") and fd["name"] in new_selected
        ]
        if has_rate_fees:
            st.markdown("### 📊 费率选择")
            st.caption("展开可选择对应费种的适用费率，实时影响预览结果。")

            rate_overrides = ctx.get("rate_overrides", {})

            for fd in has_rate_fees:
                fee_name = fd["name"]
                config = fd.get("rate_config")
                if not config:
                    continue

                rate_options = config["rate_options"]
                use_composite = config.get("use_composite_key", False)

                if use_composite:
                    # ── 施工图审查费：项目类型 → 规模（河北单费率 / 天津多费率）──
                    default_key = config.get("default_key", "")
                    current_key = rate_overrides.get(fee_name, default_key)

                    # 解析当前 key 获得 ptype/size
                    _cur_parts = str(current_key).split("|") if current_key else []
                    _cur_ptype = _cur_parts[0] if len(_cur_parts) >= 1 else "公建"
                    _cur_size = _cur_parts[1] if len(_cur_parts) >= 2 else "中型"

                    # 按项目类型分组
                    _size_map: dict[str, dict[str, dict]] = {}
                    for ro in rate_options:
                        _size_map.setdefault(ro["ptype"], {})[ro["size"]] = ro
                    # 从 rate_options 动态构建项目类型列表（支持河北/天津）
                    _all_ptypes = list(dict.fromkeys(ro["ptype"] for ro in rate_options))

                    with st.expander(
                        f"{fd['label']} — 费率选择",
                        expanded=True,
                    ):
                        st.markdown(
                            f"<small>📜 依据：{_basis_with_links(config.get('basis', ''))}</small>",
                            unsafe_allow_html=True,
                        )
                        _is_hebei_sel = _cur_ptype == "河北"
                        if _is_hebei_sel:
                            st.info("📍 河北省项目，依据发改价格〔2011〕534号，审查费率统一为 6.5%。")
                        else:
                            st.info("💡 请选择项目类型和规模，费率将自动匹配。")

                        # Step 1: 选择项目类型
                        _ptype_labels = {
                            "公建": "公建类（公共建筑）",
                            "工业": "工业类",
                            "市政": "市政类（道路/桥梁/管网等）",
                            "住宅": "住宅类（按建筑面积计费）",
                            "河北": "河北省（发改价格〔2011〕534号）",
                        }
                        try:
                            _ptype_idx = _all_ptypes.index(_cur_ptype)
                        except ValueError:
                            _ptype_idx = 0
                        sel_ptype = st.selectbox(
                            "项目类型",
                            _all_ptypes,
                            index=_ptype_idx,
                            format_func=lambda x: _ptype_labels.get(x, x),
                            key=f"cascade_shencha_ptype",
                        )

                        # Step 2: 选择规模（根据类型过滤；河北无规模区分）
                        _sizes = list(_size_map.get(sel_ptype, {}).keys())
                        _is_hebei_type = sel_ptype == "河北"
                        if _cur_ptype != sel_ptype:
                            _cur_size = _sizes[0] if _sizes else "中型"
                        try:
                            _size_idx = _sizes.index(_cur_size)
                        except ValueError:
                            _size_idx = 0
                        if not _is_hebei_type:
                            sel_size = st.selectbox(
                                "项目规模",
                                _sizes,
                                index=_size_idx,
                                key=f"cascade_shencha_size",
                            )
                        else:
                            sel_size = _sizes[_size_idx] if _sizes else "—"

                        # 构建选中项
                        sel_ro = _size_map.get(sel_ptype, {}).get(sel_size)
                        if sel_ro:
                            selected_key = sel_ro["key"]
                            selected_label = sel_ro["label"]
                            rate_overrides[fee_name] = selected_key
                            ctx["_shencha_sel"] = {
                                "ptype": sel_ro["ptype"],
                                "size": sel_ro["size"],
                                "rate": sel_ro["rate"],
                                "billing": sel_ro.get("billing", "rate"),
                            }

                            # 住宅类：显示建筑面积输入
                            if sel_ro.get("billing") == "area":
                                prev_area = ctx.get("_shencha_area") or 0.0
                                shencha_area = st.number_input(
                                    "建筑面积（m²）",
                                    min_value=0.0,
                                    value=float(prev_area) if prev_area else 0.0,
                                    step=100.0,
                                    format="%.0f",
                                    key="cascade_shencha_area",
                                    help="住宅类施工图审查费按建筑面积 × 单价计费。\n"
                                         "大型 1.9 元/m² / 中型 1.7 元/m² / 小型 1.3 元/m²",
                                )
                                ctx["_shencha_area"] = shencha_area
                                if shencha_area > 0:
                                    rate_val = sel_ro["rate"].replace(" 元/m²", "")
                                    est = shencha_area * float(rate_val)
                                    st.caption(
                                        f"预计审查费：{shencha_area:.0f} m² "
                                        f"× {rate_val} 元/m² = **{est:.2f} 元**"
                                        f"（{est/10000:.4f} 万元）"
                                    )

                            # 显示选中费率的卡片
                            st.markdown(
                                f"""<div style="
                                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                                    border-radius: 12px;
                                    padding: 16px 24px;
                                    margin: 8px 0;
                                    color: white;
                                ">
                                    <span style="font-size: 0.85rem; opacity: 0.85;">✅ {fee_name}</span><br>
                                    <span style="font-size: 1.1rem;">{selected_label}</span>
                                </div>""",
                                unsafe_allow_html=True,
                            )
                else:
                    # ── 原有逻辑：简单费率（费率值唯一）──
                    default_rate = config["default_rate"]
                    param_key = config["param_key"]

                    # 当前选中的费率
                    current_rate = rate_overrides.get(fee_name, default_rate)

                    # 构建费率标签列表
                    rate_values = [ro["rate"] for ro in rate_options]

                    # 找到当前选中费率的索引
                    try:
                        rate_idx = rate_values.index(current_rate)
                    except ValueError:
                        rate_idx = len(rate_values) // 2  # 默认中值

                    with st.expander(
                        f"{fd['label']} — 费率选择",
                        expanded=False,
                    ):
                        st.markdown(
                            f"<small>📜 依据：{_basis_with_links(config.get('basis', ''))}</small>",
                            unsafe_allow_html=True,
                        )
                        st.caption(
                            f"计费基数：{ctx['total_part1']:.0f} 万元"
                        )

                        selected_rate = st.radio(
                            "选择费率",
                            rate_values,
                            index=rate_idx,
                            format_func=lambda r, opts=rate_options: next(
                                (o.get("label") or f"{o['rate']}  →  {o['fee_wan']:.2f} 万元"
                                 for o in opts if o['rate'] == r), r),
                            key=f"cascade_rate_{fee_name}",
                            horizontal=False,
                            label_visibility="collapsed",
                        )

                        rate_overrides[fee_name] = selected_rate

                        # 显示选中费率的卡片
                        selected_opt = next(
                            (ro for ro in rate_options if ro["rate"] == selected_rate), None)
                        selected_fee = selected_opt["fee_wan"] if selected_opt else 0
                        selected_label = selected_opt.get("label", "") if selected_opt else ""

                        if selected_fee > 0:
                            fee_line = (
                                f'<span style="font-size: 1.1rem; opacity: 0.7;">→ 费用</span>'
                                f'<span style="font-size: 1.8rem; font-weight: 700;">{selected_fee:.2f} 万</span>'
                            )
                        elif selected_label:
                            fee_line = (
                                f'<span style="font-size: 1.1rem; opacity: 0.7;">→ </span>'
                                f'<span style="font-size: 1.2rem; font-weight: 700;">{selected_label}</span>'
                            )
                        else:
                            fee_line = ""

                        st.markdown(
                            f"""<div style="
                                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                                border-radius: 12px;
                                padding: 16px 24px;
                                margin: 8px 0;
                                color: white;
                            ">
                                <span style="font-size: 0.85rem; opacity: 0.85;">✅ {fee_name}</span><br>
                                <span style="font-size: 1.1rem;">费率</span>
                                <span style="font-size: 1.8rem; font-weight: 700;">{selected_rate}</span>
                                {fee_line}
                            </div>""",
                            unsafe_allow_html=True,
                        )

            ctx["rate_overrides"] = rate_overrides
            st.session_state.pending_fee_selection["rate_overrides"] = rate_overrides

            st.markdown("---")

        # ── 水土保持补偿费参数输入 ──
        if "水土保持补偿费" in new_selected:
            st.markdown("### 🏗️ 水土保持补偿费 — 参数输入")
            st.caption(
                "依据津发改价综〔2020〕351号，需输入物理参数。"
                "该费用独立于投资额计算，不参与层级汇总。"
            )

            sb_params = ctx.setdefault("shuibao_comp_params", {
                "calc_type": "general",
                "land_input": 0.0,
                "land_unit": "m²",
                "well_cnt": 0,
                "add_wells": 0,
                "extract_vol": 0.0,
                "material_vol": 0.0,
                "waste_vol": 0.0,
            })

            # 计算类型选择
            sb_type_labels = [
                "一般性生产建设项目 — 1.4 元/m²",
                "矿产资源开采（建设期）— 1.4 元/m²",
                "矿产资源开采（油气生产期）— 2000m²/井 × 1.4 元/m²",
                "矿产资源开采（其他矿产）— 0.3 元/m³",
                "取土/挖砂/采石/烧制 — 0.3 元/m³",
                "排放废弃土石渣 — 0.3 元/m³",
            ]
            sb_type_keys = [
                "general", "mining_construction", "mining_oil_gas",
                "mining_other", "material_extraction", "waste_disposal",
            ]
            cur_sb_type = sb_params.get("calc_type", "general")
            try:
                sb_type_idx = sb_type_keys.index(cur_sb_type)
            except ValueError:
                sb_type_idx = 0
            chosen_sb_type = st.selectbox(
                "计征类型",
                range(len(sb_type_labels)),
                index=sb_type_idx,
                format_func=lambda i: sb_type_labels[i],
                key="cascade_sb_type",
            )
            sb_params["calc_type"] = sb_type_keys[chosen_sb_type]
            chosen_sb_key = sb_type_keys[chosen_sb_type]

            # 参数输入（根据类型动态显示）
            land_m2 = 0.0
            if chosen_sb_key in ("general", "mining_construction"):
                uc1, uc2 = st.columns([1, 3])
                with uc1:
                    area_unit = st.selectbox(
                        "面积单位", ["m²", "亩", "公顷"],
                        key="cascade_sb_area_unit",
                        index=["m²", "亩", "公顷"].index(sb_params.get("land_unit", "m²")),
                    )
                    sb_params["land_unit"] = area_unit
                with uc2:
                    land_input = st.number_input(
                        f"征占土地面积（{area_unit}）",
                        min_value=0.0,
                        value=float(sb_params.get("land_input", 0.0)),
                        step=1.0 if area_unit == "m²" else 0.1,
                        format="%.2f",
                        key="cascade_sb_land",
                    )
                    sb_params["land_input"] = land_input
                if area_unit == "亩":
                    land_m2 = round(land_input * 666.67, 2)
                elif area_unit == "公顷":
                    land_m2 = land_input * 10000
                else:
                    land_m2 = land_input
                if land_input > 0:
                    st.caption(f"换算为 **{land_m2:,.0f} m²**")

            elif chosen_sb_key == "mining_oil_gas":
                wc1, wc2 = st.columns(2)
                with wc1:
                    well_cnt = st.number_input(
                        "油气生产井数量（口）", min_value=0,
                        value=int(sb_params.get("well_cnt", 0)),
                        step=1, key="cascade_sb_wells",
                    )
                    sb_params["well_cnt"] = well_cnt
                with wc2:
                    add_wells = st.number_input(
                        "丛式井增加井数（口）", min_value=0,
                        value=int(sb_params.get("add_wells", 0)),
                        step=1, key="cascade_sb_addwells",
                    )
                    sb_params["add_wells"] = add_wells

            elif chosen_sb_key == "mining_other":
                extract_vol = st.number_input(
                    "开采量（m³）", min_value=0.0,
                    value=float(sb_params.get("extract_vol", 0.0)),
                    step=100.0, format="%.0f", key="cascade_sb_extract",
                )
                sb_params["extract_vol"] = extract_vol

            elif chosen_sb_key == "material_extraction":
                material_vol = st.number_input(
                    "取土/挖砂/采石/烧制量（m³）", min_value=0.0,
                    value=float(sb_params.get("material_vol", 0.0)),
                    step=100.0, format="%.0f", key="cascade_sb_material",
                )
                sb_params["material_vol"] = material_vol

            elif chosen_sb_key == "waste_disposal":
                waste_vol = st.number_input(
                    "排放废弃土石渣量（m³）", min_value=0.0,
                    value=float(sb_params.get("waste_vol", 0.0)),
                    step=100.0, format="%.0f", key="cascade_sb_waste",
                )
                sb_params["waste_vol"] = waste_vol

            st.session_state.pending_fee_selection["shuibao_comp_params"] = sb_params

            st.markdown("---")

        # ── 交易服务费计费方选择 ──
        if "交易服务费" in new_selected:
            st.markdown("### 🏛️ 交易服务费 — 计费方")
            st.caption("依据津发改价管[2017]979号，招标方承担60%，中标方承担40%。")
            jiaoyi_party = ctx.setdefault("jiaoyi_party", None)
            party_options = ["合计（不区分）", "招标方（60%）", "中标方（40%）"]
            party_values = [None, "招标方", "中标方"]
            cur_party_idx = party_values.index(jiaoyi_party) if jiaoyi_party in party_values else 0
            chosen = st.radio(
                "请选择计费方",
                range(len(party_options)),
                index=cur_party_idx,
                format_func=lambda i: party_options[i],
                key="cascade_jiaoyi_party",
                horizontal=True,
            )
            ctx["jiaoyi_party"] = party_values[chosen]
            st.session_state.pending_fee_selection["jiaoyi_party"] = party_values[chosen]

            # 依赖费种展示（监理费、设计费作为计算基数）
            jianli_coefs = ctx.get("coef_overrides", {}).get("监理费", {})
            sheji_coefs = ctx.get("coef_overrides", {}).get("工程设计费", {})
            jianli_r = None
            sheji_r = None
            try:
                jianli_r = calc_jianli(
                    jianan=ctx["jianan"], shebei=ctx["shebei"],
                    professional_coef=jianli_coefs.get("professional_coef", 1.0),
                    complexity_coef=jianli_coefs.get("complexity_coef", 0.85),
                    elevation_coef=jianli_coefs.get("elevation_coef", 1.0),
                )
                jianli_val = jianli_r["结果(万元)"]
            except Exception:
                jianli_val = 0
            try:
                sheji_r = calc_sheji(
                    ctx["total_part1"],
                    professional_coef=sheji_coefs.get("professional_coef", 1.0),
                    complexity_coef=sheji_coefs.get("complexity_coef", 1.0),
                )
                sheji_val = sheji_r["结果(万元)"]
            except Exception:
                sheji_val = 0

            st.markdown("#### 📎 依赖费种（计算基数）")
            # 监理费
            jl_col1, jl_col2 = st.columns([2, 1])
            with jl_col1:
                jl_params = jianli_r.get("参数", {}) if jianli_r else {}
                coef_strs = []
                for cn, ckey in [("专业系数", "专业调整系数"), ("复杂系数", "复杂程度系数"), ("高程系数", "高程调整系数")]:
                    cv = jl_params.get(ckey, 1.0)
                    if cv is not None:
                        coef_strs.append(f"{cn}={cv}")
                st.caption("监理费" + ("：`" + "，".join(coef_strs) + "`" if coef_strs else ""))
            with jl_col2:
                st.metric(label="监理费", value=f"{jianli_val} 万元")

            # 设计费
            sj_col1, sj_col2 = st.columns([2, 1])
            with sj_col1:
                sj_params = sheji_r.get("参数", {}) if sheji_r else {}
                coef_strs = []
                for cn, ckey in [("专业系数", "专业调整系数"), ("复杂系数", "复杂程度系数"), ("附加系数", "附加调整系数")]:
                    cv = sj_params.get(ckey, 1.0)
                    if cv is not None and abs(cv) >= 0.005:
                        coef_strs.append(f"{cn}={cv}")
                st.caption("设计费" + ("：`" + "，".join(coef_strs) + "`" if coef_strs else ""))
            with sj_col2:
                st.metric(label="设计费", value=f"{sheji_val} 万元")

            st.markdown("---")

        # ── Custom Fee Input ──
        st.markdown("### ➕ 自定义费用")
        st.caption("添加需要在汇总中额外计算的费用（如检测费、评估费、拆迁费等）。")

        col_name, col_amount, col_btn = st.columns([3, 1.5, 1])
        with col_name:
            custom_name = st.text_input(
                "费用名称", key="cascade_custom_name",
                placeholder="如：旧桥检测费",
                label_visibility="collapsed",
            )
        with col_amount:
            custom_amount = st.number_input(
                "金额（万元）", min_value=0.0, step=0.1,
                format="%.2f", key="cascade_custom_amount",
                label_visibility="collapsed",
            )
        with col_btn:
            if st.button("➕ 添加", use_container_width=True, key="cascade_custom_add"):
                if custom_name.strip() and custom_amount > 0:
                    ctx["custom_fees"].append({
                        "name": custom_name.strip(),
                        "amount_wan": custom_amount,
                    })
                    st.session_state.pending_fee_selection["custom_fees"] = ctx["custom_fees"]
                    st.rerun()

        # 显示已添加的自定义费用
        if ctx["custom_fees"]:
            st.markdown("**已添加的自定义费用**：")
            for i, cf in enumerate(ctx["custom_fees"]):
                c1, c2 = st.columns([5, 1])
                with c1:
                    st.markdown(f"- **{cf['name']}**：{cf['amount_wan']:.2f} 万元")
                with c2:
                    if st.button("🗑", key=f"cascade_del_custom_{i}", help="删除此费用"):
                        ctx["custom_fees"].pop(i)
                        st.session_state.pending_fee_selection["custom_fees"] = ctx["custom_fees"]
                        st.rerun()

        st.markdown("---")

        # ── Contract Rate/Price Overrides ──
        st.markdown("### 📝 合同费率/合同价覆盖")
        st.caption(
            "输入合同约定的费率或一口价，将替代对应费种的标准计算结果。"
            "合同覆盖不影响其他费种的依赖计算。"
        )

        contract_overrides = ctx.setdefault("contract_overrides", {})

        # 可被覆盖的费种：已选中且非预备费
        overridable_fees = [
            fd for fd in ctx["fee_defs"]
            if fd["name"] in new_selected and fd["name"] != "预备费"
        ]
        overridable_names = [fd["name"] for fd in overridable_fees]
        overridable_labels = [f"{fd['label']}（{fd['name']}）" for fd in overridable_fees]

        if overridable_fees:
            col_fee, col_type = st.columns([2, 1.5])
            with col_fee:
                selected_fee_idx = st.selectbox(
                    "选择费种",
                    range(len(overridable_fees)),
                    format_func=lambda i: overridable_labels[i] if i < len(overridable_labels) else "",
                    key="contract_fee_select",
                    label_visibility="collapsed",
                )
                selected_fee_name = overridable_names[selected_fee_idx]
            with col_type:
                ov_type = st.radio(
                    "覆盖类型",
                    ["合同费率", "合同价"],
                    key="contract_type_radio",
                    horizontal=True,
                    label_visibility="collapsed",
                )

            if ov_type == "合同费率":
                col_rate, col_base = st.columns([1, 2])
                with col_rate:
                    ov_rate = st.number_input(
                        "费率(%)",
                        min_value=0.01, max_value=100.0,
                        value=2.0, step=0.01, format="%.2f",
                        key="contract_rate_input",
                    )
                with col_base:
                    base_options = ["工程费", "建安费", "项目总投资", "自定义金额", "选定费种"]
                    ov_base = st.selectbox(
                        "计费基数",
                        base_options,
                        key="contract_base_select",
                    )

                ov_base_custom = 0.0
                ov_base_fees: list[str] = []
                if ov_base == "自定义金额":
                    ov_base_custom = st.number_input(
                        "自定义基数（万元）",
                        min_value=0.0, step=1.0, format="%.2f",
                        value=float(ctx["total_part1"]),
                        key="contract_base_custom",
                    )
                elif ov_base == "选定费种":
                    # 可选的基数费种：已选中的费种（排除被覆盖的费种自身）
                    base_candidates = [
                        fd for fd in overridable_fees
                        if fd["name"] != selected_fee_name
                    ]
                    if base_candidates:
                        base_candidate_names = [fd["name"] for fd in base_candidates]
                        base_candidate_labels = [
                            f"{fd['label']}（{fd['name']}）" for fd in base_candidates
                        ]
                        ov_base_fees = st.multiselect(
                            "选择作为基数的费种",
                            options=base_candidate_names,
                            format_func=lambda n: next(
                                (l for l, fdn in zip(base_candidate_labels, base_candidate_names)
                                 if fdn == n), n),
                            key="contract_base_fees",
                            help="所选费种的标准计算值求和作为计费基数",
                        )
                    else:
                        st.caption("无可选的基数费种")
            else:
                ov_amount = st.number_input(
                    "合同价（万元）",
                    min_value=0.0, step=0.1, format="%.2f",
                    value=0.0,
                    key="contract_amount_input",
                )

            # ── 预览覆盖效果 ──
            if ov_type == "合同费率" and ov_rate > 0:
                # 计算实时基数
                _preview_base = 0.0
                _base_desc = ""
                if ov_base == "工程费":
                    _preview_base = ctx["total_part1"]
                    _base_desc = f"工程费({_preview_base:.0f}万元)"
                elif ov_base == "建安费":
                    _preview_base = ctx["jianan"]
                    _base_desc = f"建安费({_preview_base:.0f}万元)"
                elif ov_base == "项目总投资":
                    _prev = ctx.get("preview", {})
                    _raw = _prev.get("raw", {}) if _prev else {}
                    _preview_base = _raw.get("项目总投资(万元)", ctx["total_part1"])
                    _base_desc = f"项目总投资({_preview_base:.2f}万元)"
                elif ov_base == "自定义金额":
                    _preview_base = ov_base_custom
                    _base_desc = f"自定义({_preview_base:.2f}万元)"
                elif ov_base == "选定费种":
                    _prev = ctx.get("preview", {})
                    _prev_num = _prev.get("numerical", {}) if _prev else {}
                    _preview_base = sum(
                        _prev_num.get(f"{f}(万元)", 0) for f in ov_base_fees
                    )
                    _base_desc = f"选定费种({'、'.join(ov_base_fees) if ov_base_fees else '无'}，合计{_preview_base:.2f}万元)"
                _preview_val = round(_preview_base * ov_rate / 100.0, 4)
                st.caption(f"💡 预览：{ov_rate}% × {_base_desc} = **{_preview_val:.2f}** 万元")
            elif ov_type == "合同价" and ov_amount > 0:
                st.caption(f"💡 预览：合同价 **{ov_amount:.2f}** 万元")

            # ── 添加按钮 ──
            if st.button("➕ 添加合同覆盖", use_container_width=True,
                         key="contract_add_btn"):
                if ov_type == "合同费率":
                    if ov_base in ("工程费", "建安费", "项目总投资") or \
                       (ov_base == "自定义金额" and ov_base_custom > 0) or \
                       (ov_base == "选定费种" and ov_base_fees):
                        override_entry: dict = {
                            "type": "rate",
                            "rate": ov_rate,
                            "base": ov_base,
                        }
                        if ov_base == "自定义金额":
                            override_entry["base_custom"] = ov_base_custom
                        elif ov_base == "选定费种":
                            override_entry["base_fees"] = ov_base_fees
                        contract_overrides[selected_fee_name] = override_entry
                        ctx["contract_overrides"] = contract_overrides
                        st.rerun()
                    else:
                        st.warning("请完成计费基数配置")
                else:
                    if ov_amount > 0:
                        contract_overrides[selected_fee_name] = {
                            "type": "price",
                            "amount_wan": ov_amount,
                        }
                        ctx["contract_overrides"] = contract_overrides
                        st.rerun()
                    else:
                        st.warning("请输入合同价金额")

            # ── 已有的合同覆盖列表 ──
            if contract_overrides:
                st.markdown("---")
                st.markdown("**已添加的合同覆盖**：")
                preview = ctx.get("preview", {})
                prev_num = preview.get("numerical", {}) if preview else {}
                for ov_fn, ov_cfg in list(contract_overrides.items()):
                    ov_label = _FEE_LABELS.get(ov_fn, ov_fn)
                    col_info, col_del = st.columns([5, 1])
                    with col_info:
                        if ov_cfg.get("type") == "rate":
                            _rt = ov_cfg["rate"]
                            _bs = ov_cfg["base"]
                            # 实时计算基数
                            if _bs == "工程费":
                                _bv = ctx["total_part1"]
                            elif _bs == "建安费":
                                _bv = ctx["jianan"]
                            elif _bs == "项目总投资":
                                _raw = preview.get("raw", {}) if preview else {}
                                _bv = _raw.get("项目总投资(万元)", ctx["total_part1"])
                            elif _bs == "自定义金额":
                                _bv = ov_cfg.get("base_custom", 0)
                            elif _bs == "选定费种":
                                _bfs = ov_cfg.get("base_fees", [])
                                _bv = sum(prev_num.get(f"{f}(万元)", 0) for f in _bfs)
                            else:
                                _bv = ctx["total_part1"]
                            _val = round(_bv * _rt / 100.0, 4)
                            if _bs == "选定费种":
                                _bfs = ov_cfg.get("base_fees", [])
                                st.caption(
                                    f"**{ov_label}**：合同费率 {_rt}% × "
                                    f"选定费种({'、'.join(_bfs)}，合计{_bv:.2f}万元) = **{_val:.2f}** 万元"
                                )
                            else:
                                st.caption(
                                    f"**{ov_label}**：合同费率 {_rt}% × "
                                    f"{_bs}({_bv:.2f}万元) = **{_val:.2f}** 万元"
                                )
                        else:
                            st.caption(
                                f"**{ov_label}**：合同价 **{ov_cfg['amount_wan']:.2f}** 万元"
                            )
                    with col_del:
                        if st.button("🗑", key=f"contract_del_{ov_fn}",
                                     help=f"删除 {ov_label} 合同覆盖"):
                            contract_overrides.pop(ov_fn, None)
                            ctx["contract_overrides"] = contract_overrides
                            st.rerun()
        else:
            st.caption("请先勾选至少一个费种（预备费除外）")

        st.markdown("---")

        # ── 按费种独立打折 ──
        st.markdown("### 💰 费用打折")
        st.caption("为每个选中的费种单独设置打折系数（1.0 = 不打折），实时预览将反映折后金额。")

        discounts = ctx.setdefault("fee_discounts", {})

        # 优先复用预览区存储的上次计算结果（已含折扣），与预览完全一致
        _stored_num = ctx.get("_computed_numerical", {})
        if _stored_num:
            _disc_numerical = dict(_stored_num)
        else:
            # 首次加载：调引擎获取（参数与预览区一致）
            _disc_skip = set(fd["name"] for fd in ctx["fee_defs"]) - new_selected
            _disc_param = {}
            _disc_rate_map = {
                "勘察费": "勘察费费率", "劳动安全卫生评审费": "劳动安全卫生评审费费率",
                "场地准备费及临时设施费": "场地准备费费率", "工程保险费": "工程保险费费率",
            }
            for _fn, _rv in ctx.get("rate_overrides", {}).items():
                _pk = _disc_rate_map.get(_fn)
                if _pk:
                    _disc_param[_pk] = float(_rv.replace("%", ""))
            _shencha_rv = ctx.get("rate_overrides", {}).get("施工图审查费", "")
            if _shencha_rv and "|" in str(_shencha_rv):
                _parts = str(_shencha_rv).split("|")
                if len(_parts) >= 3:
                    _disc_param["施工图审查费费率"] = float(_parts[2].replace("%", "").replace("元/m²", ""))
                    _disc_param["施工图审查费项目类型"] = _parts[0]
                    _disc_param["施工图审查费项目规模"] = _parts[1]
                    _shencha_area = ctx.get("_shencha_area")
                    if _shencha_area:
                        _disc_param["施工图审查费建筑面积"] = float(_shencha_area)
            _yb_rate = st.session_state.get("cascade_yb_rate")
            if _yb_rate is not None:
                _disc_param["预备费率"] = float(_yb_rate)

            _disc_raw = {}
            try:
                _disc_raw = _calc_all_fees(
                    jianan=ctx["jianan"], shebei=ctx["shebei"],
                    project_type=ctx["project_type"], query=ctx["query"],
                    skip_fees=_disc_skip if _disc_skip else None,
                    coef_overrides=ctx.get("coef_overrides") or None,
                    param_overrides=_disc_param or None,
                    jiaoyi_party=ctx.get("jiaoyi_party"),
                    contract_overrides=ctx.get("contract_overrides") or None,
                    fee_discounts=ctx.get("fee_discounts") or None,
                    custom_fees=ctx.get("custom_fees") or None,
                    region=st.session_state.get("selected_region"),
                )
                _disc_numerical = _disc_raw["_数值"]
            except Exception:
                _disc_numerical = {}
            # 引擎不计算造价咨询费/水土保持补偿费，内联计算
            if "造价咨询费" in new_selected and "造价咨询费(万元)" not in _disc_numerical:
                try:
                    _cc_svcs = ctx.get("service_selections", {}).get("造价咨询费", [])
                    if _cc_svcs:
                        _cc_cascade_total = (
                            _disc_raw.get("项目总投资(万元)", 0) + sum(cf["amount_wan"] for cf in ctx.get("custom_fees", []))
                        )
                        if is_hebei_region(st.session_state.get("selected_region")):
                            from fee_engine import calc_cost_consulting_multi_hebei
                            _cc_prof = (ctx.get("coef_overrides", {}).get("造价咨询费", {}).get("professional_coef", 1.0))
                            _cc_multi = calc_cost_consulting_multi_hebei(
                                _cc_svcs, ctx["jianan"], total_investment=_cc_cascade_total if _cc_cascade_total > 0 else None,
                                professional_coef=_cc_prof, discount_coef=1.0)
                        else:
                            from fee_engine import calc_cost_consulting_multi
                            _cc_multi = calc_cost_consulting_multi(
                                _cc_svcs, ctx["total_part1"], jianan_only=ctx["jianan"],
                                total_investment=_cc_cascade_total if _cc_cascade_total > 0 else None)
                        _disc_numerical["造价咨询费(万元)"] = _cc_multi.get("合计(万元)", 0)
                except Exception:
                    pass
            if "水土保持补偿费" in new_selected and "水土保持补偿费(万元)" not in _disc_numerical:
                try:
                    _sb_params = ctx.get("shuibao_comp_params", {})
                    _sb_land_m2 = 0.0
                    _sb_unit = _sb_params.get("land_unit", "m²")
                    _sb_land_input = float(_sb_params.get("land_input", 0.0))
                    if _sb_unit == "亩":
                        _sb_land_m2 = round(_sb_land_input * 666.67, 2)
                    elif _sb_unit == "公顷":
                        _sb_land_m2 = _sb_land_input * 10000
                    else:
                        _sb_land_m2 = _sb_land_input
                    from fee_engine import calc_shuibao_compensation
                    _sb_result = calc_shuibao_compensation(
                        calc_type=_sb_params.get("calc_type", "general"), land_area_m2=_sb_land_m2,
                        well_count=int(_sb_params.get("well_cnt", 0)),
                        additional_wells=int(_sb_params.get("add_wells", 0)),
                        extraction_volume_m3=float(_sb_params.get("extract_vol", 0.0)),
                        material_volume_m3=float(_sb_params.get("material_vol", 0.0)),
                        waste_volume_m3=float(_sb_params.get("waste_vol", 0.0)))
                    _disc_numerical["水土保持补偿费(万元)"] = _sb_result.get("结果(万元)", 0)
                except Exception:
                    pass

        discounted_total = 0.0
        raw_total = 0.0

        selected_fee_list = sorted(
            [fd for fd in ctx["fee_defs"]
             if fd["name"] in new_selected and fd["name"] != "预备费"],
            key=lambda fd: (fd["tier"], fd["name"]),
        )

        for fd in selected_fee_list:
            fn = fd["name"]
            _stored_val = _disc_numerical.get(f"{fn}(万元)", 0)
            if _stored_val <= 0:
                continue
            is_contract = fn in ctx.get("contract_overrides", {})
            cur_discount = discounts.get(fn, 1.0)
            # 若使用存储值（已含预览区折扣），反除得到原始值显示
            if _stored_num and abs(cur_discount - 1.0) >= 0.005 and cur_discount > 0.001:
                raw_val = round(_stored_val / cur_discount, 4)
            else:
                raw_val = _stored_val
            raw_total += raw_val
            disc_col1, disc_col2 = st.columns([3, 1])
            with disc_col1:
                label_text = f"**{fd['label']}**（{raw_val:.2f} 万元）"
                if is_contract:
                    label_text += " 🔒"
                st.caption(label_text)
            with disc_col2:
                if is_contract:
                    st.caption("合同价不打折")
                    discounts[fn] = 1.0
                    new_discount = 1.0
                else:
                    new_discount = st.number_input(
                        f"打折系数",
                        min_value=0.01, max_value=2.00,
                        value=float(cur_discount), step=0.05,
                        format="%.2f",
                        key=f"discount_{fn}",
                        label_visibility="collapsed",
                    )
                    discounts[fn] = new_discount
            discounted_val = round(raw_val * new_discount, 4)
            discounted_total += discounted_val
            if abs(new_discount - 1.0) >= 0.005:
                st.caption(f"  → {discounted_val:.2f} 万元")

        # 自定义费用不打折
        custom_total = sum(cf["amount_wan"] for cf in ctx.get("custom_fees", []))
        discounted_total += custom_total
        raw_total += custom_total

        if abs(discounted_total - raw_total) > 0.005:
            st.markdown(f"**打折前二类费合计**：{raw_total:.2f} 万元")
            st.markdown(f"**打折后二类费合计**：**{discounted_total:.2f} 万元**")
            if discounted_total < raw_total:
                st.success(f"节省 **{round(raw_total - discounted_total, 2)}** 万元")
            else:
                st.warning(f"增加 **{round(discounted_total - raw_total, 2)}** 万元")
        else:
            st.info(f"二类费合计（未打折）：**{raw_total:.2f} 万元**")

        st.session_state.pending_fee_selection["fee_discounts"] = discounts

        st.markdown("---")

        # ── Real-time Preview ──
        st.markdown("### 💡 实时预览")

        try:
            # 构建 skip_fees
            all_fee_names = set(fd["name"] for fd in ctx["fee_defs"])
            skip_fees = all_fee_names - new_selected

            # 构建 param_overrides（系数覆盖 + 费率覆盖）
            param_overrides = {}
            # 将系数覆盖转换为 param_overrides 格式
            for fee_name, overrides in ctx["coef_overrides"].items():
                for k, v in overrides.items():
                    param_overrides[f"{fee_name}.{k}"] = v
            # 费率覆盖：fee_name → param_key
            rate_param_map = {
                "勘察费": "勘察费费率",
                "劳动安全卫生评审费": "劳动安全卫生评审费费率",
                "场地准备费及临时设施费": "场地准备费费率",
                "工程保险费": "工程保险费费率",
            }
            for fee_name, rate_val in ctx.get("rate_overrides", {}).items():
                pk = rate_param_map.get(fee_name)
                if pk:
                    # 费率格式如 "0.5%"，提取数字
                    rate_num = float(rate_val.replace("%", ""))
                    param_overrides[pk] = rate_num
            # 施工图审查费：使用复合键（如 "公建|中型|2.9"），需拆分传递
            _shencha_rv = ctx.get("rate_overrides", {}).get("施工图审查费", "")
            if _shencha_rv and "|" in str(_shencha_rv):
                _parts = str(_shencha_rv).split("|")
                if len(_parts) >= 3:
                    param_overrides["施工图审查费项目类型"] = _parts[0]
                    param_overrides["施工图审查费项目规模"] = _parts[1]
                    # 费率：去除 % 和 元/m² 后缀
                    _rate_raw = _parts[2].replace("%", "").replace("元/m²", "").strip()
                    param_overrides["施工图审查费费率"] = float(_rate_raw)
                    # 住宅类建筑面积
                    _shencha_area = ctx.get("_shencha_area")
                    if _shencha_area:
                        param_overrides["施工图审查费建筑面积"] = float(_shencha_area)
            # 预备费率：直接从 widget session key 读取（避免 lag）
            _yb_rate_widget = st.session_state.get("cascade_yb_rate")
            if _yb_rate_widget is not None:
                param_overrides["预备费率"] = float(_yb_rate_widget)

            # 调用引擎计算（仅选中费种 + 系数覆盖 + 费率覆盖）
            preview_raw = _calc_all_fees(
                jianan=ctx["jianan"],
                shebei=ctx["shebei"],
                project_type=ctx["project_type"],
                query=ctx["query"],
                skip_fees=skip_fees if skip_fees else None,
                coef_overrides=ctx.get("coef_overrides") or None,
                param_overrides=param_overrides or None,
                jiaoyi_party=ctx.get("jiaoyi_party"),
                contract_overrides=ctx.get("contract_overrides") or None,
                fee_discounts=ctx.get("fee_discounts") or None,
                custom_fees=ctx.get("custom_fees") or None,
                region=st.session_state.get("selected_region"),
            )

            numerical = preview_raw["_数值"]

            # ── 环评费多服务类型覆盖 ──
            if "环境影响咨询费" in new_selected and "环境影响咨询费" not in ctx.get("contract_overrides", {}):
                hp_svcs = ctx.get("service_selections", {}).get("环境影响咨询费", [])
                if hp_svcs and hp_svcs != ["编制报告书"]:
                    # 用户选择了非默认的服务类型组合，用 calc_huanping_multi 覆盖
                    from fee_engine import calc_huanping_multi
                    hp_coefs = ctx.get("coef_overrides", {}).get("环境影响咨询费", {})
                    hp_ind_coef = hp_coefs.get("industry_coef", 1.0)
                    hp_sens_coef = hp_coefs.get("sensitivity_coef", 1.0)
                    try:
                        # 环评费基数为项目总投资（引擎已含自定义费用）
                        _hp_base = preview_raw.get("项目总投资(万元)", 0)
                        hp_multi = calc_huanping_multi(
                            _hp_base if _hp_base > 0 else ctx["total_part1"],
                            hp_svcs,
                            industry_coef=hp_ind_coef,
                            sensitivity_coef=hp_sens_coef,
                        )
                        hp_total = hp_multi.get("合计(万元)", 0)
                        old_hp = numerical.get("环境影响咨询费(万元)", 0)
                        numerical["环境影响咨询费(万元)"] = hp_total
                        # 更新原始结果
                        preview_raw["原始结果"]["环境影响咨询费"] = hp_multi
                        # 重新计算 T2 小计
                        t2_keys = [k for k in ("建设管理费", "可行性研究费", "环境影响咨询费")
                                   if f"{k}(万元)" in numerical]
                        new_t2 = sum(numerical.get(f"{k}(万元)", 0) for k in t2_keys)
                        preview_raw["T2小计(万元)"] = round(new_t2, 4)
                        # 重新计算二类费合计（含自定义费用）
                        _custom_all = sum(cf["amount_wan"] for cf in ctx.get("custom_fees", []))
                        fee_total_raw = (
                            preview_raw.get("T0小计(万元)", 0)
                            + preview_raw.get("T1小计(万元)", 0)
                            + new_t2
                        )
                        preview_raw["二类费合计(万元)"] = round(fee_total_raw + _custom_all, 4)
                        # 重新计算总投资（含自定义费用）
                        preview_raw["总投资(万元)"] = round(
                            ctx["total_part1"] + fee_total_raw + _custom_all, 4)
                        preview_raw["项目总投资(万元)"] = round(
                            ctx["total_part1"] + fee_total_raw + _custom_all
                            + preview_raw.get("预备费小计(万元)", 0), 4)
                    except Exception:
                        pass  # 失败时保留原始值

            # ── 可行性研究费多服务类型计算 ──
            if "可行性研究费" in new_selected and "可行性研究费" not in ctx.get("contract_overrides", {}):
                ky_svcs = ctx.get("service_selections", {}).get("可行性研究费", [])
                if ky_svcs and ky_svcs != ["编制可研报告"]:
                    from fee_engine import calc_keyan_multi
                    _ky_coefs = ctx.get("coef_overrides", {}).get("可行性研究费", {})
                    _ky_ind_coef = _ky_coefs.get("industry_coef", 1.0)
                    _ky_comp_coef = _ky_coefs.get("complexity_coef", 1.0)
                    try:
                        _ky_total_wan = preview_raw.get("项目总投资(万元)", 0)
                        _ky_amount_yi = _ky_total_wan / 10000.0
                        ky_multi = calc_keyan_multi(
                            _ky_amount_yi,
                            ky_svcs,
                            industry_coef=_ky_ind_coef,
                            complexity_coef=_ky_comp_coef,
                        )
                        ky_total = ky_multi.get("合计(万元)", 0)
                        old_ky = numerical.get("可行性研究费(万元)", 0)
                        numerical["可行性研究费(万元)"] = ky_total
                        preview_raw["原始结果"]["可行性研究费"] = ky_multi
                        # 重新计算 T2 小计
                        t2_keys = [k for k in ("建设管理费", "可行性研究费", "环境影响咨询费")
                                   if f"{k}(万元)" in numerical]
                        new_t2 = sum(numerical.get(f"{k}(万元)", 0) for k in t2_keys)
                        preview_raw["T2小计(万元)"] = round(new_t2, 4)
                        # 重新计算二类费合计
                        fee_total_raw = (
                            preview_raw.get("T0小计(万元)", 0)
                            + preview_raw.get("T1小计(万元)", 0)
                            + new_t2
                        )
                        _custom_all = sum(cf["amount_wan"] for cf in ctx.get("custom_fees", []))
                        preview_raw["二类费合计(万元)"] = round(fee_total_raw + _custom_all, 4)
                        # 重新计算总投资（含自定义费用）
                        preview_raw["总投资(万元)"] = round(
                            ctx["total_part1"] + fee_total_raw + _custom_all, 4)
                        preview_raw["项目总投资(万元)"] = round(
                            ctx["total_part1"] + fee_total_raw + _custom_all
                            + preview_raw.get("预备费小计(万元)", 0), 4)
                    except Exception:
                        pass  # 失败时保留原始值

            # ── 水土保持补偿费（独立于 CC，需在迭代前计算以纳入建管费和预备费基数）──
            sb_fee_wan = 0.0
            if "水土保持补偿费" in new_selected:
                sb_params = ctx.get("shuibao_comp_params", {})
                sb_type = sb_params.get("calc_type", "general")
                try:
                    from fee_engine import calc_shuibao_compensation
                    _sb_land_m2 = 0.0
                    _sb_unit = sb_params.get("land_unit", "m²")
                    _sb_land_input = float(sb_params.get("land_input", 0.0))
                    if _sb_unit == "亩":
                        _sb_land_m2 = round(_sb_land_input * 666.67, 2)
                    elif _sb_unit == "公顷":
                        _sb_land_m2 = _sb_land_input * 10000
                    else:
                        _sb_land_m2 = _sb_land_input
                    sb_result = calc_shuibao_compensation(
                        calc_type=sb_type,
                        land_area_m2=_sb_land_m2,
                        well_count=int(sb_params.get("well_cnt", 0)),
                        additional_wells=int(sb_params.get("add_wells", 0)),
                        extraction_volume_m3=float(sb_params.get("extract_vol", 0.0)),
                        material_volume_m3=float(sb_params.get("material_vol", 0.0)),
                        waste_volume_m3=float(sb_params.get("waste_vol", 0.0)),
                    )
                    sb_fee_wan = sb_result.get("结果(万元)", 0)
                    numerical["水土保持补偿费(万元)"] = sb_fee_wan
                    preview_raw["原始结果"]["水土保持补偿费"] = sb_result
                except Exception:
                    pass

            # ── 造价咨询费多服务类型计算 ──
            if "造价咨询费" in new_selected and "造价咨询费" not in ctx.get("contract_overrides", {}):
                cc_svcs = ctx.get("service_selections", {}).get("造价咨询费", [])
                if cc_svcs:
                    try:
                        # 对于需要总投资的子项（概算审核/概算编制等），
                        # 使用 cascade 引擎已计算的项目总投资作为基数
                        # 自定义费用应计入项目总投资
                        _cascade_total = (
                            preview_raw.get("项目总投资(万元)", 0)
                            + sum(cf["amount_wan"]
                                  for cf in ctx.get("custom_fees", []))
                        )
                        if is_hebei_region(st.session_state.get("selected_region")):
                            from fee_engine import calc_cost_consulting_multi_hebei
                            _cc_prof = (ctx.get("coef_overrides", {})
                                        .get("造价咨询费", {})
                                        .get("professional_coef", 1.0))
                            cc_multi = calc_cost_consulting_multi_hebei(
                                cc_svcs,
                                ctx["jianan"],  # 河北规则基数为建安费（不含设备费）
                                total_investment=_cascade_total if _cascade_total > 0 else None,
                                professional_coef=_cc_prof,
                                discount_coef=1.0,
                            )
                        else:
                            from fee_engine import calc_cost_consulting_multi
                            cc_multi = calc_cost_consulting_multi(
                                cc_svcs,
                                ctx["total_part1"],
                                jianan_only=ctx["jianan"],
                                total_investment=_cascade_total if _cascade_total > 0 else None,
                            )
                        cc_total = cc_multi.get("合计(万元)", 0)
                        # 添加/覆盖到 numerical
                        old_cc = numerical.get("造价咨询费(万元)", 0)
                        numerical["造价咨询费(万元)"] = cc_total
                        # 更新原始结果
                        preview_raw["原始结果"]["造价咨询费"] = cc_multi
                        # 重新计算 T0 小计（造价咨询费是 Tier 0）
                        t0_keys = [
                            "监理费", "工程设计费", "勘察费",
                            "劳动安全卫生评审费", "场地准备费及临时设施费", "工程保险费",
                            "造价咨询费",
                        ]
                        new_t0 = sum(numerical.get(f"{k}(万元)", 0) for k in t0_keys)
                        preview_raw["T0小计(万元)"] = round(new_t0, 4)

                        # ── 迭代收敛：建设管理费 ↔ 造价咨询费(概算审核等) ──
                        # 建设管理费和概算审核都依赖项目总投资，形成循环依赖，
                        # 需要迭代至收敛（阈值 0.005 万元）
                        from fee_engine import (
                            calc_jianshe_guanli, _extract_numeric_value as _ext_num,
                            _detect_keyan_industry, calc_keyan, calc_keyan_multi,
                            JIANSHE_GUANLI_RATES,
                            _cumulative_tiered, _match_custom_fee_deductions,
                        )
                        t1_total = preview_raw.get("T1小计(万元)", 0)
                        # 自定义费用应计入项目总投资（影响建设管理费和概算审核基数）
                        _custom_total = sum(
                            cf["amount_wan"] for cf in ctx.get("custom_fees", []))
                        # 识别管线切改费/建设用地费，在建设管理费基数中扣除
                        _gl_deductions = _match_custom_fee_deductions(
                            ctx.get("custom_fees", []))
                        _gl_qg = _gl_deductions.get("管线切改费", 0.0)
                        _gl_js = _gl_deductions.get("建设用地费", 0.0)
                        # 预备费率：直接从 widget 读取
                        _yb_rate = float(st.session_state.get("cascade_yb_rate", 5.0))
                        _prev_total = 0.0
                        # 含自定义费用 + 水保费（均为二类费）
                        _curr_total = round(
                            ctx["total_part1"] + new_t0 + t1_total
                            + preview_raw.get("T2小计(万元)", 0) + _custom_total
                            + sb_fee_wan, 4)
                        # 先用当前 total 计算预备费，得到完整项目总投资
                        _yb = round((_curr_total) * _yb_rate / 100.0, 4)
                        _curr_total = round(_curr_total + _yb, 4)

                        for _iter_i in range(25):
                            # 保存本轮起始 total 用于敛散判断
                            _prev_total = _curr_total

                            # 1) 重算建设管理费 — 基数 = 项目总投资 − 建管费自身 − 切改费 − 用地费
                            #    有合同覆盖的费种保持不变
                            if "建设管理费" not in ctx.get("contract_overrides", {}):
                                _gl_old = numerical.get("建设管理费(万元)", 0)
                                _gl_base = _curr_total - _gl_old - _gl_qg - _gl_js
                                _gl_r = calc_jianshe_guanli(_gl_base)
                                numerical["建设管理费(万元)"] = _ext_num(_gl_r)

                            # 2) 重算可行性研究费（多服务类型用 calc_keyan_multi）
                            #    有合同覆盖的费种保持不变
                            if "可行性研究费" not in ctx.get("contract_overrides", {}):
                                _keyan_ind, _keyan_coef = _detect_keyan_industry(ctx["query"])
                                _ky_comp_coef = (ctx.get("coef_overrides", {})
                                                 .get("可行性研究费", {})
                                                 .get("complexity_coef", 1.0))
                                _ky_svcs = ctx.get("service_selections", {}).get("可行性研究费", [])
                                if _ky_svcs and _ky_svcs != ["编制可研报告"]:
                                    # 用户选择了多服务类型，用 calc_keyan_multi
                                    _ky_multi = calc_keyan_multi(
                                        _curr_total / 10000.0, _ky_svcs,
                                        industry_coef=_keyan_coef,
                                        complexity_coef=_ky_comp_coef,
                                    )
                                    numerical["可行性研究费(万元)"] = _ky_multi.get("合计(万元)", 0)
                                else:
                                    _amount_yi = _curr_total / 10000.0
                                    _keyan_r = calc_keyan(
                                        _amount_yi, service_type="编制可研报告",
                                        industry_coef=_keyan_coef, industry_name=_keyan_ind,
                                    )
                                    numerical["可行性研究费(万元)"] = _ext_num(_keyan_r)

                            # 3) 更新 T2
                            _t2_keys = ["建设管理费", "可行性研究费", "环境影响咨询费"]
                            _new_t2 = sum(
                                numerical.get(f"{k}(万元)", 0) for k in _t2_keys)

                            # 4) 用最新项目总投资重算造价咨询费（概算审核等依赖总投资）
                            #    有合同覆盖的费种保持不变
                            if "造价咨询费" not in ctx.get("contract_overrides", {}):
                                if is_hebei_region(st.session_state.get("selected_region")):
                                    _cc_multi = calc_cost_consulting_multi_hebei(
                                        cc_svcs, ctx["jianan"],
                                        total_investment=_curr_total,
                                        professional_coef=_cc_prof, discount_coef=1.0,
                                    )
                                else:
                                    _cc_multi = calc_cost_consulting_multi(
                                        cc_svcs, ctx["total_part1"],
                                        jianan_only=ctx["jianan"],
                                        total_investment=_curr_total,
                                    )
                                _cc_total = _cc_multi.get("合计(万元)", 0)
                                numerical["造价咨询费(万元)"] = _cc_total
                                preview_raw["原始结果"]["造价咨询费"] = _cc_multi

                            # 5) 更新 T0（CC 变了）
                            _new_t0 = sum(
                                numerical.get(f"{k}(万元)", 0) for k in t0_keys)

                            # 6) 重算汇总（含自定义费用 + 水保费）
                            _fee_total = _new_t0 + t1_total + _new_t2
                            _yb = round(
                                (ctx["total_part1"] + _fee_total + _custom_total
                                 + sb_fee_wan)
                                * _yb_rate / 100.0, 4)
                            _curr_total = round(
                                ctx["total_part1"] + _fee_total + _yb
                                + _custom_total + sb_fee_wan, 4)

                            # 7) 收敛判定（在重算后进行，确保最后一轮结果基于收敛 total）
                            if abs(_curr_total - _prev_total) < 0.005:
                                break

                        # 收敛后写入汇总值（含自定义费用 + 水保费）
                        cc_multi = preview_raw["原始结果"]["造价咨询费"]
                        cc_total = numerical["造价咨询费(万元)"]
                        preview_raw["T0小计(万元)"] = round(_new_t0, 4)
                        preview_raw["T2小计(万元)"] = round(_new_t2, 4)
                        preview_raw["二类费合计(万元)"] = round(
                            _fee_total + _custom_total + sb_fee_wan, 4)
                        preview_raw["总投资(万元)"] = round(
                            ctx["total_part1"] + _fee_total + _custom_total
                            + sb_fee_wan, 4)
                        preview_raw["预备费小计(万元)"] = round(_yb, 4)
                        numerical["预备费(万元)"] = _yb
                        preview_raw["项目总投资(万元)"] = round(
                            ctx["total_part1"] + _fee_total + _custom_total
                            + sb_fee_wan + _yb, 4)
                    except Exception as _e:
                        import traceback
                        print(f"[CC convergence ERROR] {type(_e).__name__}: {_e}", flush=True)
                        traceback.print_exc()
                        pass  # 失败时保留原始值

            # ── 若造价咨询费未跑迭代，需手动将水保费补入汇总值 ──
            if sb_fee_wan > 0:
                _cc_iter_ran = ("造价咨询费" in new_selected
                                and "造价咨询费" not in ctx.get("contract_overrides", {})
                                and bool(ctx.get("service_selections", {}).get("造价咨询费", [])))
                if not _cc_iter_ran:
                    preview_raw["二类费合计(万元)"] = round(
                        preview_raw.get("二类费合计(万元)", 0) + sb_fee_wan, 4)
                    preview_raw["总投资(万元)"] = round(
                        preview_raw.get("总投资(万元)", 0) + sb_fee_wan, 4)
                    preview_raw["项目总投资(万元)"] = round(
                        preview_raw.get("项目总投资(万元)", 0) + sb_fee_wan, 4)

            # ── 对非 T0 费种统一补应用折扣（T0 已由引擎应用）──
            _final_disc = ctx.get("fee_discounts", {})
            _T0_NAMES = {"监理费", "工程设计费", "勘察费",
                         "劳动安全卫生评审费", "场地准备费及临时设施费", "工程保险费"}
            _any_disc_applied = False
            if _final_disc:
                for _fn, _fd in _final_disc.items():
                    if _fn in _T0_NAMES or _fn == "预备费" or abs(_fd - 1.0) < 0.005:
                        continue
                    _key = f"{_fn}(万元)"
                    if _key in numerical and numerical[_key] > 0:
                        numerical[_key] = round(numerical[_key] * _fd, 4)
                        _any_disc_applied = True
            # 重算层级小计以反映折扣变动（含自定义费用）
            if _any_disc_applied:
                _tk0 = ["监理费", "工程设计费", "勘察费",
                        "劳动安全卫生评审费", "场地准备费及临时设施费",
                        "工程保险费", "造价咨询费"]
                _tk1 = ["交易服务费", "施工图审查费", "招标代理费"]
                _tk2 = ["建设管理费", "可行性研究费", "环境影响咨询费"]
                _nt0 = sum(numerical.get(f"{k}(万元)", 0) for k in _tk0)
                _nt1 = sum(numerical.get(f"{k}(万元)", 0) for k in _tk1)
                _nt2 = sum(numerical.get(f"{k}(万元)", 0) for k in _tk2)
                _nt_custom = sum(cf["amount_wan"] for cf in ctx.get("custom_fees", []))
                preview_raw["T0小计(万元)"] = round(_nt0, 4)
                preview_raw["T1小计(万元)"] = round(_nt1, 4)
                preview_raw["T2小计(万元)"] = round(_nt2, 4)
                preview_raw["二类费合计(万元)"] = round(
                    _nt0 + _nt1 + _nt2 + _nt_custom + sb_fee_wan, 4)

            # ── 存储计算结果供打折区和预览区共用 ──
            ctx["_computed_raw"] = preview_raw
            ctx["_computed_numerical"] = dict(numerical)
            ctx["_computed_sb_fee_wan"] = sb_fee_wan

            # 按层级显示
            for tier in [0, 1, 2, 4]:
                tier_fees = [
                    fd for fd in ctx["fee_defs"]
                    if fd["tier"] == tier and fd["name"] in new_selected
                ]
                if not tier_fees:
                    continue
                subtotal_raw = preview_raw.get(f'T{tier}小计(万元)', 0)
                subtotal_str = f"<small style='color:#888;'>{subtotal_raw:.2f} 万元</small>" if subtotal_raw > 0 else ""
                st.markdown(
                    f"**{tier_labels.get(tier, '')}**  {subtotal_str}",
                    unsafe_allow_html=True,
                )
                for fd in tier_fees:
                    fn = fd["name"]
                    val = numerical.get(f"{fn}(万元)")
                    if val is not None:
                        # 显示系数/费率标注
                        note_parts = []
                        if fn in ctx["coef_overrides"]:
                            coefs = ctx["coef_overrides"][fn]
                            for k, v in coefs.items():
                                if abs(v - 1.0) > 0.005:
                                    note_parts.append(f"{k}={v}")
                        if fn in ctx.get("rate_overrides", {}):
                            note_parts.append(f"费率={ctx['rate_overrides'][fn]}")
                        if fn in ctx.get("service_selections", {}):
                            svcs = ctx["service_selections"][fn]
                            if svcs and svcs != ["编制报告书"]:
                                note_parts.append(f"{len(svcs)}项服务")
                        if fn in ctx.get("contract_overrides", {}):
                            note_parts.append("合同")
                        # 打折标注（T0 折扣已由引擎应用，此处仅标注不重复乘）
                        disc = ctx.get("fee_discounts", {}).get(fn, 1.0)
                        if abs(disc - 1.0) >= 0.005:
                            note_parts.append(f"打折={disc:.2f}")
                        note_str = ""
                        if note_parts:
                            note_str = f" <small style='color:#888;'>({'，'.join(note_parts)})</small>"
                        st.markdown(
                            f"- {fd['label']}：**{val:.2f}** 万元{note_str}",
                            unsafe_allow_html=True,
                        )

            # 预备费（含费率输入）——二类费合计已含水保费，预备费基数完整
            _yb_rate_disp = float(st.session_state.get("cascade_yb_rate", 5.0))
            _yb_ctr_disp = (ctx.get("contract_overrides") or {}).get("预备费", {})
            _fee_for_yb = preview_raw.get("二类费合计(万元)", 0)
            if _yb_ctr_disp and _yb_ctr_disp.get("type") == "price":
                yb_val = numerical.get("预备费(万元)")
            else:
                yb_val = round((ctx["total_part1"] + _fee_for_yb) * _yb_rate_disp / 100.0, 4)
                numerical["预备费(万元)"] = yb_val
            if yb_val is not None and yb_val > 0:
                yb_col1, yb_col2 = st.columns([3, 1])
                with yb_col1:
                    st.markdown(f"**预备费**：**{yb_val:.2f}** 万元")
                with yb_col2:
                    # 从 widget key 读取当前值（首次默认 5.0）
                    _current_yb = st.session_state.get("cascade_yb_rate", 5.0)
                    st.number_input(
                        "预备费率(%)",
                        min_value=0.0, max_value=20.0,
                        value=float(_current_yb),
                        step=0.5, format="%.1f",
                        key="cascade_yb_rate",
                        help="(一类费+二类费)×预备费率",
                        label_visibility="visible",
                    )
            else:
                # 预备费未选中，不显示
                pass

            # 水土保持补偿费（独立于投资额的费种）
            if sb_fee_wan > 0:
                st.markdown(f"**🏗️ 水土保持补偿费**：**{sb_fee_wan:.2f}** 万元"
                            f" <small style='color:#888;'>({sb_fee_wan * 10000:,.0f} 元，含10%中央收入)</small>",
                            unsafe_allow_html=True)

            # 自定义费用
            custom_total = sum(cf["amount_wan"] for cf in ctx["custom_fees"])
            if ctx["custom_fees"]:
                st.markdown("**自定义费用**：")
                for cf in ctx["custom_fees"]:
                    st.markdown(f"- {cf['name']}：**{cf['amount_wan']:.2f}** 万元")

            # ── 汇总指标 ──
            st.markdown("---")
            fee_total_raw = preview_raw.get("二类费合计(万元)", 0)
            pt = preview_raw.get("项目总投资(万元)", 0)

            # 二类费合计（已含自定义费用 + 水土保持补偿费，迭代中已汇入）
            fee_total_with_custom = fee_total_raw
            # 预备费已在上方统一计算（yb_val），直接复用
            yb_total = yb_val if yb_val is not None else preview_raw.get("预备费小计(万元)", 0)
            # 项目总投资 = 一类费 + 二类费 + 预备费
            project_total_with_custom = round(ctx["total_part1"] + fee_total_with_custom + yb_total, 4)

            n_cols = 3
            col1, col2, col3 = st.columns(n_cols)
            col1.metric("二类费合计", f"{_round2(fee_total_with_custom)} 万元")
            col2.metric("预备费", f"{_round2(yb_total)} 万元")
            col3.metric("项目总投资", f"{_round2(project_total_with_custom)} 万元")

            # 存储预览结果供确认按钮使用
            ctx["preview"] = {
                "raw": preview_raw,
                "custom_total": custom_total,
                "fee_total_with_custom": fee_total_with_custom,
                "project_total_with_custom": project_total_with_custom,
                "yubei_total": yb_total,
                "numerical": numerical,
                "sb_fee_wan": sb_fee_wan,
            }
            st.session_state.pending_fee_selection["preview"] = ctx["preview"]

        except Exception as e:
            st.error(f"预览计算失败：{e}")
            import traceback
            st.code(traceback.format_exc())
            ctx["preview"] = None
            st.session_state.pending_fee_selection["preview"] = None

        st.markdown("---")

        # ── 导出 Excel ──
        st.markdown("---")
        try:
            excel_bytes = _build_cascade_excel(ctx)
            st.download_button(
                label="📥 导出 Excel",
                data=excel_bytes,
                file_name=f"二类费计算汇总_{ctx.get('project_type', '项目')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_cascade_excel",
            )
        except Exception:
            pass  # 导出失败不影响主流程

        # ── Confirm / Cancel ──
        st.markdown("---")
        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            if st.button("✅ 确认结果", type="primary", use_container_width=True,
                         key="confirm_fee_selection"):
                if not ctx["selected_fees"] and not ctx["custom_fees"]:
                    st.warning("请至少选择一项费用或添加自定义费用")
                    st.stop()

                # 构建响应文本
                preview = ctx.get("preview")
                numerical = preview["numerical"] if preview else {}
                fee_discounts = ctx.get("fee_discounts", {})
                custom_fees = ctx.get("custom_fees", [])
                fee_defs = ctx["fee_defs"]

                # 按层级生成费用明细（含各自打折，预备费不计入二类费合计）
                tier_lines = []
                discounted_fee_total = 0.0
                raw_fee_total = 0.0
                for tier in [0, 1, 2, 3]:
                    tier_fees = [
                        fd for fd in fee_defs
                        if fd["tier"] == tier and fd["name"] in ctx["selected_fees"]
                    ]
                    if not tier_fees:
                        continue
                    for fd in tier_fees:
                        fn = fd["name"]
                        val = numerical.get(f"{fn}(万元)")
                        if val is not None:
                            is_yubei = (fn == "预备费")
                            if not is_yubei:
                                raw_fee_total += val
                                disc = fee_discounts.get(fn, 1.0)
                                disc_val = round(val * disc, 4)
                                discounted_fee_total += disc_val
                            else:
                                disc = 1.0
                                disc_val = val
                            coef_note = ""
                            note_parts = []
                            if fn in ctx["coef_overrides"]:
                                coefs = ctx["coef_overrides"][fn]
                                for k, v in coefs.items():
                                    if abs(v - 1.0) > 0.005:
                                        note_parts.append(f"{k}={v}")
                            if fn in ctx.get("rate_overrides", {}):
                                note_parts.append(f"费率={ctx['rate_overrides'][fn]}")
                            if fn in ctx.get("service_selections", {}):
                                svcs = ctx["service_selections"][fn]
                                if svcs and svcs != ["编制报告书"]:
                                    note_parts.append(f"{'、'.join(svcs)}")
                            if fn in ctx.get("contract_overrides", {}):
                                ov_cfg = ctx["contract_overrides"][fn]
                                if ov_cfg.get("type") == "rate":
                                    note_parts.append(f"合同费率{ov_cfg['rate']}%")
                                else:
                                    note_parts.append("合同价")
                            if abs(disc - 1.0) >= 0.005:
                                note_parts.append(f"打折={disc:.2f}")
                            if note_parts:
                                coef_note = f"（{'，'.join(note_parts)}）"
                            display_val = disc_val if abs(disc - 1.0) >= 0.005 else val
                            tier_lines.append(f"- **{fd['label']}**{coef_note}：{display_val:.2f} 万元")

                # 自定义费用
                custom_lines = []
                custom_total_val = 0.0
                if custom_fees:
                    custom_lines.append("")
                    custom_lines.append("**自定义费用**：")
                    for cf in custom_fees:
                        custom_lines.append(f"- **{cf['name']}**：{cf['amount_wan']:.2f} 万元")
                        custom_total_val += cf["amount_wan"]

                # 水土保持补偿费
                sb_fee_wan = preview.get("sb_fee_wan", 0) if preview else 0
                sb_lines = []
                if sb_fee_wan > 0:
                    sb_params = ctx.get("shuibao_comp_params", {})
                    sb_detail = numerical.get("水土保持补偿费(万元)", sb_fee_wan)
                    sb_lines.append("")
                    sb_lines.append(f"**🏗️ 水土保持补偿费**：{sb_detail:.2f} 万元"
                                    f"（{sb_detail * 10000:,.0f} 元，含10%中央收入）")

                # 打折文本
                discount_text = ""
                if abs(discounted_fee_total - raw_fee_total) > 0.005:
                    discount_text = (
                        f"\n\n**打折后二类费合计**：{discounted_fee_total:.2f} 万元"
                        f"（打折前 {raw_fee_total:.2f} 万元）"
                    )

                # 加上自定义费用 + 水土保持补偿费
                discounted_with_custom = discounted_fee_total + custom_total_val + sb_fee_wan

                # 预备费文本
                yb_val = preview["yubei_total"] if preview else 0
                yb_text = ""
                if yb_val > 0:
                    yb_text = f"\n\n**预备费（基本预备费）**：{yb_val:.2f} 万元"

                project_total_val = preview["project_total_with_custom"] if preview else 0

                final_response = (
                    f"## 多费种联算结果\n\n"
                    f"计费基数：建安费 **{ctx['jianan']}** 万 + 设备费 **{ctx['shebei']}** 万 "
                    f"= **{ctx['total_part1']}** 万元 ｜ "
                    f"项目类型：**{ctx['project_type']}**\n\n"
                    f"### 各项费用\n\n"
                    + "\n".join(tier_lines) +
                    (f"\n\n**二类费合计**：{discounted_with_custom:.2f} 万元"
                     if not discount_text else "")
                    + (f"\n".join(custom_lines) if custom_lines else "")
                    + (f"\n".join(sb_lines) if sb_lines else "")
                    + discount_text
                    + yb_text
                    + f"\n\n**项目总投资**：{project_total_val:.2f} 万元"
                )

                st.session_state.messages.append({
                    "role": "assistant",
                    "content": final_response,
                })
                del st.session_state.pending_fee_selection
                st.rerun()

        with col_btn2:
            if st.button("🗑 取消", use_container_width=True,
                         key="cancel_fee_selection"):
                del st.session_state.pending_fee_selection
                st.rerun()

# ===== 输入框 =====
st.divider()
st.markdown("### 输入你的问题")

if "current_query" in st.session_state:
    prompt = st.session_state.current_query
    del st.session_state.current_query
else:
    prompt = st.chat_input("请输入你的造价问题，例如：白皮松高度3.5米多少钱？")

if prompt:
    # 新提问时清除旧的待处理选择
    st.session_state.pop("pending_rate_select", None)
    st.session_state.pop("pending_coef_select", None)
    st.session_state.pop("pending_dependent_fee", None)
    st.session_state.pop("pending_huanping", None)
    st.session_state.pop("pending_keyan", None)
    st.session_state.pop("pending_shuibao_compensation", None)
    st.session_state.pop("pending_jiaoyi_party", None)
    st.session_state.pop("pending_fee_selection", None)
    # 添加用户消息
    st.session_state.messages.append({"role": "user", "content": prompt})

    # 调试：将检测结果存入 session_state，跨 rerun 持久化
    try:
        debug_fee = detect_and_calculate(prompt, region=st.session_state.get("selected_region"))
        st.session_state.debug_info = {
            "prompt": prompt[:80],
            "fee_type": debug_fee.get("fee_type") if debug_fee else "None",
            "has_amount": debug_fee.get("has_amount") if debug_fee else "N/A",
            "费种": debug_fee.get("费种", "") if debug_fee else "",
            "结果": str(debug_fee.get("结果(万元)", debug_fee.get("结果(元)", ""))) if debug_fee else "",
        }
    except Exception as e:
        st.session_state.debug_info = {"error": str(e)}
        debug_fee = None

    # 关键：fee_result 必须在后续分支中使用
    fee_result = debug_fee

    with st.chat_message("user"):
        st.markdown(prompt)

    # 生成回答
    with st.chat_message("assistant"):
        try:
            fee_result = detect_and_calculate(prompt, region=st.session_state.get("selected_region"))
        except Exception as e:
            import traceback
            st.code(traceback.format_exc())
            fee_result = None

        with st.spinner("正在检索数据并生成回答..."):
            # 二类费规则引擎
            # fee_result = detect_and_calculate(prompt)  # 已在上方调用

            if fee_result and fee_result.get("has_amount"):
                # === 多费种迭代模式路由 ===
                mode = fee_result.get("mode")
                if mode == "cascade":
                    raw = fee_result.get("_engine_raw", {})
                    params = fee_result["输入参数"]
                    jianan = params["建安工程费(万元)"]
                    shebei = params["设备购置费(万元)"]
                    project_type = params["项目类型"]
                    total_part1 = params["第一部分工程费(万元)"]

                    # 始终用最新代码重建 fee_defs（支持热更新 service_config）
                    fee_defs = _build_fee_selection_meta(raw, prompt, region=st.session_state.get("selected_region"))
                    # 🔧 DEBUG
                    for _fd in fee_defs:
                        if _fd["name"] == "施工图审查费":
                            _rc = _fd.get("rate_config") or {}
                            print(f"DEBUG _build_fee_selection_meta: query={prompt[:50]} basis={_rc.get('basis')} default={_rc.get('default_key')} n_opts={len(_rc.get('rate_options',[]))}")

                    if "pending_fee_selection" not in st.session_state:
                        # 首次：初始化交互式费种选择面板（交易服务费默认不选中）
                        selected_fees = set(
                            fd["name"] for fd in fee_defs
                            if fd["name"] != "交易服务费"
                        )

                        coef_overrides = {}
                        for fd in fee_defs:
                            if fd["has_coefs"] and fd["coef_config"]:
                                defaults = {}
                                for c in fd["coef_config"]["coefs"]:
                                    defaults[c["param_name"]] = c["current"]
                                coef_overrides[fd["name"]] = defaults

                        st.session_state.pending_fee_selection = {
                            "query": prompt,
                            "jianan": jianan,
                            "shebei": shebei,
                            "project_type": project_type,
                            "total_part1": total_part1,
                            "fee_defs": fee_defs,
                            "selected_fees": selected_fees,
                            "coef_overrides": coef_overrides,
                            "rate_overrides": {},
                            "service_selections": {
                                "环境影响咨询费": ["编制报告书"],
                                "可行性研究费": ["编制可研报告"],
                                "造价咨询费": (
                                    ["预算编制"]
                                    if is_hebei_region(st.session_state.get("selected_region"))
                                    else ["编制施工图预算"]
                                ),
                            },
                            "custom_fees": [],
                            "contract_overrides": {},
                            "fee_discounts": {fd["name"]: 1.0 for fd in fee_defs},
                            "preview": None,
                            "shuibao_comp_params": {
                                "calc_type": "general",
                                "land_input": 0.0, "land_unit": "m²",
                                "well_cnt": 0, "add_wells": 0,
                                "extract_vol": 0.0, "material_vol": 0.0,
                                "waste_vol": 0.0,
                            },
                        }
                    else:
                        # 更新已有的 session：替换 fee_defs（保持用户已选状态）
                        ctx = st.session_state.pending_fee_selection
                        ctx["fee_defs"] = fee_defs
                        ctx["query"] = prompt
                        ctx["jianan"] = jianan
                        ctx["shebei"] = shebei
                        ctx["total_part1"] = total_part1
                        ctx["project_type"] = project_type
                        # 补上可能缺失的 service_selections
                        svc = ctx.setdefault("service_selections", {})
                        svc.setdefault("环境影响咨询费", ["编制报告书"])
                        svc.setdefault("可行性研究费", ["编制可研报告"])
                        if "造价咨询费" not in svc:
                            svc["造价咨询费"] = (
                                ["预算编制"] if is_hebei_region(st.session_state.get("selected_region"))
                                else ["编制施工图预算"]
                            )
                        # 补上可能缺失的 coef_overrides（新费种）
                        for fd in fee_defs:
                            if fd["has_coefs"] and fd["coef_config"]:
                                if fd["name"] not in ctx["coef_overrides"]:
                                    defaults = {}
                                    for c in fd["coef_config"]["coefs"]:
                                        defaults[c["param_name"]] = c["current"]
                                    ctx["coef_overrides"][fd["name"]] = defaults
                        # 补上可能缺失的 fee_discounts
                        discounts = ctx.setdefault("fee_discounts", {})
                        for fd in fee_defs:
                            discounts.setdefault(fd["name"], 1.0)
                        # 补上可能缺失的 contract_overrides
                        ctx.setdefault("contract_overrides", {})
                        # 补上可能缺失的 shuibao_comp_params
                        ctx.setdefault("shuibao_comp_params", {
                            "calc_type": "general",
                            "land_input": 0.0, "land_unit": "m²",
                            "well_cnt": 0, "add_wells": 0,
                            "extract_vol": 0.0, "material_vol": 0.0,
                            "waste_vol": 0.0,
                        })

                    n_fees = len(st.session_state.pending_fee_selection["fee_defs"])
                    response = (
                        f"## 多费种联算\n\n"
                        f"> 📋 该模式支持 {n_fees} 项二类费\n\n"
                        f"请滚动到页面下方 **📋 二类费选择 — 交互式联算** 区域，"
                        f"勾选需要计算的费种后可调整系数、添加自定义费用。"
                    )
                elif mode == "iteration":
                    response = _render_iteration_result(fee_result)
                elif mode == "comparison":
                    response = _render_comparison_result(fee_result)

                if mode is None:
                    # === 引擎精确计算：单费种直接展示，不经过 LLM ===
                    needs_dep = fee_result.get("needs_dependent_config", False)
                    is_sheji = fee_result.get("fee_type") == "工程设计费"
                    is_rate_selectable = fee_result.get("is_rate_selectable", False)
                    is_coef_selectable = fee_result.get("is_coef_selectable", False)

                    if needs_dep:
                        # === 依赖费种交互式配置（招标代理费 & 施工图审查费）===
                        st.session_state.pending_dependent_fee = {
                            "target_fee": fee_result["target_fee"],
                            "target_fee_name": fee_result["target_fee_name"],
                            "dependent_fees": fee_result["dependent_fees"],
                            "base_params": fee_result["base_params"],
                            "query": prompt,
                            "step": "config",
                            "final_result": None,
                            "discount_coef": 1.0,
                            "dep_discounts": {},
                            "dep_custom_amounts": {},
                        }
                        fee_name = fee_result.get("费种", "")
                        n_deps = len(fee_result.get("dependent_fees", []))
                        dep_names = "、".join(
                            d.get("fee_label", d.get("fee_type", ""))
                            for d in fee_result["dependent_fees"]
                        )
                        response = (
                            f"## {fee_name}\n\n"
                            f"> 🔗 该费种依赖 {n_deps} 个其他费种的计算结果\n\n"
                            f"请滚动到页面下方 **🔗 依赖费种配置** 区域，"
                            f"先配置 **{dep_names}** 的参数后点击确认计算。"
                        )
                    elif fee_result.get("needs_huanping_select"):
                        # === 环评费多服务类型选择 ===
                        st.session_state.pending_huanping = {
                            "amount_wan": fee_result["amount_wan"],
                            "estimated_investment": fee_result.get("estimated_investment", fee_result["amount_wan"]),
                            "has_explicit_investment": fee_result.get("has_explicit_investment", False),
                            "industry_coef": fee_result.get("industry_coef", 1.0),
                            "industry_name": fee_result.get("industry_name", ""),
                            "sensitivity_coef": fee_result.get("sensitivity_coef", 1.0),
                            "query": prompt,
                            "discount_coef": 1.0,
                        }
                        response = (
                            f"## 环境影响咨询费\n\n"
                            f"> 🌿 该费种包含 4 项服务类型\n\n"
                            f"请滚动到页面下方 **🌿 环评费 — 服务类型选择** 区域，"
                            f"选择需要计算的服务类型并调整系数后点击确认。"
                        )
                    elif fee_result.get("needs_keyan_select"):
                        # === 可行性研究费多服务类型选择 ===
                        st.session_state.pending_keyan = {
                            "amount_yi": fee_result["amount_yi"],
                            "industry_coef": fee_result.get("industry_coef", 1.0),
                            "industry_name": fee_result.get("industry_name", ""),
                            "complexity_coef": fee_result.get("complexity_coef", 1.0),
                            "query": prompt,
                            "discount_coef": 1.0,
                        }
                        response = (
                            f"## 建设项目前期工作咨询费\n\n"
                            f"> 📊 该费种包含 4 项服务类型\n\n"
                            f"请滚动到页面下方 **📊 建设项目前期工作咨询费 — 服务类型选择** 区域，"
                            f"选择需要计算的服务类型并调整系数后点击确认。"
                        )
                    elif fee_result.get("needs_jiaoyi_party_select"):
                        # === 交易服务费计费方选择 ===
                        jiaoyi_result = fee_result
                        st.session_state.pending_jiaoyi_party = {
                            "fee_result": jiaoyi_result,
                            "query": prompt,
                        }
                        # 依赖费种文本（含系数详情）
                        deps_info = fee_result.get("依赖费种", {})
                        dep_lines = []
                        # 监理费
                        if deps_info.get("监理费(万元)"):
                            jl_line = f"- 监理费：**{deps_info['监理费(万元)']} 万元**"
                            jl_params = deps_info.get("监理费_参数", {})
                            if jl_params:
                                coef_strs = []
                                for cn in ["专业调整系数", "复杂程度系数", "高程调整系数"]:
                                    cv = jl_params.get(cn, 1.0)
                                    if cv is not None and abs(cv - 1.0) >= 0.005:
                                        coef_strs.append(f"{cn}={cv}")
                                if coef_strs:
                                    jl_line += f"（{'，'.join(coef_strs)}）"
                            dep_lines.append(jl_line)
                        # 设计费
                        if deps_info.get("设计费(万元)"):
                            sj_line = f"- 设计费：**{deps_info['设计费(万元)']} 万元**"
                            sj_params = deps_info.get("设计费_参数", {})
                            if sj_params:
                                coef_strs = []
                                for cn in ["专业调整系数", "复杂程度系数", "附加调整系数"]:
                                    cv = sj_params.get(cn, 1.0)
                                    if cv is not None and abs(cv) >= 0.005:
                                        coef_strs.append(f"{cn}={cv}")
                                if coef_strs:
                                    sj_line += f"（{'，'.join(coef_strs)}）"
                            dep_lines.append(sj_line)
                        dep_section = ""
                        if dep_lines:
                            dep_section = "### 📎 依赖费种（计算基数）\n\n" + "\n".join(dep_lines) + "\n\n"
                        response = (
                            f"## 工程建设交易服务费\n\n"
                            f"{dep_section}"
                            f"> 🏛️ 请选择计费方\n\n"
                            f"请滚动到页面下方 **🏛️ 交易服务费 — 计费方选择** 区域，"
                            f"选择招标方或中标方后点击确认。"
                        )
                    elif fee_result.get("needs_shuibao_compensation_select"):
                        # === 水土保持补偿费参数输入 ===
                        st.session_state.pending_shuibao_compensation = {
                            "calc_type": fee_result.get("calc_type", "general"),
                            "query": prompt,
                            "land_input": 0.0,
                            "well_cnt": 0,
                            "add_wells": 0,
                            "extract_vol": 0.0,
                            "material_vol": 0.0,
                            "waste_vol": 0.0,
                        }
                        response = (
                            f"## 水土保持补偿费\n\n"
                            f"> 🏗️ 该费种需要输入物理参数\n\n"
                            f"请滚动到页面下方 **🏗️ 水土保持补偿费 — 参数输入** 区域，"
                            f"选择计征类型并输入参数后点击确认。"
                        )
                    elif is_rate_selectable:
                        # === 交互式费率选择：存入 session state，在聊天区外渲染 ===
                        st.session_state.pending_rate_select = {
                            "fee_result": fee_result,
                            "query": prompt,
                        }
                        fee_name = fee_result.get("费种", "")
                        n_rates = len(fee_result.get("费率明细", []))
                        response = (
                            f"## {fee_name}\n\n"
                            f"> ℹ️ 该费种支持交互式费率选择\n\n"
                            f"请滚动到页面下方 **🎯 费率选择** 区域，"
                            f"从 {n_rates} 档费率中选择适用费率后点击确认。"
                        )
                    elif is_coef_selectable:
                        # === 交互式系数选择：存入 session state，在聊天区外渲染 ===
                        st.session_state.pending_coef_select = {
                            "coef_metadata": fee_result.get("coef_metadata", {}),
                            "fee_result": fee_result,
                            "query": prompt,
                        }
                        fee_name = fee_result.get("费种", "")
                        n_coefs = len(fee_result.get("coef_metadata", {}).get("coefs", []))
                        response = (
                            f"## {fee_name}\n\n"
                            f"> ℹ️ 该费种支持交互式系数调整\n\n"
                            f"请滚动到页面下方 **🎛️ 系数调整** 区域，"
                            f"调整 {n_coefs} 个系数后点击确认。"
                        )
                    else:
                        st.markdown("### 计算结果（程序精确计算）")
                        _render_engine_card(fee_result)

                    if needs_dep or is_rate_selectable or is_coef_selectable or fee_result.get("needs_huanping_select") or fee_result.get("needs_keyan_select") or fee_result.get("needs_shuibao_compensation_select") or fee_result.get("needs_jiaoyi_party_select"):
                        pass  # 已在上面通过 pending_* 处理完成
                    elif is_sheji:
                        st.divider()
                        _render_sheji_static(fee_result)
                        response = _build_sheji_text(fee_result)
                    else:
                        # 非设计费：也跳过 LLM，用静态确认
                        st.divider()
                        fee_name = fee_result.get("费种", "")
                        result_val = fee_result.get("结果(万元)") or fee_result.get("结果(元)")
                        unit = "万元" if "结果(万元)" in fee_result else "元"
                        basis = fee_result.get("依据", "")
                        desc = fee_result.get("说明", "")
                        ft = fee_result.get("fee_type", "")

                        # ── 打折系数（所有非 pending 费种统一在此渲染）──
                        base_fee_wan = _get_fee_numeric(fee_result)
                        default_discount = fee_result.get("_discount_coef", 1.0)
                        discount_coef = 1.0
                        discounted_fee_wan = base_fee_wan
                        if base_fee_wan is not None:
                            st.markdown("### 💰 费用打折")
                            discount_coef = st.number_input(
                                "打折系数（1.0 = 不打折，0.8 = 打八折）",
                                min_value=0.01, max_value=2.00,
                                value=default_discount, step=0.05,
                                format="%.2f",
                                key=f"discount_{ft}",
                                help="输入打折系数调整最终费用。",
                            )
                            discounted_fee_wan = round(base_fee_wan * discount_coef, 4)
                            if abs(discount_coef - 1.0) < 0.005:
                                st.info(f"**不打折**，最终费用：**{discounted_fee_wan} 万元**")
                            elif discount_coef < 1.0:
                                st.warning(
                                    f"打折系数 **{discount_coef:.2f}** → "
                                    f"{base_fee_wan:.2f} 万 × {discount_coef:.2f} = "
                                    f"**{discounted_fee_wan} 万元**"
                                    f"（节省 {round(base_fee_wan - discounted_fee_wan, 4)} 万元）"
                                )
                            else:
                                st.warning(
                                    f"上浮系数 **{discount_coef:.2f}** → "
                                    f"{base_fee_wan:.2f} 万 × {discount_coef:.2f} = "
                                    f"**{discounted_fee_wan} 万元**"
                                    f"（增加 {round(discounted_fee_wan - base_fee_wan, 4)} 万元）"
                                )
                            st.markdown("---")

                        # ── 打折文本（公共作用域，所有非 pending 费种共用）──
                        discount_text = ""
                        discounted_display = result_val
                        discounted_unit = unit
                        if base_fee_wan is not None and abs(discount_coef - 1.0) >= 0.005:
                            discount_text = (
                                f"\n\n**打折系数**：{discount_coef:.2f}\n\n"
                                f"**打折后费用**：{discounted_fee_wan} 万元"
                                f"（{base_fee_wan:.2f} 万 × {discount_coef:.2f}）"
                            )
                            discounted_display = discounted_fee_wan
                            discounted_unit = "万元"

                        # 施工图审查费（津价管[2011]46号 + 建市[2007]86号）
                        is_shencha = ft == "施工图审查费"
                        # 环评费（计价格[2002]125号 — 四项服务类型全部输出）
                        is_huanping = ft == "环境影响咨询费"
                        # 建设项目前期工作咨询费（计价格[1999]1283号 — 内插法详细步骤）
                        is_keyan = ft == "可行性研究费"
                        # 粗略估算类费种（《市政工程设计概算编制办法》）
                        is_rough = ft in (
                            "勘察费", "劳动安全卫生评审费",
                            "场地准备费及临时设施费", "工程保险费",
                        )
                        # 招标代理费多类型（5 类自动汇总）
                        is_zhaobiao_multi = fee_result.get("is_zhaobiao_multi", False)

                        if is_zhaobiao_multi:
                            detail_list = fee_result.get("明细", [])
                            total_fee = fee_result.get("合计(万元)", 0)
                            deps = fee_result.get("依赖费种", {})

                            # 依赖费种展示
                            st.markdown("### 依赖费种（自动计算）")
                            dep_cols = st.columns(3)
                            dep_items = [
                                ("监理费", deps.get("监理费(万元)", 0)),
                                ("设计费", deps.get("设计费(万元)", 0)),
                                ("勘察费", deps.get("勘察费(万元)", 0)),
                            ]
                            for i, (name, val) in enumerate(dep_items):
                                with dep_cols[i]:
                                    st.metric(label=name, value=f"{val} 万元")

                            # 费用明细表
                            st.markdown("### 招标代理费明细")
                            detail_rows = []
                            for d in detail_list:
                                dtype = d.get("类型", "")
                                dbase = d.get("基数(万元)", 0)
                                dsrc = d.get("基数来源", "")
                                dfee = d.get("费用(万元)", 0)
                                note = d.get("说明", "")
                                if note:
                                    detail_rows.append(f"| **{dtype}** | {dsrc} | {dbase:.2f} | ⚠️ {note} |")
                                else:
                                    detail_rows.append(f"| **{dtype}** | {dsrc} | {dbase:.2f} | **{dfee}** |")
                            st.markdown(
                                "| 类型 | 基数来源 | 基数（万元） | 费用（万元） |\n"
                                "|:--|:--|:--|:--|\n" + "\n".join(detail_rows)
                            )

                            # 各子项分档计算
                            with st.expander("📐 查看各子项分档计算过程"):
                                for d in detail_list:
                                    dtype = d.get("类型", "")
                                    dfee = d.get("费用(万元)", 0)
                                    steps = d.get("计算步骤", [])
                                    note = d.get("说明", "")
                                    if note:
                                        st.warning(f"**{dtype}**：{note}")
                                        continue
                                    if steps:
                                        st.markdown(f"#### {dtype}（{dfee} 万元）")
                                        step_rows = []
                                        for s in steps:
                                            qujian = s.get("区间", "")
                                            amt = s.get("金额(万元)", "")
                                            rate = s.get("费率(%)", "")
                                            fee_v = s.get("费用(万元)", "")
                                            step_rows.append(f"| {qujian} | {amt} | {rate}% | **{fee_v}** |")
                                        st.markdown(
                                            "| 区间（万元） | 金额（万元） | 费率 | 费用（万元） |\n"
                                            "|:--|:--|:--|:--|\n" + "\n".join(step_rows)
                                        )
                                        st.markdown("")

                            st.markdown(f"### 💰 招标代理费合计：**{total_fee} 万元**")

                            # 打折
                            discount_text = ""
                            discounted_display = total_fee
                            discounted_unit = "万元"
                            if total_fee and total_fee > 0:
                                discount_coef = st.number_input(
                                    "打折系数",
                                    min_value=0.01, max_value=2.00,
                                    value=default_discount, step=0.05,
                                    format="%.2f",
                                    key=f"discount_zhaobiao",
                                    help="输入打折系数调整最终费用。",
                                )
                                discounted_total = round(total_fee * discount_coef, 4)
                                if abs(discount_coef - 1.0) < 0.005:
                                    st.info(f"**不打折**，最终总费用：**{discounted_total} 万元**")
                                elif discount_coef < 1.0:
                                    st.warning(
                                        f"打折系数 **{discount_coef:.2f}** → "
                                        f"{total_fee} × {discount_coef:.2f} = **{discounted_total} 万元**"
                                        f"（节省 {round(total_fee - discounted_total, 4)} 万元）"
                                    )
                                    discount_text = (
                                        f"\n\n**打折系数**：{discount_coef:.2f}\n\n"
                                        f"**打折后总费用**：{discounted_total} 万元"
                                    )
                                else:
                                    st.warning(
                                        f"上浮系数 **{discount_coef:.2f}** → **{discounted_total} 万元**"
                                    )
                                    discount_text = (
                                        f"\n\n**上浮系数**：{discount_coef:.2f}\n\n"
                                        f"**上浮后总费用**：{discounted_total} 万元"
                                    )
                                discounted_display = discounted_total

                            st.success(
                                f"以上为《招标代理业务收费管理暂行办法》（计价格[2002]1980号）精确结果。\n\n"
                                f"招标代理费合计：**{discounted_display} {discounted_unit}**{discount_text}"
                            )

                            # 构建持久化 response
                            dep_md = "\n".join(
                                f"- {name}：**{val} 万元**"
                                for name, val in dep_items
                            )
                            detail_md = ""
                            for d in detail_list:
                                dtype = d.get("类型", "")
                                dsrc = d.get("基数来源", "")
                                dbase = d.get("基数(万元)", 0)
                                dfee = d.get("费用(万元)", 0)
                                note = d.get("说明", "")
                                if note:
                                    detail_md += f"- **{dtype}**（{dsrc}）：{note}\n"
                                else:
                                    detail_md += (
                                        f"- **{dtype}**：基数 {dsrc} {dbase:.2f} 万 → **{dfee} 万元**\n"
                                    )
                            response = (
                                f"## 招标代理服务费\n\n"
                                f"**依据**：{_basis_md_links('《招标代理业务收费管理暂行办法》（计价格[2002]1980号）')}\n\n"
                                f"### 依赖费种\n\n{dep_md}\n\n"
                                f"### 费用明细\n\n{detail_md}\n"
                                f"### 💰 合计：**{discounted_display} {discounted_unit}**{discount_text}\n\n"
                                f"共 5 类招标代理费，可上下浮动 20%。"
                            )

                        elif is_shencha or is_huanping or is_rough or is_keyan:
                            # 构建完整 markdown 响应（跨 rerun 持久化）
                            mid_val = fee_result.get("结果中值(万元)")
                            mid_text = f"（中值约 **{mid_val} 万元**）" if mid_val else ""

                            # 构建计算过程
                            steps = fee_result.get("计算步骤", [])
                            steps_md = ""
                            if steps:
                                steps_md = "### 计算过程\n\n"
                                for i, s in enumerate(steps, 1):
                                    step_name = s.get("步骤", "")
                                    formula = s.get("公式", "")
                                    result_step = s.get("结果", "")
                                    steps_md += f"**{i}. {step_name}**：{formula} → **{result_step}**\n\n"

                            # 构建费率明细表
                            detail = fee_result.get("费率明细", [])
                            detail_md = ""
                            if detail:
                                detail_md = "### 费率-费用对照表\n\n"
                                detail_md += "| 费率 | 费用（万元） |\n"
                                detail_md += "|------|-------------|\n"
                                for d in detail:
                                    detail_md += f"| {d['费率']} | **{d['费用(万元)']}** |\n"
                                detail_md += "\n"

                            if is_shencha:
                                response = (
                                    f"## {fee_name}\n\n"
                                    f"**依据**：{_basis_md_links(basis)}\n\n"
                                    f"{steps_md}"
                                    f"---\n\n"
                                    f"### 计算结果\n\n"
                                    f"审查费：**{discounted_display} {discounted_unit}**{discount_text}\n\n"
                                    f"{desc}"
                                )
                            elif is_huanping:
                                # 四种服务类型结果表
                                all_svc = fee_result.get("全部服务类型结果", {})
                                svc_table = ""
                                if all_svc:
                                    svc_table = "### 四种服务类型全部结果\n\n"
                                    svc_table += "| 服务类型 | 费用范围（万元） | 中值（万元） |\n"
                                    svc_table += "|----------|:--:|:--:|\n"
                                    for svc_name in ["编制报告书", "编制报告表", "评估报告书", "评估报告表"]:
                                        svc_r = all_svc.get(svc_name, {})
                                        svc_table += (
                                            f"| **{svc_name}** "
                                            f"| {svc_r.get('结果(万元)', '-')} "
                                            f"| {svc_r.get('结果中值(万元)', '-')} |\n"
                                        )
                                    svc_table += "\n"
                                response = (
                                    f"## {fee_name}\n\n"
                                    f"**依据**：{_basis_md_links(basis)}\n\n"
                                    f"{steps_md}"
                                    f"---\n\n"
                                    f"{svc_table}"
                                    f"### 计算结果\n\n"
                                    f"基准收费：**{discounted_display} {discounted_unit}**{discount_text}\n\n"
                                    f"{desc}"
                                )
                            elif is_keyan:
                                # 服务类型结果表（2项或4项）
                                all_svc = fee_result.get("全部服务类型结果", {})
                                svc_table = ""
                                if all_svc:
                                    n = len(all_svc)
                                    if n == 4:
                                        title = "### 四种服务类型全部结果"
                                    elif n == 2:
                                        first_key = list(all_svc.keys())[0]
                                        if "可研" in first_key:
                                            title = "### 可研报告相关服务类型结果"
                                        else:
                                            title = "### 项目建议书相关服务类型结果"
                                    else:
                                        title = "### 服务类型结果"
                                    svc_table = f"{title}\n\n"
                                    svc_table += "| 服务类型 | 基准价（万元） | 最终费用（万元） |\n"
                                    svc_table += "|----------|:--:|:--:|\n"
                                    for svc_name in all_svc:
                                        svc_r = all_svc[svc_name]
                                        base = svc_r.get("基准价(万元)", "-")
                                        fee = svc_r.get("结果(万元)", "-")
                                        svc_table += f"| **{svc_name}** | {base} | {fee} |\n"
                                    svc_table += "\n"
                                n_svc = len(all_svc) if all_svc else 0
                                if n_svc > 0:
                                    response = (
                                        f"## {fee_name}\n\n"
                                        f"**依据**：{_basis_md_links(basis)}\n\n"
                                        f"{svc_table}"
                                        f"{steps_md}"
                                        f"---\n\n"
                                        f"### 计算结果\n\n"
                                        f"最终费用：**{discounted_display} {discounted_unit}**{discount_text}\n\n"
                                        f"{desc}"
                                    )
                                else:
                                    response = (
                                        f"## {fee_name}\n\n"
                                        f"**依据**：{_basis_md_links(basis)}\n\n"
                                        f"{steps_md}"
                                        f"---\n\n"
                                        f"### 计算结果\n\n"
                                        f"最终费用：**{discounted_display} {discounted_unit}**{discount_text}\n\n"
                                        f"{desc}"
                                    )
                            else:
                                # === 粗略估算类费种 — 交互式费率选择 ===
                                detail = fee_result.get("费率明细", [])
                                if detail:
                                    rate_opts = [d['费率'] for d in detail]
                                    fee_map = {d['费率']: d['费用(万元)'] for d in detail}
                                    mid_idx = len(detail) // 2

                                    st.markdown("### 请选择适用费率（间隔 0.1%）")
                                    selected_rate = st.selectbox(
                                        "费率",
                                        rate_opts,
                                        index=mid_idx,
                                        key=f"rate_select_{ft}",
                                        label_visibility="collapsed",
                                    )
                                    selected_fee = fee_map[selected_rate]
                                    st.success(
                                        f"选定费率 **{selected_rate}** → "
                                        f"**{fee_name}**：**{selected_fee} 万元**"
                                    )

                                    with st.expander("查看完整费率对照表"):
                                        st.markdown(detail_md)
                                else:
                                    selected_rate = "中值"
                                    selected_fee = mid_val

                                response = (
                                    f"## {fee_name}\n\n"
                                    f"**依据**：{_basis_md_links(basis)}\n\n"
                                    f"{steps_md}"
                                    f"{detail_md}"
                                    f"---\n\n"
                                    f"### 计算结果\n\n"
                                    f"选定费率：**{selected_rate}**\n\n"
                                    f"费用：**{discounted_display} {discounted_unit}**{discount_text}\n\n"
                                    f"{desc}"
                                )
                        else:
                            # ── 简单费种：保存到 session state，在聊天区外渲染打折交互 ──
                            if ft == "造价咨询费":
                                st.session_state.pending_cost_consulting = {
                                    "fee_result": fee_result,
                                    "query": prompt,
                                    "default_discount": fee_result.get("_discount_coef", 1.0),
                                }
                                response = (
                                    f"## 造价咨询费\n\n"
                                    f"> ℹ️ 请选择需要的服务子项\n\n"
                                    f"请滚动到页面下方 **📋 造价咨询服务选择** 区域勾选服务子项后点击确认。"
                                )
                            else:
                                st.session_state.pending_simple_fee = {
                                    "fee_result": fee_result,
                                    "query": prompt,
                                    "default_discount": fee_result.get("_discount_coef", 1.0),
                                }
                                fee_name = fee_result.get("费种", "")
                                response = (
                                    f"## {fee_name}\n\n"
                                    f"> ℹ️ 该费种支持打折调整\n\n"
                                    f"请滚动到页面下方 **💰 费用打折** 区域调整打折系数后点击确认。"
                                )

            elif fee_result and not fee_result.get("has_amount"):
                # === 无金额参考模式 ===
                is_sheji = (fee_result.get("fee_type") == "工程设计费")

                if is_sheji:
                    import pandas as pd

                    st.markdown("### 附表一：工程设计收费基价表")
                    st.caption("依据：《工程勘察设计收费管理规定》（计价格[2002]10号）单位：万元")
                    rate_table = fee_result.get("费率表", [])
                    if rate_table:
                        header = rate_table[0]
                        rows = rate_table[1:]
                        df_rate = pd.DataFrame(rows, columns=header)
                        st.table(df_rate)
                        st.caption("注：计费额 > 2000000 万元的，以计费额乘以 1.6% 的收费率计算收费基价。")

                    st.markdown("---")
                    st.markdown("### 附表二：工程设计收费专业调整系数表")
                    sheji_table_data = [
                        ("1、矿山采选工程", "黑色、黄金、化学、非金属及其他矿采选工程", "1.1"),
                        ("", "采煤工程，有色、铀矿采选工程", "1.2"),
                        ("", "选煤及其他煤炭工程", "1.3"),
                        ("2、加工冶炼工程", "各类冷加工工程", "1"),
                        ("", "船舶水工工程", "1.1"),
                        ("", "各类冶炼、热加工、压力加工工程", "1.2"),
                        ("", "核加工工程", "1.3"),
                        ("3、石油化工工程", "石油、化工、石化、化纤、医药工程", "1.2"),
                        ("", "核化工工程", "1.6"),
                        ("4、水利电力工程", "风力发电、其他水利工程", "0.8"),
                        ("", "火电工程", "1"),
                        ("", "核电常规岛、水电、水库、送变电工程", "1.2"),
                        ("", "核能工程", "1.6"),
                        ("5、交通运输工程", "机场场道工程", "0.8"),
                        ("", "公路、城市道路工程", "0.9"),
                        ("", "机场空管和助航灯光、轻轨工程", "1"),
                        ("", "水运、地铁、桥梁、隧道工程", "1.1"),
                        ("", "索道工程", "1.3"),
                        ("6、建筑市政工程", "邮政工艺工程", "0.8"),
                        ("", "建筑、市政、电信工程", "1"),
                        ("", "人防、园林绿化、广电工艺工程", "1.1"),
                        ("7、农业林业工程", "农业工程", "0.9"),
                        ("", "林业工程", "0.8"),
                    ]
                    df_coef = pd.DataFrame(sheji_table_data, columns=["工程类型", "具体专业", "专业调整系数"])
                    st.table(df_coef)
                    st.info(
                        "工程设计费调整系数共三个：\n\n"
                        "1. **专业调整系数**（上表 附表二）\n"
                        "2. **工程复杂程度调整系数**：I级 0.85 / II级 1.0 / III级 1.15\n"
                        "3. **附加调整系数**：多个时合并计算（相加 − 个数 + 1）\n\n"
                        "计算公式：基本设计收费 = 收费基价（附表一） × 专业调整系数（附表二） × 复杂程度系数 × 附加调整系数\n\n"
                        "⚠️ 工程设计费**不含**高程调整系数（高程系数仅用于监理费 发改价格[2007]670号）"
                    )
                    st.divider()
                    st.caption("以下为 AI 补充说明（仅作解释性描述，数字以上表为准）：")

                is_kancha = (fee_result.get("fee_type") == "勘察费")

                if is_kancha:
                    # 勘察费（无金额时）：展示计算方法和费率说明
                    import pandas as pd
                    st.markdown("### 工程勘察费 — 计算方法")
                    st.info(
                        "**工程勘察费**依据《工程勘察设计收费管理规定》（计价格[2002]10号）"
                        "的 **工程勘察收费标准** 部分计算。\n\n"
                        "⚠️ **与工程设计费的重要区别**：工程勘察费按**实物工作量**定额计费，"
                        "不是按投资额比例。\n\n"
                        "**精确计算公式**：\n"
                        "- 工程勘察收费 = 工程勘察收费基准价 × (1 ± 20%)\n"
                        "- 工程勘察收费基准价 = 实物工作收费 + 技术工作收费\n"
                        "- 实物工作收费 = 收费基价 × 实物工作量 × 附加调整系数\n"
                        "- 技术工作收费 = 实物工作收费 × 技术工作收费比例\n\n"
                        "**粗略估算方法**（《市政工程设计概算编制办法》，中国计划出版社）：\n"
                        "- 通用项目：第一部分工程费 × **0.8%~1.1%**\n"
                        "- 建筑项目：第一部分工程费 × **0.3%~0.5%**\n\n"
                        "💡 提供建安费和设备费金额，程序可按上述百分比法粗略估算。"
                        "精确计算请提供勘察类型和实物工作量。"
                    )

                    jianan_detected = fee_result.get("检测到建安费(万元)")
                    shebei_detected = fee_result.get("检测到设备费(万元)")
                    amt_detected = fee_result.get("检测到金额(万元)")
                    if jianan_detected is not None:
                        st.metric("建安工程费", f"{jianan_detected} 万元")
                        if shebei_detected is not None:
                            st.metric("设备购置费", f"{shebei_detected} 万元")
                    elif amt_detected is not None:
                        st.metric("检测到金额", f"{amt_detected} 万元")

                    st.divider()
                    st.markdown("**需明确的参数（精确计算）**：")
                    st.markdown(
                        "1. 勘察类型（工程测量/岩土工程勘察/水文地质勘察/工程物探等 16 大类）\n"
                        "2. 实物工作量（钻孔深度、测量面积/比例尺、取样数量等）\n"
                        "3. 复杂程度等级（简单/中等/复杂）\n"
                        "4. 附加调整系数（气温/高程/带状/水域等）"
                    )
                    st.caption("详细费率表见知识库《计价格[2002]10号》工程勘察收费标准章节。")

                    # 构建响应文本
                    response = (
                        f"工程勘察费依据《工程勘察设计收费管理规定》（计价格[2002]10号）"
                        f"的工程勘察收费标准计算。\n\n"
                        f"⚠️ 与工程设计费不同，勘察费按**实物工作量**定额计费"
                        f"（如钻探米数、测量面积等），不是按投资额比例。\n\n"
                        f"**粗略估算**（《市政工程设计概算编制办法》）："
                        f"通用 0.8%~1.1%，建筑 0.3%~0.5%。\n"
                        f"**精确计算**需提供勘察类型（16大类）、实物工作量、复杂程度、附加调整系数。\n\n"
                        f"详细费率表见知识库《计价格[2002]10号》工程勘察收费标准章节。"
                    )

                else:
                    history = [
                        {"role": m["role"], "content": m["content"]}
                        for m in st.session_state.messages[:-1]
                        if m["role"] in ("user", "assistant")
                    ]
                    response = engine.chat(prompt, history)
                    st.markdown(response)

                with st.expander("查看计费依据"):
                    st.markdown(f"**{fee_result.get('费种', '')}**")
                    st.markdown(f"<small>依据：{_basis_with_links(fee_result.get('依据', ''))}</small>", unsafe_allow_html=True)
                    st.caption(f"计费方式：{fee_result.get('计费方式', '')}")
                    rate_table = fee_result.get("费率表", [])
                    if rate_table:
                        import pandas as pd
                        header = rate_table[0]
                        rows = rate_table[1:]
                        df = pd.DataFrame(rows, columns=header)
                        st.table(df)
                    auto_coefs = fee_result.get("auto_detected_coefs", {})
                    if auto_coefs:
                        st.markdown("**引擎自动检测的系数**：")
                        for k, v in auto_coefs.items():
                            st.markdown(f"- {k}：**{v}**")
                    steps = fee_result.get("计算步骤")
                    if steps:
                        st.markdown("**分档计算明细**：")
                        for s in steps:
                            st.markdown(
                                f"- {s.get('区间', '')}：{s.get('金额(万元)', '')}万元 "
                                f"× {s.get('费率(%)', '')}% = **{s.get('费用(万元)', '')}万元**"
                            )
                    adjustment = fee_result.get("计费额调整")
                    if adjustment and adjustment.get("触发调整"):
                        st.info(adjustment.get("说明", ""))
                    if "分摊" in fee_result:
                        st.caption(fee_result["分摊"])

                    results = engine.search(prompt, top_k=5)
                    if results:
                        st.divider()
                        st.markdown("**数据库匹配**：")
                        for r in results:
                            u = r.get("unit", "元/株")
                            st.markdown(
                                f"- [{r['category']}] {r['name']}（{r['spec']}）"
                                f" -> 综合指标 **{r['comprehensive']}**{u}"
                            )

            else:
                # === 正常 LLM 对话 ===
                history = [
                    {"role": m["role"], "content": m["content"]}
                    for m in st.session_state.messages[:-1]
                    if m["role"] in ("user", "assistant")
                ]
                response = engine.chat(prompt, history)
                st.markdown(response)

                with st.expander("查看检索到的数据"):
                    results = engine.search(prompt, top_k=5)
                    if results:
                        for r in results:
                            u = r.get("unit", "元/株")
                            st.markdown(
                                f"- [{r['category']}] {r['name']}（{r['spec']}）"
                                f" -> 综合指标 **{r['comprehensive']}**{u}"
                            )
                    else:
                        st.caption("未匹配到相关数据")

    st.session_state.messages.append({"role": "assistant", "content": response})
    st.rerun()